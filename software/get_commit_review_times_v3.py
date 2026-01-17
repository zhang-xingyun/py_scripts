#!/usr/bin/env python3
import logging
import os
import json
import datetime
import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from gitlab_app import models


def generate_excel(ids_data, filename, name):
    print(filename)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = name

    for index, value in enumerate(ids_data[0].keys()):
        worksheet.cell(1, (index + 1), str(value))
    for row, one_data in enumerate(ids_data):
        # print(row)
        # print(one_data)
        for colum, value in enumerate(one_data.values()):
            # tmp_value = ILLEGAL_CHARACTERS_RE.sub(r'', str(value))
            tmp_value = value
            if isinstance(value, str):
                tmp_value = ILLEGAL_CHARACTERS_RE.sub(r'', str(value))
            else:
                tmp_value = value
            # print(type(tmp_value),tmp_value)
            worksheet.cell((row + 2), (colum + 1), tmp_value)
    file_path = '/'.join(filename.split('/')[:-1])

    print(file_path)
    if not os.path.exists(file_path):
        os.makedirs(file_path)
    workbook.save((filename + '.xlsx'))
    print('finish')


def gen_commit_times_by_department(dep_level: int, dep_name: str):
    person_list = []
    if dep_level == 1:
        person_list = models.People.objects.filter(businessUnit=dep_name)
    if dep_level == 2:
        person_list = models.People.objects.filter(division=dep_name)
    if dep_level == 3:
        person_list = models.People.objects.filter(department=dep_name)
    if dep_level == 4:
        person_list = models.People.objects.filter(org4=dep_name)

    logging.debug(person_list)
    person_id_list = [person.id for person in person_list]
    commit_list = models.Commit.objects.filter(author_id__in=person_id_list, author_date__gt="2023-06-28")
    data = list()
    for commit in commit_list:
        total = commit.additions + commit.deletions
        if total > 1000 or commit.files > 50:
            continue
        tmp = dict()
        branches = [branch.branch_name for branch in
                    commit.branches.all()]
        tmp['commit_hash'] = commit.commit_hash
        tmp['author'] = commit.author.user_id
        tmp['additions'] = commit.additions
        tmp['deletions'] = commit.deletions
        tmp['total_code_line'] = commit.additions + commit.deletions
        tmp['files'] = commit.files
        tmp['author_date'] = str(commit.author_date)
        tmp['repo_id'] = commit.repo.repo_id
        tmp['cr_ph'] = commit.cr_ph
        tmp['branches'] = str(branches)
        data.append(tmp)
    print(data)
    generate_excel(data, 'report/commit_times-' + dep_name, 'commit_times')


def gen_commit_times_by_repo():
    repo_list = [8189]

    logging.debug(repo_list)
    commit_list = models.Commit.objects.filter(repo__repo_id__in=repo_list)
    data = list()
    for commit in commit_list:
        tmp = dict()
        tmp['commit_hash'] = commit.commit_hash
        tmp['author'] = commit.author.user_id
        tmp['additions'] = commit.additions
        tmp['deletions'] = commit.deletions
        tmp['total_code_line'] = commit.additions + commit.deletions
        tmp['files'] = commit.files
        tmp['author_date'] = str(commit.author_date)
        tmp['repo_id'] = commit.repo.repo_id
        tmp['cr_ph'] = commit.cr_ph
        data.append(tmp)
    print(data)
    generate_excel(data, 'report/commit_times_repo', 'commit_times')

def gen_commit_times_by_repo_branch():
    repo_branch_list = {
    }
    repo_list = repo_branch_list.keys()
    print(repo_list)
    commit_list = models.Commit.objects.filter(
        repo__repo_id__in=repo_branch_list.keys(),
        submitter_date__gt="2022-06-01")
    print(commit_list)
    data = list()
    for commit in commit_list:
        tmp = dict()
        exist_branches = list()
        branches = commit.branches.all()
        print(type(branches))
        print(branches)

        branch_name_list = [br.branch_name for br in branches]
        for branch in repo_branch_list[commit.repo.repo_id]:
            if branch in branch_name_list:
                exist_branches.append(branch)
        if not exist_branches:
            continue
        tmp["branch_list"] = str(exist_branches)
        if not commit.jira_ids:
            commit.jira_ids = "[]"
        tmp["jira_ids"] = str(commit.jira_ids)
        tmp['commit_hash'] = commit.commit_hash
        # tmp['author'] = commit.author.user_id
        # tmp['department'] = commit.author.org4
        tmp['author_date'] = str(commit.author_date)
        tmp['submitter_date'] = str(commit.submitter_date)
        tmp['repo_id'] = commit.repo.repo_id
        tmp['repo_url'] = commit.repo.web_url
        # tmp['cr_ph'] = commit.cr_ph
        data.append(tmp)
    print(data)
    generate_excel(data, 'report/commit_times', 'commit_times')


def gen_ph_cr_times_by_department(dep_level: int, dep_name: str):
    person_list = []
    if dep_level == 1:
        person_list = models.People.objects.filter(businessUnit=dep_name)
    if dep_level == 2:
        person_list = models.People.objects.filter(division=dep_name)
    if dep_level == 3:
        person_list = models.People.objects.filter(department=dep_name)
    if dep_level == 4:
        person_list = models.People.objects.filter(org4=dep_name)
    logging.debug(person_list)
    person_name_list = [person.user_id for person in person_list]
    cr_list = models.CodeReview.objects.filter(reviewer__in=person_name_list,
                                               source='PH', accept_time__gt="2023-06-28")
    data = list()
    for cr in cr_list:

        comments = list()
        if cr.comments:
            print(cr.comments)
            try:
                comm_json = json.loads(cr.comments)
            except Exception as e:
                print("json loads error:" + str(e))
                comm_json = list()
            for comm in comm_json:
                comm_tmp = dict()
                comm_tmp['time'] = datetime.datetime.utcfromtimestamp(
                    comm['time'])
                comm_tmp['comment'] = comm['comment']
                comments.append(comm_tmp)
        if len(comments) == 0:
            tmp = dict()
            tmp['reviewer'] = cr.reviewer
            tmp['ph_id'] = cr.source_id
            tmp['start_time'] = str(cr.start_time)
            tmp['accept_time'] = str(cr.accept_time)
            tmp['comments_mum'] = cr.comments_mum
            tmp['duration'] = cr.duration
            tmp['repo_id'] = cr.repo_id
            tmp['comments_date'] = ""
            tmp['comments'] = ""
            data.append(tmp)
        else:
            for comment in comments:
                tmp = dict()
                tmp['reviewer'] = cr.reviewer
                tmp['ph_id'] = cr.source_id
                tmp['start_time'] = str(cr.start_time)
                tmp['accept_time'] = str(cr.accept_time)
                tmp['comments_mum'] = cr.comments_mum
                tmp['duration'] = cr.duration
                tmp['repo_id'] = cr.repo_id
                tmp['comments_date'] = str(comment['time'])
                tmp['comments'] = str(comment['comment'])
                data.append(tmp)
    print(data)
    generate_excel(data, 'report/ph_cr_times-' + dep_name, 'ph_cr_times')


def gen_ph_cr_times_by_repo():
    repo_list = [7429]
    logging.debug(repo_list)
    cr_list = models.CodeReview.objects.filter(repo_id__in=repo_list)
    data = list()
    for cr in cr_list:

        comments = list()
        if cr.comments:
            print(cr.comments)
            try:
                comm_json = json.loads(cr.comments)
            except Exception as e:
                print("json loads error:" + str(e))
                comm_json = list()
            for comm in comm_json:
                comm_tmp = dict()
                comm_tmp['time'] = datetime.datetime.utcfromtimestamp(
                    comm['time'])
                comm_tmp['comment'] = comm['comment']
                comments.append(comm_tmp)
        if len(comments) == 0:
            tmp = dict()
            tmp['reviewer'] = cr.reviewer
            tmp['ph_id'] = cr.ph_id
            tmp['start_time'] = str(cr.start_time)
            tmp['accept_time'] = str(cr.accept_time)
            tmp['comments_mum'] = cr.comments_mum
            tmp['duration'] = cr.duration
            tmp['repo_id'] = cr.repo_id
            tmp['comments_date'] = ""
            tmp['comments'] = ""
            data.append(tmp)
        else:
            for comment in comments:
                tmp = dict()
                tmp['reviewer'] = cr.reviewer
                tmp['ph_id'] = cr.ph_id
                tmp['start_time'] = str(cr.start_time)
                tmp['accept_time'] = str(cr.accept_time)
                tmp['comments_mum'] = cr.comments_mum
                tmp['duration'] = cr.duration
                tmp['repo_id'] = cr.repo_id
                tmp['comments_date'] = str(comment['time'])
                tmp['comments'] = str(comment['comment'])
                data.append(tmp)
    print(data)
    generate_excel(data, 'report/ph_cr_times-repo', 'ph_cr_times')

def gen_mr_cr_times_by_department(dep_level: int, dep_name: str):
    person_list = []
    if dep_level == 1:
        person_list = models.People.objects.filter(businessUnit=dep_name)
    if dep_level == 2:
        person_list = models.People.objects.filter(division=dep_name)
    if dep_level == 3:
        person_list = models.People.objects.filter(department=dep_name)
    if dep_level == 4:
        person_list = models.People.objects.filter(org4=dep_name)
    logging.debug(person_list)
    person_name_list = [person.user_id for person in person_list]
    cr_list = models.CodeReview.objects.filter(reviewer__in=person_name_list,
                                               source='MR', accept_time__gt="2023-06-28")
    data = list()
    for cr in cr_list:

        comments = list()
        if cr.comments:
            print(cr.comments)
            try:
                comm_json = json.loads(cr.comments)
            except Exception as e:
                print("json loads error:" + str(e))
                comm_json = list()
            for comm in comm_json:
                comm_tmp = dict()
                comm_tmp['time'] = datetime.datetime.utcfromtimestamp(
                    comm['time'])
                comm_tmp['comment'] = comm['comment']
                comments.append(comm_tmp)
        if len(comments) == 0:
            tmp = dict()
            tmp['reviewer'] = cr.reviewer
            tmp['ph_id'] = cr.source_id
            tmp['start_time'] = str(cr.start_time)
            tmp['accept_time'] = str(cr.accept_time)
            tmp['comments_mum'] = cr.comments_mum
            tmp['duration'] = cr.duration
            tmp['repo_id'] = cr.repo_id
            tmp['comments_date'] = ""
            tmp['comments'] = ""
            data.append(tmp)
        else:
            for comment in comments:
                tmp = dict()
                tmp['reviewer'] = cr.reviewer
                tmp['ph_id'] = cr.source_id
                tmp['start_time'] = str(cr.start_time)
                tmp['accept_time'] = str(cr.accept_time)
                tmp['comments_mum'] = cr.comments_mum
                tmp['duration'] = cr.duration
                tmp['repo_id'] = cr.repo_id
                tmp['comments_date'] = str(comment['time'])
                tmp['comments'] = str(comment['comment'])
                data.append(tmp)
    print(data)
    generate_excel(data, 'report/mr_cr_times-' + dep_name, 'mr_cr_times')


def gen_mr_cr_times_by_gitlab_path(gitlab_path: str):
    repos = models.Repo.objects.filter(web_url__icontains=gitlab_path)
    repo_ids = [item.repo_id for item in repos]
    cr_list = models.CodeReview.objects.filter(repo_id__in=repo_ids, source='MR')
    data = list()
    for cr in cr_list:

        comments = list()
        if cr.comments:
            print(cr.comments)
            try:
                comm_json = json.loads(cr.comments)
            except Exception as e:
                print("json loads error:" + str(e))
                comm_json = list()
            for comm in comm_json:
                comm_tmp = dict()
                comm_tmp['time'] = datetime.datetime.utcfromtimestamp(
                    comm['time'])
                comm_tmp['comment'] = comm['comment']
                comments.append(comm_tmp)
        if len(comments) == 0:
            tmp = dict()
            tmp['reviewer'] = cr.reviewer
            tmp['ph_id'] = cr.source_id
            tmp['start_time'] = str(cr.start_time)
            tmp['accept_time'] = str(cr.accept_time)
            tmp['comments_mum'] = cr.comments_mum
            tmp['duration'] = cr.duration
            tmp['repo_id'] = cr.repo_id
            tmp['comments_date'] = ""
            tmp['comments'] = ""
            data.append(tmp)
        else:
            for comment in comments:
                tmp = dict()
                tmp['reviewer'] = cr.reviewer
                tmp['ph_id'] = cr.source_id
                tmp['start_time'] = str(cr.start_time)
                tmp['accept_time'] = str(cr.accept_time)
                tmp['comments_mum'] = cr.comments_mum
                tmp['duration'] = cr.duration
                tmp['repo_id'] = cr.repo_id
                tmp['comments_date'] = str(comment['time'])
                tmp['comments'] = str(comment['comment'])
                data.append(tmp)
    print(data)
    generate_excel(data, 'report/mr_cr_times-' + '土星', 'mr_cr_times')


def gen_ph_cr_times_by_gitlab_path(gitlab_path: str):
    repos = models.Repo.objects.filter(web_url__icontains=gitlab_path)
    repo_ids = [item.repo_id for item in repos]
    cr_list = models.CodeReview.objects.filter(repo_id__in=repo_ids, source='PH')
    data = list()
    for cr in cr_list:

        comments = list()
        if cr.comments:
            print(cr.comments)
            try:
                comm_json = json.loads(cr.comments)
            except Exception as e:
                print("json loads error:" + str(e))
                comm_json = list()
            for comm in comm_json:
                comm_tmp = dict()
                comm_tmp['time'] = datetime.datetime.utcfromtimestamp(
                    comm['time'])
                comm_tmp['comment'] = comm['comment']
                comments.append(comm_tmp)
        if len(comments) == 0:
            tmp = dict()
            tmp['reviewer'] = cr.reviewer
            tmp['ph_id'] = cr.source_id
            tmp['start_time'] = str(cr.start_time)
            tmp['accept_time'] = str(cr.accept_time)
            tmp['comments_mum'] = cr.comments_mum
            tmp['duration'] = cr.duration
            tmp['repo_id'] = cr.repo_id
            tmp['comments_date'] = ""
            tmp['comments'] = ""
            data.append(tmp)
        else:
            for comment in comments:
                tmp = dict()
                tmp['reviewer'] = cr.reviewer
                tmp['ph_id'] = cr.source_id
                tmp['start_time'] = str(cr.start_time)
                tmp['accept_time'] = str(cr.accept_time)
                tmp['comments_mum'] = cr.comments_mum
                tmp['duration'] = cr.duration
                tmp['repo_id'] = cr.repo_id
                tmp['comments_date'] = str(comment['time'])
                tmp['comments'] = str(comment['comment'])
                data.append(tmp)
    print(data)
    generate_excel(data, 'report/ph_cr_times-' + '土星', 'ph_cr_times')


def gen_commit_times_by_gitlab_path(gitlab_path: str):
    repos = models.Repo.objects.filter(web_url__icontains=gitlab_path)
    repo_ids = [item.repo_id for item in repos]
    commit_list = models.Commit.objects.filter(repo__repo_id__in=repo_ids)
    data = list()
    for commit in commit_list:
        total = commit.additions + commit.deletions
        # if total > 1000 or commit.files > 50:
        #     continue
        tmp = dict()
        user_id = ''
        if commit.author:
            user_id = commit.author.user_id
        branches = [branch.branch_name for branch in
                    commit.branches.all()]
        tmp['commit_hash'] = commit.commit_hash
        tmp['author'] = user_id
        tmp['additions'] = commit.additions
        tmp['deletions'] = commit.deletions
        tmp['total_code_line'] = commit.additions + commit.deletions
        tmp['files'] = commit.files
        tmp['author_date'] = str(commit.author_date)
        tmp['repo_id'] = commit.repo.repo_id
        tmp['cr_ph'] = commit.cr_ph
        tmp['branches'] = str(branches)
        data.append(tmp)
    print(data)
    generate_excel(data, 'report/commit_times-' + '土星', 'commit_times')


def gen_files_times_by_gitlab_path(gitlab_path: str):
    repos = models.Repo.objects.filter(web_url__icontains=gitlab_path)
    repo_ids = [item.repo_id for item in repos]
    commit_list = models.Commit.objects.filter(repo__repo_id__in=repo_ids)
    data = list()
    for commit in commit_list:
        file_list = json.loads(commit.file_list)
        for file in file_list:
            tmp = dict()
            # branches = [branch.branch_name for branch in
            #             commit.branches.all()]
            user_id = ''
            if commit.author:
                user_id = commit.author.user_id
            path_level = file['file'].split('/')
            first_level_path = '/'
            if len(path_level) > 1:
                first_level_path = path_level[0]
            # tmp['author'] = user_id
            tmp['first_level_path'] = first_level_path
            tmp['file'] = file['file']
            tmp['additions'] = file['addtions']
            tmp['deletions'] = file['deletions']
            tmp['commit_hash'] = commit.commit_hash
            tmp['author_date'] = str(commit.author_date)
            tmp['repo_id'] = commit.repo.repo_id
            # tmp['cr_ph'] = commit.cr_ph
            # tmp['branches'] = str(branches)
            data.append(tmp)
    print(data)
    generate_excel(data, 'report/commit_files_times-' + '土星', 'commit_times')


def run():
    # gen_commit_times_by_department(1, "软件平台")
    # gen_cr_times_by_department(1, "软件平台")
    # gen_ph_cr_times_by_department(1, "软件平台")
    # gen_mr_cr_times_by_department(1, "软件平台")
    # gen_commit_times_by_repo_branch()
    # gen_commit_times_by_repo()
    # gen_ph_cr_times_by_repo()
    gen_mr_cr_times_by_gitlab_path("https://gitlab.test.com/proj/saturn")
    gen_ph_cr_times_by_gitlab_path("https://gitlab.test.com/proj/saturn")
    gen_commit_times_by_gitlab_path("https://gitlab.test.com/proj/saturn")
    gen_files_times_by_gitlab_path("https://gitlab.test.com/proj/saturn")
