#!/usr/bin/env python3.6
# -*- coding: utf-8 -*-
# coding=utf-8
import os
import sys
import re
import logging
import time
import yaml
import requests
import pprint
import json
import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import datetime
from gitlab_app import models

proj_map = {
}


class JenkinsData:
    def __init__(self):
        self.file = ""

    def read_data(self):
        s_date = datetime.datetime.strptime('2022-07-25 18:0:0',
                                            '%Y-%m-%d %H:%M:%S')
        e_date = datetime.datetime.strptime('2022-07-26 18:0:0',
                                            '%Y-%m-%d %H:%M:%S')
        jobs = models.JenkinsBuildRecord.objects.filter(
            job_start_time__gte=s_date, job_start_time__lte=e_date)
        output = dict()
        if jobs:
            for job in jobs:
                job_name = job.job_name
                # if not job_name == 'platform_ap_toolchain/HAT':
                #     continue
                job_url = job.job_build_url
                job_result = job.job_result
                job_start_time = job.job_start_time
                job_buildableTimeMillis = \
                    job.job_buildableTimeMillis / 1000 / 60
                job_buildingDurationMillis = \
                    job.job_buildingDurationMillis / 1000 / 60
                job_executingTimeMillis = \
                    job.job_executingTimeMillis / 1000 / 60
                # job_waitingTimeMillis = job.job_waitingTimeMillis / 1000 / 60

                print(job_start_time)
                if job_name not in output.keys():
                    output[job_name] = dict()
                    output[job_name]['项目'] = 'unknow'
                    for key in proj_map.keys():
                        if re.search(key, job_name):
                            output[job_name]['项目'] = proj_map[key]
                            break
                    output[job_name]['名字'] = job_name
                    output[job_name]['链接'] = job_url
                    output[job_name]['次数'] = int()
                    output[job_name]['成功'] = int()
                    output[job_name]['失败'] = int()
                    output[job_name]['取消'] = int()
                    output[job_name]['总时长'] = int()
                    output[job_name]['平均总时长'] = int()
                    output[job_name]['总等待时间'] = int()
                    output[job_name]['平均等待时间'] = int()
                    output[job_name]['总执行时间'] = int()
                    output[job_name]['平均执行时间'] = int()
                    output[job_name]['备注'] = str()
                    # output[job_name]['waitingTimeMillis'] = int()
                    # output[job_name]['average_waitingTimeMillis'] = int()
                output[job_name]['次数'] = output[job_name]['次数'] + 1
                if job_result == 'SUCCESS':
                    output[job_name]['成功'] = output[job_name]['成功'] + 1
                if job_result == 'FAILURE':
                    output[job_name]['失败'] = output[job_name]['失败'] + 1
                if job_result == 'ABORTED':
                    output[job_name]['取消'] = output[job_name]['取消'] + 1
                output[job_name]['总时长'] = \
                    round(output[job_name]['总时长'] +
                          job_buildingDurationMillis, 2)
                output[job_name]['平均总时长'] = \
                    round(output[job_name]['总时长'] / output[
                        job_name]['次数'], 2)
                output[job_name]['总等待时间'] = \
                    round(output[job_name]['总等待时间'] +
                          job_buildableTimeMillis, 2)
                output[job_name]['平均等待时间'] = \
                    round(output[job_name]['总等待时间'] / output[
                        job_name]['次数'], 2)
                output[job_name]['总执行时间'] = round(
                    output[job_name][
                        '总执行时间'] + job_executingTimeMillis, 2)
                output[job_name]['平均执行时间'] = round(
                    output[job_name][
                        '总执行时间'] / output[job_name]['次数'], 2)
                # output[job_name]['waitingTimeMillis'] = round(
                #     output[job_name][
                #         'waitingTimeMillis'] + job_waitingTimeMillis, 2)
                # output[job_name]['average_waitingTimeMillis'] = round(
                #     output[job_name][
                #         'waitingTimeMillis'] / output[job_name]['次数'], 2)

        test_2 = sorted(output.items(), key=lambda x: x[1]['次数'], reverse=True)
        test_4 = [item[1] for item in test_2]
        print(test_4)
        target_name = 'jenkins_data'
        self.generate_excel(test_4, 'report/' + target_name, target_name)

    def generate_excel(self, ids_data, filename, name):
        print(filename)
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = name

        for index, value in enumerate(ids_data[0].keys()):
            worksheet.cell(1, (index + 1), str(value))
        for row, one_data in enumerate(ids_data):
            # print(row)
            # print(one_data)
            for colum, value in enumerate(one_data.values()):
                # tmp_value = ILLEGAL_CHARACTERS_RE.sub(r'', str(value))
                tmp_value = value
                if isinstance(value, str):
                    tmp_value = ILLEGAL_CHARACTERS_RE.sub(r'', str(value))
                else:
                    tmp_value = value
                # print(type(tmp_value),tmp_value)
                worksheet.cell((row + 2), (colum + 1), tmp_value)
        file_path = '/'.join(filename.split('/')[:-1])
        print(file_path)
        if not os.path.exists(file_path):
            os.makedirs(file_path)
        workbook.save((filename + '.xlsx'))
        print('finish')


def run():
    jd = JenkinsData()
    jd.read_data()
