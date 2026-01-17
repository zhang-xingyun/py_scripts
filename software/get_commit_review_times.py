#!/usr/bin/env python3
import logging
import os
import json
import datetime
import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from gitlab_app import models


def generate_excel(ids_data, filename, name):
    print(filename)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = name

    for index, value in enumerate(ids_data[0].keys()):
        worksheet.cell(1, (index + 1), str(value))
    for row, one_data in enumerate(ids_data):
        # print(row)
        # print(one_data)
        for colum, value in enumerate(one_data.values()):
            # tmp_value = ILLEGAL_CHARACTERS_RE.sub(r'', str(value))
            tmp_value = value
            if isinstance(value, str):
                tmp_value = ILLEGAL_CHARACTERS_RE.sub(r'', str(value))
            else:
                tmp_value = value
            # print(type(tmp_value),tmp_value)
            worksheet.cell((row + 2), (colum + 1), tmp_value)
    file_path = '/'.join(filename.split('/')[:-1])

    print(file_path)
    if not os.path.exists(file_path):
        os.makedirs(file_path)
    workbook.save((filename + '.xlsx'))
    print('finish')


def gen_commit_times_by_department(dep_level: int, dep_name: str):
    person_list = []
    if dep_level == 1:
        person_list = models.People.objects.filter(businessUnit=dep_name)
    if dep_level == 2:
        person_list = models.People.objects.filter(division=dep_name)
    if dep_level == 3:
        person_list = models.People.objects.filter(department=dep_name)
    if dep_level == 4:
        person_list = models.People.objects.filter(org4=dep_name)

    logging.debug(person_list)
    person_id_list = [person.id for person in person_list]
    commit_list = models.Commit.objects.filter(author_id__in=person_id_list)
    data = list()
    for commit in commit_list:
        tmp = dict()
        tmp['commit_hash'] = commit.commit_hash
        tmp['author'] = commit.author.user_id
        tmp['additions'] = commit.additions
        tmp['deletions'] = commit.deletions
        tmp['total_code_line'] = commit.additions + commit.deletions
        tmp['files'] = commit.files
        tmp['author_date'] = str(commit.author_date)
        tmp['repo_id'] = commit.repo_id
        tmp['cr_ph'] = commit.cr_ph
        data.append(tmp)
    print(data)
    generate_excel(data, 'report/commit_times-' + dep_name, 'commit_times')


def gen_commit_times_by_repo():
    repo_list = [7572, 7624, 7573, 7620]
    commit_list = models.Commit.objects.filter(repo__repo_id__in=repo_list,
                                               submitter_date__gt="2023-02-17")
    data = list()
    for commit in commit_list:
        tmp = dict()
        tmp['commit_hash'] = commit.commit_hash
        # tmp['author'] = commit.author.user_id
        # tmp['department'] = commit.author.org4
        # tmp['author_date'] = str(commit.author_date)
        tmp['submitter_date'] = str(commit.submitter_date)
        # tmp['repo_id'] = commit.repo.repo_id
        tmp['repo'] = commit.repo.name
        # tmp['cr_ph'] = commit.cr_ph
        data.append(tmp)
    print(data)
    generate_excel(data, 'report/commit_times', 'commit_times')


def gen_cr_times_by_department(dep_level: int, dep_name: str):
    person_list = []
    if dep_level == 1:
        person_list = models.People.objects.filter(businessUnit=dep_name)
    if dep_level == 2:
        person_list = models.People.objects.filter(division=dep_name)
    if dep_level == 3:
        person_list = models.People.objects.filter(department=dep_name)
    if dep_level == 4:
        person_list = models.People.objects.filter(org4=dep_name)
    logging.debug(person_list)
    person_name_list = [person.user_id for person in person_list]
    cr_list = models.CodeReview.objects.filter(reviewer__in=person_name_list)
    data = list()
    for cr in cr_list:

        comments = list()
        if cr.comments:
            print(cr.comments)
            try:
                comm_json = json.loads(cr.comments)
            except Exception as e:
                print("json loads error:" + str(e))
                comm_json = list()
            for comm in comm_json:
                comm_tmp = dict()
                comm_tmp['time'] = datetime.datetime.utcfromtimestamp(
                    comm['time'])
                comm_tmp['comment'] = comm['comment']
                comments.append(comm_tmp)
        if len(comments) == 0:
            tmp = dict()
            tmp['reviewer'] = cr.reviewer
            tmp['ph_id'] = cr.ph_id
            tmp['start_time'] = str(cr.start_time)
            tmp['accept_time'] = str(cr.accept_time)
            tmp['comments_mum'] = cr.comments_mum
            tmp['duration'] = cr.duration
            tmp['repo_id'] = cr.repo_id
            tmp['comments_date'] = ""
            tmp['comments'] = ""
            data.append(tmp)
        else:
            for comment in comments:
                tmp = dict()
                tmp['reviewer'] = cr.reviewer
                tmp['ph_id'] = cr.ph_id
                tmp['start_time'] = str(cr.start_time)
                tmp['accept_time'] = str(cr.accept_time)
                tmp['comments_mum'] = cr.comments_mum
                tmp['duration'] = cr.duration
                tmp['repo_id'] = cr.repo_id
                tmp['comments_date'] = str(comment['time'])
                tmp['comments'] = str(comment['comment'])
                data.append(tmp)
    print(data)
    generate_excel(data, 'report/cr_times-' + dep_name, 'cr_times')


def run():
    gen_commit_times_by_department(1, "软件平台")
    gen_cr_times_by_department(1, "软件平台")
    # gen_commit_times_by_repo()
