import pickle
import os
import time
import sys
import shutil
import argparse
import datetime
import textwrap
import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import gitlab
from gitlab_app import models


class Report():
    def __init__(self):
        self.raw_data_path = '/opt/gitlab_report/raw_data/'
        self.raw_data_path2 = '/opt/gitlab_report/raw_data2/'
        self.branch_rule = '^master$|^develop$|^main$|^release-.+|' \
                           '^sprint-.+|^feature-.+|^bugfix-.+|^hotfix-.+|' \
                           '^cicd-.+|^test-.+|^tool-.+|^dev-.+|^rel-.+|' \
                           '^feat-.+'
        self.gl = gitlab.Gitlab.from_config('trigger', ['/data/wwwroot/gitlab_data/pyecharts_django/python-gitlab.cfg'])
        self.gl.auth()

        self.current_time = int(time.time())

    def read_pkl(self, file_name):
        f_obj = open(file_name, 'rb')
        return pickle.load(f_obj)

    def generate_excel(self, ids_data, filename, name):
        print(filename)
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = name

        for index, value in enumerate(ids_data[0].keys()):
            worksheet.cell(1, (index + 1), str(value))
        for row, one_data in enumerate(ids_data):
            # print(row)
            # print(one_data)
            for colum, value in enumerate(one_data.values()):
                # tmp_value = ILLEGAL_CHARACTERS_RE.sub(r'', str(value))
                tmp_value = value
                if isinstance(value, str):
                    tmp_value = ILLEGAL_CHARACTERS_RE.sub(r'', str(value))
                else:
                    tmp_value = value
                # print(type(tmp_value),tmp_value)
                worksheet.cell((row + 2), (colum + 1), tmp_value)
        file_path = '/'.join(filename.split('/')[:-1])
        print(file_path)
        if not os.path.exists(file_path):
            os.makedirs(file_path)
        workbook.save((filename + '.xlsx'))
        print('finish')

    def get_inactive_branch(self, project_id, project_url):
        output = list()
        branches_attributes_file_path = \
            self.raw_data_path + '/repo/' + \
            str(project_id) + '/branches_attributes.pkl'
        if not os.path.exists(branches_attributes_file_path):
            return
        branches_attributes = self.read_pkl(branches_attributes_file_path)

        events = list()
        events_file_path = self.raw_data_path + '/repo/' + str(
            project_id) + '/events_attributes.pkl'
        if os.path.exists(events_file_path):
            events = self.read_pkl(events_file_path)
        #
        for branch_name, branch in branches_attributes.items():
            freeze = False
            # if not branch['name'] == 'release-hongjing':
            #     continue
            print(branch['name'])
            last_commit_date = ''
            if branch['commit']:
                last_commit_date = branch['commit']['created_at'].split('T')[0]
            print('last commit date:' + last_commit_date)
            if branch['protected']:
                continue
            branch_create = 'more_unknown'
            author_username = 'more_unknown'
            for event in events:
                if event['action_name'] == 'pushed new' and \
                        event['push_data']['ref'] == branch_name:
                    branch_create = event['created_at'].split('T')[0]
                    author_username = event['author_username']
            tmp_dict = dict()
            tmp_dict['repo_id'] = project_id
            tmp_dict['branch_name'] = branch_name
            tmp_dict['freeze'] = freeze
            tmp_dict['type'] = 'frozen branch'
            tmp_dict['web_url'] = project_url
            tmp_dict['author_username'] = author_username
            tmp_dict['branch_create'] = branch_create
            tmp_dict['branch_last_commit'] = last_commit_date

            tmp_dict['inactive_day'] = 0
            if tmp_dict['branch_last_commit']:
                tmp_dict['inactive_day'] \
                    = int((self.current_time - time.mktime(time.strptime(
                        tmp_dict['branch_last_commit'], '%Y-%m-%d'))) / (
                        24 * 60 * 60))

            block_branch_ret = False
            if branch_create == 'more_unknown':
                block_branch_ret = True
            else:
                block_branch_create_day = \
                    int((self.current_time - time.mktime(time.strptime(
                        tmp_dict['branch_create'], '%Y-%m-%d'))) / (
                        24 * 60 * 60))
                if block_branch_create_day > 186:
                    block_branch_ret = True
            if tmp_dict['inactive_day'] > 186 and block_branch_ret:
                output.append(tmp_dict)
        # if len(output) > 0:
        #     self.generate_excel(output, 'report/branch', 'branch')
        return output

    def freeze_repo_branch(self):
        need_be_frozen_repo_list = list()
        need_be_frozen_branch_list = list()
        current_time = int(time.time())
        projects = self.gl.projects.list(
            all=True,
            retry_transient_errors=True)
        # projects = list()
        # project = self.gl.projects.get(2096)
        # projects.append(project)
        for project in projects:
            if project.attributes['path_with_namespace'].startswith(
                    "third_party") \
                    or project.attributes['path_with_namespace'].startswith(
                    "3rd") \
                    or project.attributes['path_with_namespace'].startswith(
                    "tmg") \
                    or project.attributes['path_with_namespace'].startswith(
                    "sdp_public"):
                continue
            if not project.attributes['archived']:
                inactive_day = int(
                    (current_time - time.mktime(
                        time.strptime(
                            project.attributes['last_activity_at'].split('T')[
                                0], '%Y-%m-%d'))) / (24 * 60 * 60))
                print('repo inactive days:' + str(inactive_day))
                print('last active day:' + str(
                    project.attributes['last_activity_at']))
                if inactive_day > 186:
                    temp_repo = dict()
                    temp_repo['repo_id'] = project.attributes['id']
                    temp_repo['web_url'] = project.attributes['web_url']
                    temp_repo['inactive_day'] = inactive_day
                    temp_repo['type'] = 'frozen repo'
                    temp_repo['branch_last_commit'] = '--'
                    temp_repo['branch_create'] = '--'
                    temp_repo['branch_name'] = '--'
                    need_be_frozen_repo_list.append(temp_repo)
                    continue

                # 判断分支的活跃度
                inactive_branch = self.get_inactive_branch(
                    project.attributes['id'], project.attributes['web_url'])
                if inactive_branch:
                    need_be_frozen_branch_list.extend(inactive_branch)

        if need_be_frozen_repo_list:
            self.generate_excel(need_be_frozen_repo_list, 'report/freeze_repo',
                                'freeze_repo')
            self.freeze_project_list(need_be_frozen_repo_list)
        if need_be_frozen_branch_list:
            self.generate_excel(need_be_frozen_branch_list,
                                'report/freeze_branch', 'freeze_branch')
            self.protect_branch_list(need_be_frozen_branch_list)

    def freeze_project_list(self, repo_dic_list):
        failed = list()
        success = list()
        for repo_dic in repo_dic_list:
            id = repo_dic['repo_id']
            gitlab_project = None
            try:
                gitlab_project = self.gl.projects.get(id)
                print('project is:' + str(gitlab_project))
            except BaseException:
                print('cannot find the gitlab_project with id:' + str(id))
                failed.append(id)
            if gitlab_project is not None:
                gitlab_project.archive()
                gitlab_project.save()
                self.record_frozen_actions(repo_dic)
                print('id: ' + str(id) + ' was frozen successfully!')
                success.append(id)
        if len(failed):
            print("The following ids failed:" + str(failed))

    def protect_branch_list(self, branch_dic_list):
        for branch_dic in branch_dic_list:
            id = branch_dic['repo_id']
            branch_name = branch_dic['branch_name']
            gitlab_project = None
            failed = list()
            success = list()
            try:
                gitlab_project = self.gl.projects.get(
                    id,
                    retry_transient_errors=True
                )
                # print('project is:' + str(gitlab_project))
            except BaseException:
                print('cannot find the gitlab_project with id:' + str(id))
                failed.append(id)
            if gitlab_project is not None:
                try:
                    # branch = gitlab_project.branches.get(
                    #     branch_name,
                    #     retry_transient_errors=True
                    # )
                    gitlab_project.protectedbranches.create({
                        'name': branch_name,
                        'merge_access_level': 0,
                        'push_access_level': 0},
                        retry_transient_errors=True)
                    gitlab_project.save()
                    print('id: ' + str(id) + ' branch: ' + str(
                        branch_name) + ' was protected successfully!')
                    success.append(id)
                    self.record_frozen_actions(branch_dic)
                except Exception as e:
                    print('id: ' + str(id) + 'branch: ' + str(
                        branch_name) + 'failed with exception:' + str(e))
            if len(failed):
                print("The following ids failed:" + str(failed))

    def record_frozen_actions(self, element):
        try:
            now_time = datetime.datetime.now()
            models.FrozenRepoBranch.objects.create(
                repo_id=element['repo_id'],
                web_url=element['web_url'],
                branch_name=element['branch_name'],
                branch_create_date=element['branch_create'],
                branch_last_commit_date=element['branch_last_commit'],
                inactive_day=element['inactive_day'],
                type=element['type'],
                frozen_date=now_time,
            )
            print("record_frozen_actions successful: ")
        except Exception as e:
            print('record_frozen_actions error: ', str(e))


def main():
    report = Report()
    report.freeze_repo_branch()
    # report.get_inactive_branch(7484)


def run():
    main()
