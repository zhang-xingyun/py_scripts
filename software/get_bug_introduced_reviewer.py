# -*-coding:utf-8一*-

import json
import logging
import os
import re
import shutil
import requests
import gitlab
import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from django.db.models import Q
from gitlab_app import models

result_path = os.path.join('result', 'diff.txt')
middle_file = os.path.join('middle', 'middle.txt')

logging.basicConfig(
    format='%(levelname)s:%(asctime)s : %(lineno)d :%(message)s',
    datefmt='%Y/%m/%d %H:%M:%S', level=logging.INFO)


def parse_json_file(file):
    if not os.path.exists(file):
        logging.error("{}文件不存在！".format(file))
        exit(1)
    file = open(file, "r", encoding="utf8")
    data_json = json.load(file)
    file.close()
    return data_json


def generate_excel(ids_data, filename, name):
    print(filename)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = name

    for index, value in enumerate(ids_data[0].keys()):
        worksheet.cell(1, (index + 1), str(value))
    for row, one_data in enumerate(ids_data):
        # print(row)
        # print(one_data)
        for colum, value in enumerate(one_data.values()):
            # tmp_value = ILLEGAL_CHARACTERS_RE.sub(r'', str(value))
            tmp_value = value
            if isinstance(value, str):
                tmp_value = ILLEGAL_CHARACTERS_RE.sub(r'', str(value))
            else:
                tmp_value = value
            # print(type(tmp_value),tmp_value)
            worksheet.cell((row + 2), (colum + 1), tmp_value)
    file_path = '/'.join(filename.split('/')[:-1])

    print(file_path)
    if not os.path.exists(file_path):
        os.makedirs(file_path)
    workbook.save((filename + '.xlsx'))
    print('finish')


class Cov():
    def __init__(self):
        self.gl = gitlab.Gitlab.from_config('trigger', ['/data/wwwroot/gitlab_data/pyecharts_django/python-gitlab.cfg'])
        self.ph_api_token = ''
        self.ph_api_url = 'https://cr.test.com/'
        self.result = list()
        self.gl.auth()

    def update_commit_message(self):
        cbs = models.CodeReviewBugIntroduce.objects.all()
        for cb in cbs:
            # if cb.id < 44737:
            #     continue
            if cb.ph_id:
                if not cb.commit_msg:
                    ph_info = self.get_ph_info(cb.ph_id)
                    title = ph_info.get('title', '')
                    if title:
                        cb.commit_msg = title
                        cb.save()


    def gitlab_get(self, repo_id, commit_id, file_path):
        # path = f"{mr_ins.manager.path}/{mr_ins.get_id()}/commits"
        # filepath = j5%2Fhb_api_isp.c
        # ref=228f7530fa2474d38c58153e75a5df9a1195a0cf
        filepath = file_path.replace("/", "%2F")
        path = "https://gitlab.test.com/api/v4/projects/%s/repository/files" \
               "/%s/blame?ref=%s" % (
                   repo_id, filepath, commit_id)
        try:
            result = self.gl.http_request("get", path)
            result_json = result.json()
            return result_json
        except Exception as e:
            logging.error("gitlab get failed:" + str(e))
            return ''

    def get_file_name_and_line(self):
        """获取改动文件路径和改动代码行"""
        data_dict = dict()
        pattern_1 = re.compile('\\+\\+\\+\\ (b/)?.*')
        pattern_2 = re.compile('---\\s+a/(.*)')
        pattern_3 = re.compile(
            '@@\\ -([0-9]+)(,[0-9]+)?\\ \\+([0-9]+)(,[0-9]+)?\\ @@.*')
        pattern_4 = re.compile('^($esc\\[[0-9;]*m)*([\\ +-])')
        logging.debug('start get file line')
        try:
            file_name = ''
            line = 0
            with open(middle_file, 'rb') as f2:
                while True:
                    code = f2.readline().decode('utf-8')
                    if code == '':
                        break
                    if pattern_1.search(code):
                        logging.debug('++++find')
                        continue
                    # 获取文件路径
                    elif pattern_2.search(code):
                        logging.debug('---find')
                        file_name = pattern_2.search(code).group(1)
                        logging.debug(file_name)
                        data_dict[file_name] = list()
                    # 获取变动代码行数  @@ -3046,6 +3049,135 @@
                    elif pattern_3.search(code):
                        line = pattern_3.search(code).group(1)
                        line = int(line)
                        print('line is:' + str(line))
                    elif pattern_4.search(code):
                        print(code)
                        ff = pattern_4.search(code).group(2)
                        logging.debug(ff)
                        if ff == '-':
                            data_dict[file_name].append(line)
                        if ff != '+':
                            line = line + 1

            logging.debug(data_dict)
        except Exception as e:
            logging.debug(e)
        print(data_dict)
        logging.debug('end get file line')
        return data_dict

    def parse_commit_hash(self, one_commit):
        tmp = dict()
        line_list = one_commit.strip().split("\n")
        tmp['subject'] = line_list[0]
        tmp['subject_len'] = len(line_list[0])
        #
        if tmp['subject'].find('cr_id_skip') > 0:
            tmp['cr_id_skip'] = True
        else:
            tmp['cr_id_skip'] = False
        #
        if tmp['subject'].find('3rd_skip') > 0:
            tmp['3rd_skip'] = True
        else:
            tmp['3rd_skip'] = False
        #
        if tmp['subject'].find('no_3rd_skip') > 0:
            tmp['no_3rd_skip'] = True
        else:
            tmp['no_3rd_skip'] = False
        tmp['body'] = '\n'.join(line_list[0:])
        tmp['reviewer:'] = list()
        for i in tmp['body'].split('\n'):
            if i.find('Reviewed By:') == 0:
                reviewer = i.replace('Reviewed By:', '')
                reviewer = list(
                    map(lambda x: x.strip(),
                        reviewer.split(',')))
                reviewer_person_list = list()
                for review in reviewer:
                    if not review.startswith("#"):
                        reviewer_person_list.append(review)
                tmp['reviewer'] = reviewer_person_list
                break
        ph_list = re.findall(
            'Differential Revision: https://cr.test.com/D(\d+)', tmp['body'])
        if ph_list:
            tmp['ph_id'] = ph_list[0]
        return tmp

    def get_all_commits_bug_introduce(self):
        bug_introduce_db_last = models.CodeReviewBugIntroduce.objects.last()
        last_commit_hash_id = models.Commit.objects.filter(
            commit_hash=bug_introduce_db_last.commit).last().id
        commits = models.Commit.objects.filter(id__gt=last_commit_hash_id)
        # commits = models.Commit.objects.all()
        for commit in commits:
            print("start:" + str(commit.id))
            if not commit.jira_ids:
                continue
            # if commit.commit_hash !=
            # "7a8609c6206a2e4ef8c801254cd563ebb6e98208":
            #     continue
            try:
                repo_id = commit.repo.repo_id
                cr_in_db = models.CodeReviewBugIntroduce.objects.filter(
                    commit=commit.commit_hash,
                    repo_id=repo_id
                )
                if cr_in_db:
                    continue
                repo_link = commit.repo.web_url
                if not commit.jira_ids:
                    continue
                for jira_id in json.loads(commit.jira_ids):
                    # print(jira_id)
                    jira = models.Jira.objects.filter(jira_id=jira_id)

                    if jira:
                        if jira[0].issue_type == 'Bug' \
                                or jira[0].issue_type == 'OBug':
                            print(jira[0])
                            commit.cr_ph = commit.cr_ph.replace('\'', '\"')
                            if commit.cr_ph:
                                commit.cr_ph = commit.cr_ph.replace('\'', '\"')
                                ph_json = json.loads(commit.cr_ph)
                                if not ph_json:
                                    continue
                                ph_id = json.loads(commit.cr_ph)[0]
                                introduced_data = self.deal_code_review_result(
                                    ph_id, repo_id)
                                # print(introduced_data)
                                for introduced in introduced_data:
                                    tmp = dict()
                                    tmp['jira_id'] = jira[0].jira_id
                                    tmp['commit'] = commit.commit_hash
                                    tmp['repo_link'] = repo_link
                                    tmp['ph_id'] = ph_id
                                    tmp['file_path'] = introduced['file_path']
                                    tmp['introduced_hash'] = introduced['hash']
                                    tmp['introduced_msg'] = introduced[
                                        'message']
                                    tmp['introduced_ph'] = introduced[
                                        'intro_ph_id']
                                    tmp['introduced_reviewer'] = json.dumps(
                                        introduced['reviewer'])
                                    tmp['introduced_authored_date'] = \
                                        introduced[
                                        'authored_date']
                                    tmp['introduced_committed_date'] = \
                                        introduced[
                                        'committed_date']
                                    tmp['introduced_author'] = introduced[
                                        'author']
                                    tmp['introduced_committer'] = introduced[
                                        'committer']
                                    tmp['repo_id'] = repo_id
                                    tmp['author_date'] = commit.author_date
                                    author_name = ''
                                    if commit.author:
                                        author_name = commit.author.user_id
                                    tmp['author_name'] = author_name
                                    tmp[
                                        'submitter_date'] = \
                                        commit.submitter_date
                                    print(tmp)
                                    self.update_data_in_db(tmp)
                                    self.result.append(tmp)
            except Exception as e:
                logging.error("Error report:" + str(e))
        generate_excel(self.result, 'report/bug_introduced', 'bug_introduced')

    def update_data_in_db(self, data):
        try:
            commit = data['commit']
            jira_id = data['jira_id']
            introduced_hash = data['introduced_hash']
            file_path = data['file_path']
            repo_id = data['repo_id']
            data.pop('commit')
            data.pop('jira_id')
            data.pop('introduced_hash')
            data.pop('file_path')
            data.pop('repo_id')
            models.CodeReviewBugIntroduce.objects.update_or_create(
                data,
                commit=commit,
                jira_id=jira_id,
                introduced_hash=introduced_hash,
                file_path=file_path,
                repo_id=repo_id
            )
            success_msg = 'Insert cr introduce success: ' + str(
                introduced_hash)
            logging.info(success_msg)
        except Exception as e:
            error_msg = 'Insert cr introduce error: ' + str(e)
            logging.error(error_msg)

    def get_ph_info(self, revision_id):
        data = dict()
        data['api.token'] = self.ph_api_token
        data['ids[0]'] = revision_id
        ph_info = dict()
        ph_info['title'] = ''
        try:
            differential_query = requests.post(
                self.ph_api_url + '/api/differential.query', data=data).json()
            title = differential_query['result'][0]['title']
            ph_info['title'] = title
        except Exception as e:
            logging.error("get ph info error:" + str(e))
        logging.info(ph_info)
        return ph_info

    def deal_code_review_result(self, ph_id, repo_id):
        """静态扫描结果写入文件"""
        print("ph_id：" + str(ph_id))
        create_dir('result')
        create_dir('middle')
        diff_id = self.get_diff_id(ph_id)
        self.diff_result(diff_id)
        self.write_diff_file()
        reviewer_result = set()
        commit_result = set()
        return_result = list()
        file_list = self.get_file_name_and_line()
        print(file_list)
        diff = self.get_diff_querydiff(diff_id)
        if not diff:
            return return_result
        parent = diff['sourceControlBaseRevision']
        print("parent:" + parent)
        # repo_id = 5885
        for file_path in file_list.keys():
            print(file_path + ":")
            blame_data = self.gitlab_get(repo_id, parent, file_path)
            # print("blame_data:")
            # print(blame_data)
            if not blame_data:
                continue
            key_id = 0
            for blame in blame_data:
                for line in blame["lines"]:
                    key_id = key_id + 1
                    if key_id in file_list[file_path]:
                        commit_file = blame["commit"]["id"] + '_' + file_path
                        if commit_file not in commit_result:
                            return_tmp = dict()
                            return_tmp['file_path'] = file_path
                            message = blame["commit"]["message"]
                            commit_data = self.parse_commit_hash(message)
                            # if "reviewer" in commit_data.keys():
                            #     for reviewer in commit_data["reviewer"]:
                            #         reviewer_result.add(reviewer)
                            name_regex = ".+\..+"
                            if re.search(name_regex,
                                         blame["commit"]["author_name"]):
                                author = blame["commit"]["author_name"].lower()
                            else:
                                author = \
                                    blame["commit"]["author_name"].split("@")[
                                        0]
                            if re.search(name_regex,
                                         blame["commit"]["committer_name"]):
                                committer = blame["commit"][
                                    "committer_name"].lower()
                            else:
                                committer = \
                                    blame["commit"]["author_name"].split("@")[
                                        0]
                            return_tmp['author'] = author
                            return_tmp['committer'] = committer
                            return_tmp['authored_date'] = blame["commit"][
                                "authored_date"]
                            return_tmp['committed_date'] = blame["commit"][
                                "committed_date"]
                            return_tmp['hash'] = blame["commit"]["id"]
                            return_tmp['message'] = commit_data['subject']
                            return_tmp['reviewer'] = list()
                            if "reviewer" in commit_data.keys():
                                return_tmp['reviewer'] = commit_data[
                                    "reviewer"]
                            return_tmp['intro_ph_id'] = 0
                            if "ph_id" in commit_data.keys():
                                return_tmp['intro_ph_id'] = commit_data[
                                    "ph_id"]
                            # print("-------------add:" + return_tmp['hash'])
                            return_result.append(return_tmp)
                        commit_result.add(commit_file)
        remove_dir('middle')
        remove_dir('result')

        # print("本次提交关联的reviewer: " + str(list(reviewer_result)))
        print("本次提交关联的commit: " + str(list(commit_result)))
        # print(return_result)
        return return_result

    def diff_result(self, diff_id):
        """请求增量接口"""
        url = "https://cr.test.com/api/differential.getrawdiff"
        token = 'api.token=&diffID={id}'
        payload = token.format(id=diff_id)
        headers = {
            'Accept': 'text/html,application/xhtml+xml,application/xml',
            'Accept-Language': 'zh-CN,zh;q=0.9',
        }
        try:
            response = requests.request("POST", url, headers=headers,
                                        data=payload)
            # print(response.content)
            with open(result_path, 'wb+') as f4:
                f4.write(response.content)
                print('-' * 10, response.content)
        except Exception as e:
            logging.debug(e)

    def get_diff_querydiff(self, diff_id):
        url = "https://cr.test.com/api/differential.querydiffs"
        token = ''
        data = dict()
        data['api.token'] = token
        data['ids[0]'] = diff_id
        headers = {
            'Accept': 'text/html,application/xhtml+xml,application/xml',
            'Accept-Language': 'zh-CN,zh;q=0.9',
        }
        try:
            response = requests.request("POST", url, headers=headers,
                                        data=data).json()
            return response['result'][diff_id]
        except Exception as e:
            logging.debug(e)

    def get_diff_id(self, ph_id):
        url = "https://cr.test.com/api/differential.query"
        token = ''
        data = dict()
        data['api.token'] = token
        data['ids[0]'] = ph_id

        try:
            differential_query = requests.post(url, data=data).json()
            return differential_query['result'][0]['diffs'][0]
        except Exception as e:
            logging.debug(e)

    def write_diff_file(self):
        """curl请求结果写入middle.txt"""
        try:
            with open(result_path, 'rb') as f:
                data = f.readline().decode('utf-8')
                data = json.loads(data)
                print(data)
                with open(middle_file, 'wb+') as f1:
                    f1.write(data['result'].encode('utf-8'))

        except Exception as e:
            logging.debug(e)


def create_dir(dir_name):
    if os.path.exists(dir_name):
        shutil.rmtree(dir_name)
    os.mkdir(dir_name)


def remove_dir(directory):
    """删除文件"""
    del_list = os.listdir(directory)
    try:
        for f in del_list:
            file_path = os.path.join(directory, f)
            if os.path.isfile(file_path):
                os.remove(file_path)
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)
    except Exception as e:
        logging.debug(e)


def write_diff_file():
    """curl请求结果写入middle.txt"""
    try:
        with open(result_path, 'rb') as f:
            data = f.readline().decode('utf-8')
            data = json.loads(data)
            print(data)
            with open(middle_file, 'wb+') as f1:
                f1.write(data['result'].encode('utf-8'))

    except Exception as e:
        logging.debug(e)


def run():
    cov = Cov()
    cov.get_all_commits_bug_introduce()
    cov.update_commit_message()
    # cov.get_ph_info('264962')
