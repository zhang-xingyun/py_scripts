import os
import logging
import json
import urllib.parse
from configparser import ConfigParser
from pygerrit2 import GerritRestAPI, HTTPBasicAuth
import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
logger = logging.getLogger()
logger.setLevel(logging.INFO)  # 设置打印级别
formatter = logging.Formatter(
    '%(asctime)s %(filename)s %(funcName)s [line:%(lineno)d] '
    '%(levelname)s %(message)s')

# 设置屏幕打印的格式
sh = logging.StreamHandler()
sh.setFormatter(formatter)
logger.addHandler(sh)

# 设置log保存
# fh = logging.FileHandler("gerrit_comments.log", encoding='utf8')
# fh.setFormatter(formatter)
# logger.addHandler(fh)


def generate_excel(ids_data, filename, name):
    print(filename)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = name

    for index, value in enumerate(ids_data[0].keys()):
        worksheet.cell(1, (index + 1), str(value))
    for row, one_data in enumerate(ids_data):
        # print(row)
        # print(one_data)
        for colum, value in enumerate(one_data.values()):
            # tmp_value = ILLEGAL_CHARACTERS_RE.sub(r'', str(value))
            tmp_value = value
            if isinstance(value, str):
                tmp_value = ILLEGAL_CHARACTERS_RE.sub(r'', str(value))
            else:
                tmp_value = value
            # print(type(tmp_value),tmp_value)
            worksheet.cell((row + 2), (colum + 1), tmp_value)
    file_path = '/'.join(filename.split('/')[:-1])

    print(file_path)
    if not os.path.exists(file_path):
        os.makedirs(file_path)
    workbook.save((filename + '.xlsx'))
    print('finish')


class Gerrit(object):
    def __init__(self, username, password):
        self.base_url = "https://gerrit.test.com:8443"
        self.username = username
        self.password = password
        self.auth = HTTPBasicAuth(self.username, self.password)
        self.rest = GerritRestAPI(self.base_url, self.auth)

    def _json_query(self, path):
        try:
            all_data = self.rest.get(path)
            return all_data
        except Exception as e:
            logger.error("Query with error:" + str(e))
            return None

    def get_projects(self):
        projects = self._json_query("projects/")
        return projects

    def get_change(self, change_number):
        raw = self._json_query("changes/?q=change:{}".format(change_number))
        if not raw or len(raw) == 0:
            return None
        return raw[0]

    def get_cs_changes(self):
        # 查找cs开头的项目中的changes，包含revision和commit数据
        raw = self._json_query(
            "changes/?q=projects:cs&no-limit&o=CURRENT_REVISION&o"
            "=CURRENT_COMMIT")
        print(len(raw))
        return raw

    def get_change_messages(self, change):
        raw = self._json_query(
            "changes/{}/messages".format(change['id'])
        )
        return raw

    def get_change_raw_diff(self, project, number, revision, path):
        print("fdsgf")
        raw = self._json_query(
            "changes/{}~{}/revisions/{}/files/{"
            "}/diff?context=ALL&intraline&whitespace=IGNORE_NONE".format(
                project,
                number,
                revision,
                urllib.parse.quote(path, safe=""),
            )
        )
        print(raw)

    def get_change_message_comments(self, change):
        raw = self._json_query(
            "changes/{}/comments".format(change['id'])
        )
        print(raw)
        if not raw:
            return list()
        commit_author = ''
        commit_committer = ''
        if change['current_revision']:
            current_revision = change['current_revision']
            print(change['revisions'][current_revision][
                'commit'])
            commit_author = change['revisions'][current_revision][
                'commit']['author']['email'].split('@')[0]
            commit_committer = change['revisions'][current_revision][
                'commit']['committer']['email'].split('@')[0]
        # dict with revision as key -> dict with path as key -> list of
        # comments on that rev/path.
        comments = list()
        path_line_set = set()
        for (path, comment_raw_list) in raw.items():
            for comment_raw in comment_raw_list:
                if comment_raw['author']['username'] == 'robot':
                    continue
                if 'line' in comment_raw.keys():
                    line = comment_raw['line']
                else:
                    line = ''
                path_line = path + '_' + str(line)
                if path_line in path_line_set:
                    continue
                link = self.base_url + '/c/' + change['project'] + '/+/' + str(
                    change['_number']) + '/' + str(
                    comment_raw['patch_set']) + '/' + path + '#' + str(line)
                comment_temp = dict()
                comment_temp['chang_number'] = change['_number']
                comment_temp['commit_author'] = commit_author
                comment_temp['commit_committer'] = commit_committer
                comment_temp['comment_author'] = comment_raw['author'][
                    'username']
                comment_temp['link'] = link
                comment_temp['project'] = change['project']
                comment_temp['statue'] = change['status']
                comment_temp['subject'] = change['subject']
                comment_temp['path'] = path
                comment_temp['message'] = comment_raw['message']
                comment_temp['commit_id'] = comment_raw['commit_id']
                comment_temp['unresolved'] = comment_raw['unresolved']
                comment_temp['patch_set'] = comment_raw['patch_set']
                comment_temp['updated'] = comment_raw['updated']
                comments.append(comment_temp)
                path_line_set.add(path_line)
        logger.info(comments)
        return comments


def main():
    cp = ConfigParser()
    config_path = os.path.dirname(os.path.abspath(__file__))
    print(config_path)
    cfg = os.path.join(config_path, 'gerrit_config.cfg')
    cp.read(cfg)
    username = cp.get('gerrit', 'username')
    password = cp.get('gerrit', 'password')
    gerrit = Gerrit(username, password)
    # gerrit.get_change_raw_diff("cs%2fbuild%2fdevice", 2886, 1, "")
    gerrit.get_projects()
    changes = gerrit.get_cs_changes()
    print(changes)
    total_comments = list()
    for change in changes:
        total_comments.extend(gerrit.get_change_message_comments(change))
    generate_excel(total_comments, 'report/gerrit_comments', 'gerrit_comments')


def run():
    main()
