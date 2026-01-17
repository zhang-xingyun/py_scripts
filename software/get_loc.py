#!/usr/bin/env python3
import logging
import os
import json
import datetime
import pickle
import time
import re
import subprocess
import threading
import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from gitlab_app import models


def generate_excel(ids_data, filename, name):
    print(filename)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = name

    for index, value in enumerate(ids_data[0].keys()):
        worksheet.cell(1, (index + 1), str(value))
    for row, one_data in enumerate(ids_data):
        # print(row)
        # print(one_data)
        for colum, value in enumerate(one_data.values()):
            # tmp_value = ILLEGAL_CHARACTERS_RE.sub(r'', str(value))
            tmp_value = value
            if isinstance(value, str):
                tmp_value = ILLEGAL_CHARACTERS_RE.sub(r'', str(value))
            else:
                tmp_value = value
            # print(type(tmp_value),tmp_value)
            worksheet.cell((row + 2), (colum + 1), tmp_value)
    file_path = '/'.join(filename.split('/')[:-1])

    print(file_path)
    if not os.path.exists(file_path):
        os.makedirs(file_path)
    workbook.save((filename + '.xlsx'))
    print('finish')


def gen_commit_times_by_department(dep_level: int, dep_name: str):
    person_list = []
    if dep_level == 1:
        person_list = models.People.objects.filter(businessUnit=dep_name)
    if dep_level == 2:
        person_list = models.People.objects.filter(division=dep_name)
    if dep_level == 3:
        person_list = models.People.objects.filter(department=dep_name)
    if dep_level == 4:
        person_list = models.People.objects.filter(org4=dep_name)

    logging.debug(person_list)
    person_id_list = [person.id for person in person_list]
    commit_list = models.Commit.objects.filter(author_id__in=person_id_list)
    data = list()
    for commit in commit_list:
        tmp = dict()
        tmp['commit_hash'] = commit.commit_hash
        tmp['author'] = commit.author.user_id
        tmp['additions'] = commit.additions
        tmp['deletions'] = commit.deletions
        tmp['total_code_line'] = commit.additions + commit.deletions
        tmp['files'] = commit.files
        tmp['author_date'] = str(commit.author_date)
        tmp['repo_id'] = commit.repo.repo_id
        tmp['cr_ph'] = commit.cr_ph
        data.append(tmp)
    print(data)
    generate_excel(data, 'report/commit_times-' + dep_name, 'commit_times')


def gen_commit_times_by_repo():
    repo_list = [7572, 7624, 7573, 7620]
    commit_list = models.Commit.objects.filter(repo__repo_id__in=repo_list,
                                               submitter_date__gt="2023-02-17")
    data = list()
    for commit in commit_list:
        tmp = dict()
        tmp['commit_hash'] = commit.commit_hash
        # tmp['author'] = commit.author.user_id
        # tmp['department'] = commit.author.org4
        # tmp['author_date'] = str(commit.author_date)
        tmp['submitter_date'] = str(commit.submitter_date)
        # tmp['repo_id'] = commit.repo.repo_id
        tmp['repo'] = commit.repo.name
        # tmp['cr_ph'] = commit.cr_ph
        data.append(tmp)
    print(data)
    generate_excel(data, 'report/commit_times', 'commit_times')


def parse_commit_info(commit_message_raw):
    jira_reg = r"([A-Z][A-Z0-9_]+-[1-9][0-9]*)"
    output = dict()
    output['jira_ids'] = list()
    id_check_list = commit_message_raw[commit_message_raw.find(
        '[') + 1:commit_message_raw.find(']')].split(' ')
    for id_check in id_check_list:
        if re.match(jira_reg, id_check):
            output['jira_ids'].append(id_check)
    #
    if commit_message_raw.find('cr_id_skip') > 0:
        output['cr_id_skip'] = True
    else:
        output['cr_id_skip'] = False
    #
    if commit_message_raw.find('3rd_skip') > 0:
        output['3rd_skip'] = True
    else:
        output['3rd_skip'] = False
    #
    if commit_message_raw.find('no_3rd_skip') > 0:
        output['no_3rd_skip'] = True
    else:
        output['no_3rd_skip'] = False
    return output

def parse_commit_author(author):
        tmp_author = author
        ret = True
        if tmp_author.find('test') < 0:
            ret = False
        if tmp_author.find('<') > 0:
            tmp_author = tmp_author.split('<')[1]
        if tmp_author.find('@') > 0:
            tmp_author = tmp_author.split('@')[0]
        return ret, tmp_author

def get_single_commit_info(commit_message_raw):
    commit_message = commit_message_raw.split('\n')
    print(commit_message)
    if commit_message[1].find('Merge:') == 0:
        tmp_dict = dict()
        print("Merge commit: " + str(commit_message[1]))
        author = commit_message[2].replace('Author:', '').strip()
        # if author.find('horizon') < 0 and \
        #   author.find('hogpu') < 0 and \
        #   author.find('hogpu') < 0 :
        #    continue
        ret, author = parse_commit_author(author)
        # if not ret:
        #     return tmp_dict
        author_date = commit_message[3].replace(
            'AuthorDate:', '').split('+')[0].strip()
        author_date = author_date.split('-')[0].strip()
        # print(author_date)
        author_date = time.strftime("%Y-%m-%d", time.strptime(author_date))
        # except Exception as e:
        #    print(e,'1111111111111111111111111111111')
        #    print(commit_message_raw)
        #
        submitter = commit_message[4].replace(
            'Commit:', '').split('+')[0].strip()
        submitter_data = commit_message[5].replace(
            'CommitDate:', '').split('+')[0].strip()
        submitter_data = submitter_data.split('-')[0].strip()
        submitter_data = time.strftime(
            "%Y-%m-%d", time.strptime(submitter_data))

        commit_info = parse_commit_info(
            commit_message[7])
        tmp_dict["author_date"] = author_date
        tmp_dict["submitter_data"] = submitter_data
        tmp_dict["jira_ids"] = commit_info["jira_ids"]
        tmp_dict["is_merge"] = True
        return tmp_dict
    else:
        tmp_dict = dict()
        author = commit_message[1].replace('Author:', '').strip()
        # if author.find('horizon') < 0 and \
        #   author.find('hogpu') < 0 and \
        #   author.find('hogpu') < 0 :
        #    continue
        ret, author = parse_commit_author(author)
        print(author)
        print(ret)
        # if not ret:
        #     return tmp_dict
        author_date = commit_message[2].replace(
            'AuthorDate:', '').split('+')[0].strip()
        author_date = author_date.split('-')[0].strip()
        # print(author_date)
        author_date = time.strftime("%Y-%m-%d", time.strptime(author_date))
        # except Exception as e:
        #    print(e,'1111111111111111111111111111111')
        #    print(commit_message_raw)
        #
        submitter = commit_message[3].replace(
            'Commit:', '').split('+')[0].strip()
        submitter_data = commit_message[4].replace(
            'CommitDate:', '').split('+')[0].strip()
        submitter_data = submitter_data.split('-')[0].strip()
        submitter_data = time.strftime(
            "%Y-%m-%d", time.strptime(submitter_data))

        commit_info = parse_commit_info(
            commit_message[6])
        # print(author_date)
        # print(commit_info)
        tmp_dict["author_date"] = author_date
        tmp_dict["submitter_data"] = submitter_data
        tmp_dict["jira_ids"] = commit_info["jira_ids"]
        tmp_dict["is_merge"] = False
        return tmp_dict


def gen_commit_times_by_repo_branch():
    raw_data_base = '/opt/gitlab_report/raw_data2/repo/'
    repo_branch_list = {
    }
    repo_list = repo_branch_list.keys()
    print(repo_list)
    data = list()
    sigal_data = set()
    for repo in repo_list:
        # commits = open(raw_data_base + str(repo) + '/commit_attributes.pkl', 'rb')
        f_obj = open('/mnt/hgfs/sharedir/commit_attributes.pkl',
                       'rb')
        # f_obj = open(raw_data_base + str(repo) + '/commit_attributes.pkl',
        #              'rb')
        commits = pickle.load(f_obj)
        for h, v in commits.items():
            if v['commit_id'] + str(v['project_id']) in list(sigal_data):
                continue
            sigal_data.add(v['commit_id'] + str(v['project_id']))
            # if v['commit_id'] == "d126c9bc896596c7eff8b90028375f4ab332d5e7":
            #     print(v)
            commit_info = get_single_commit_info(v['commit_info'])
            print("return commit_info")
            print(commit_info)
            if commit_info:
                commit_info['commit_id'] = v['commit_id']
                commit_info['repo_id'] = v['project_id']
                branch_info = v['branch_info']
                exist_branches = list()
                for branch in repo_branch_list[v['project_id']]:
                    if branch_info and branch in list(branch_info):
                        exist_branches.append(branch)
                commit_info['branch_list'] = str(exist_branches)
                if commit_info['jira_ids']:
                    print(commit_info['jira_ids'])
                commit_info['jira_ids'] = str(commit_info['jira_ids'])
                if commit_info['author_date'] and exist_branches and commit_info['author_date'] > "2022-06-01":
                    data.append(commit_info)
                # print(data)
    # print(data)
    generate_excel(data, 'report/commit_times', 'commit_times')


    # commit_list = models.Commit.objects.filter(
    #     repo__repo_id__in=repo_branch_list.keys(),
    #     submitter_date__gt="2022-06-01")
    # print(commit_list)
    # data = list()
    # for commit in commit_list:
    #     tmp = dict()
    #     exist_branches = list()
    #     branches = commit.branches.all()
    #     print(type(branches))
    #     print(branches)
    #
    #     branch_name_list = [br.branch_name for br in branches]
    #     for branch in repo_branch_list[commit.repo.repo_id]:
    #         if branch in branch_name_list:
    #             exist_branches.append(branch)
    #     if not exist_branches:
    #         continue
    #     tmp["branch_list"] = str(exist_branches)
    #     if not commit.jira_ids:
    #         commit.jira_ids = "[]"
    #     tmp["jira_ids"] = str(commit.jira_ids)
    #     tmp['commit_hash'] = commit.commit_hash
    #     # tmp['author'] = commit.author.user_id
    #     # tmp['department'] = commit.author.org4
    #     tmp['author_date'] = str(commit.author_date)
    #     tmp['submitter_date'] = str(commit.submitter_date)
    #     tmp['repo_id'] = commit.repo.repo_id
    #     tmp['repo_url'] = commit.repo.web_url
    #     # tmp['cr_ph'] = commit.cr_ph
    #     data.append(tmp)
    # print(data)
    # generate_excel(data, 'report/commit_times', 'commit_times')


def gen_cr_times_by_department(dep_level: int, dep_name: str):
    person_list = []
    if dep_level == 1:
        person_list = models.People.objects.filter(businessUnit=dep_name)
    if dep_level == 2:
        person_list = models.People.objects.filter(division=dep_name)
    if dep_level == 3:
        person_list = models.People.objects.filter(department=dep_name)
    if dep_level == 4:
        person_list = models.People.objects.filter(org4=dep_name)
    logging.debug(person_list)
    person_name_list = [person.user_id for person in person_list]
    cr_list = models.CodeReview.objects.filter(reviewer__in=person_name_list)
    data = list()
    for cr in cr_list:

        comments = list()
        if cr.comments:
            print(cr.comments)
            try:
                comm_json = json.loads(cr.comments)
            except Exception as e:
                print("json loads error:" + str(e))
                comm_json = list()
            for comm in comm_json:
                comm_tmp = dict()
                comm_tmp['time'] = datetime.datetime.utcfromtimestamp(
                    comm['time'])
                comm_tmp['comment'] = comm['comment']
                comments.append(comm_tmp)
        if len(comments) == 0:
            tmp = dict()
            tmp['reviewer'] = cr.reviewer
            tmp['ph_id'] = cr.ph_id
            tmp['start_time'] = str(cr.start_time)
            tmp['accept_time'] = str(cr.accept_time)
            tmp['comments_mum'] = cr.comments_mum
            tmp['duration'] = cr.duration
            tmp['repo_id'] = cr.repo_id
            tmp['comments_date'] = ""
            tmp['comments'] = ""
            data.append(tmp)
        else:
            for comment in comments:
                tmp = dict()
                tmp['reviewer'] = cr.reviewer
                tmp['ph_id'] = cr.ph_id
                tmp['start_time'] = str(cr.start_time)
                tmp['accept_time'] = str(cr.accept_time)
                tmp['comments_mum'] = cr.comments_mum
                tmp['duration'] = cr.duration
                tmp['repo_id'] = cr.repo_id
                tmp['comments_date'] = str(comment['time'])
                tmp['comments'] = str(comment['comment'])
                data.append(tmp)
    print(data)
    generate_excel(data, 'report/cr_times-' + dep_name, 'cr_times')


def get_loc_by_commit(repo_id, branch_name):
    commits = models.Commit.objects.filter(repo__repo_id=repo_id)
    total_line = 0
    for commit in commits:
        branches = commit.branches.all()
        if branch_name in [branch.branch_name for branch in branches]:
            print('repo:' + str(commit.repo.repo_id) + ' hash:' + commit.commit_hash + 'add:' + str(commit.additions) + 'delete:' + str(commit.deletions))
            total_line = total_line + commit.additions - commit.deletions
    print("repo:" + str(repo_id) + " total_line:" + str(total_line))


def runcmd(command):
    # print('run cmd---- ', command)
    try:
        ret = subprocess.run(
            command, shell=True,
            stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    except Exception as e:
        print(e)
        print('run cmd exception change to gpk ---- ', command)
        return (False, '')
    if ret.returncode != 0:
        # print('error', command, ret.stdout)
        print("error:", command, ret.returncode)
        # sys.exit(1)
        return (False, ''.join(map(chr, ret.stdout)))
    return (True, ''.join(map(chr, ret.stdout)))

def parse_loc_count(loc_info):
    print(time.ctime(), 'parse_loc_count')
    tmp_dict = dict()
    for line in loc_info.split('\n'):
        line_parse = line.split()
        if len(line_parse) != 6:
            continue
        if line_parse[0] == 'Language':
            continue
        # print(line_parse)
        if line_parse[0] not in tmp_dict.keys():
            tmp_dict[line_parse[0]] = dict()
        tmp_dict[line_parse[0]]['Files'] = line_parse[1]
        tmp_dict[line_parse[0]]['Lines'] = line_parse[2]
        tmp_dict[line_parse[0]]['Blank'] = line_parse[3]
        tmp_dict[line_parse[0]]['Comment'] = line_parse[4]
        tmp_dict[line_parse[0]]['Code'] = line_parse[5]
    return tmp_dict

def parse_loc_count(loc_info):
    print(time.ctime(), 'parse_loc_count')
    tmp_dict = dict()
    for line in loc_info.split('\n'):
        line_parse = line.split()
        if len(line_parse) != 6:
            continue
        if line_parse[0] == 'Language':
            continue
        # print(line_parse)
        if line_parse[0] not in tmp_dict.keys():
            tmp_dict[line_parse[0]] = dict()
        tmp_dict[line_parse[0]]['Files'] = line_parse[1]
        tmp_dict[line_parse[0]]['Lines'] = line_parse[2]
        tmp_dict[line_parse[0]]['Blank'] = line_parse[3]
        tmp_dict[line_parse[0]]['Comment'] = line_parse[4]
        tmp_dict[line_parse[0]]['Code'] = line_parse[5]
    return tmp_dict


def get_repo_loc(repo_ids, repos):
    while True:
        try:
            repo_id = repo_ids.pop()
        except Exception as e:
            logging.error(e)
            if len(repo_ids) == 0:
                break
            continue

        # tmp['repo_link'] = repo.web_url
        print(repo_id)
        debug_repos = [8189, 8485, 9236, 8791, 9263]
        if repo_id not in debug_repos:
            continue
        try:
            repo_dict = dict()
            repo_path = '/scm/gitlab_report/raw_data/repo/' + str(
                repo_id) + '/tmp_repo/'
            print(repo_path)
            if not os.path.exists(repo_path):
                print("no exist:" + repo_path)
                continue
            os.chdir(repo_path)
            runcmd('git pull ')
            if not ret:
                print("pull failed")
                continue
            branches = models.Branch.objects.filter(repo__repo_id=repo_id)
            for branch_name in [branch.branch_name for branch in branches]:
                tmp = dict()
                tmp['repo'] = repo_id
                tmp['branch'] = branch_name
                tmp['total_line'] = 0
                print(branch_name)

                ret, co = runcmd('git checkout ' + branch_name)

                if not ret:
                    print("checkout failed")
                    continue
                ret, count_info = runcmd('/root/tmp/gitlab/gitlab_report/loc')
                if ret:
                    loc_tmp = parse_loc_count(count_info)
                    if loc_tmp:
                        total_line = loc_tmp['Total']['Lines']
                        tmp['total_line'] = total_line
                        totol_data.append(tmp)
                        print('repo:' + str(
                            repo.repo_id) + ' branch:' + branch_name +
                              'total_line:' + str(
                            total_line))
                else:
                    print('loc', 'error', ssh_url_to_repo)
        except Exception as e:
            continue


def get_loc():
    get_single_project_thread = 6
    totol_data = list()
    repos = models.Repo.objects.all()
    # threads = list()
    # repo_ids = [repo.repo_id for repo in repos]
    # tmp_dic = dict()
    # for i in range(get_single_project_thread):
    #     tr = threading.Thread(
    #         target=get_repo_loc, args=(repo_ids, tmp_dic))
    #     threads.append(tr)
    # for i in threads:
    #     i.start()
    # for i in threads:
    #     i.join()
    for repo in repos:
        #debug_repo = [8189, 8485, 9236, 8791, 9263]
        print(repo.id)
        #if not repo.repo_id in debug_repo:
        #    continue
        try:
            repo_dict = dict()
            repo_path = '/scm/gitlab_report/raw_data/repo/' + str(
                repo.repo_id) + '/tmp_repo/'
            # print(repo_path)
            if not os.path.exists(repo_path):
                # print("no exist:" + repo_path)
                continue

            os.chdir(repo_path)
            branches = models.Branch.objects.filter(repo__repo_id=repo.repo_id)
            for branch_name in [branch.branch_name for branch in branches]:
                print(branch_name)
                tmp = dict()
                tmp['repo_id'] = repo.repo_id
                tmp['web_url'] = repo.web_url
                tmp['branch'] = branch_name
                ret, count_info = runcmd('git checkout ' + branch_name)
                if not ret:
                    continue
                ret, count_info = runcmd('/root/tmp/gitlab/gitlab_report/loc')
                if ret:
                    loc_tmp = parse_loc_count(count_info)
                    if loc_tmp:
                        # print(loc_tmp)
                        total_line = loc_tmp.get('Total', dict()).get('Lines', 0)
                        total_comment = loc_tmp.get('Total', dict()).get('Comment', 0)
                        total_code = loc_tmp.get('Total', dict()).get('Code', 0)
                        total_files = loc_tmp.get('Total', dict()).get('Files', 0)
                        c_plus_lines = loc_tmp.get('C++', dict()).get('Lines', 0)
                        c_plus_comment = loc_tmp.get('C++', dict()).get('Comment', 0)
                        c_header_lines = loc_tmp.get('C/C++ Header', dict()).get('Lines', 0)
                        c_header_comment = loc_tmp.get('C/C++ Header', dict()).get('Comment', 0)
                        c_lines = loc_tmp.get('C', dict()).get('Lines', 0)
                        c_comment = loc_tmp.get('C', dict()).get('Comment', 0)
                        c_c_plus_lines = int(c_plus_lines) + int(c_header_lines) + int(c_lines)
                        c_c_plus_comment = int(c_plus_comment) + int(c_header_comment) + int(c_comment)
                        tmp["total_lines"] = total_line
                        tmp["total_comment"] = total_comment
                        tmp["total_code"] = total_code
                        tmp["total_files"] = total_files

                        tmp["c_plus_lines"] = c_plus_lines
                        tmp["c_plus_comment"] = c_plus_comment
                        tmp["c_header_lines"] = c_header_lines
                        tmp["c_header_comment"] = c_header_comment
                        tmp["c_lines"] = c_lines
                        tmp["c_comment"] = c_comment
                        tmp["c_c_plus_lines"] = c_c_plus_lines
                        tmp["c_c_plus_comment"] = c_c_plus_comment

                        totol_data.append(tmp)
                        print('repo:' + str(repo.repo_id) + ' branch:' + branch_name + 'total_line:' + str(total_line))
                else:
                    print('loc', 'error', ssh_url_to_repo)
        except Exception as e:
            print("error:" + str(e))
            continue

        # print(totol_data)
    date_now = datetime.datetime.today()
    gen_date = date_now.strftime("%Y-%m-%d")
    for item in totol_data:
        try:
            branch_name = item['branch']
            repo_id = item['repo_id']
            item.pop('branch')
            item.pop('repo_id')
            logging.info("insert:" + str(repo_id) + str(branch_name))
            models.RepoLoc.objects.update_or_create(
                item,
                repo_id=repo_id,
                branch_name=branch_name,
                cacu_date=gen_date)
        except Exception as e:
            logging.error('Insert loc error: ', str(e))
    # save_pkl("/tmp/line.pkl", totol_data)
    # generate_excel(totol_data, '/tmp/repolines', 'repo_lines')


def gen_loc_by_pkl():
    f_obj = open('/mnt/hgfs/sharedir/line.pkl', 'rb')
    lines = pickle.load(f_obj)
    generate_excel(lines, 'report/repo_lines', 'repo_lines')


def run():
    # gen_commit_times_by_department(1, "软件平台")
    # gen_cr_times_by_department(1, "软件平台")
    # gen_commit_times_by_repo()
    # gen_commit_times_by_repo_branch()
    # get_loc_by_commit(8189, 'master')
    get_loc()
    # gen_loc_by_pkl()

