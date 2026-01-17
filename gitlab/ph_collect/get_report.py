from collections import OrderedDict
import sys
import pickle
import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import copy
import time
import gc

def save_pkl(file_name,data):
    f_obj = open(file_name, 'wb')
    pickle.dump(data, f_obj)
    f_obj.close()

def read_pkl(file_name):
    f_obj = open(file_name, 'rb')
    return pickle.load(f_obj)

def clean_user():
    output = dict()
    data = read_pkl('users.pkl')
    print(len(data))
    for i in data:
        output[i['phid']] = dict()
        output[i['phid']]['id'] = i['id']
        output[i['phid']]['type'] = i['type']
        output[i['phid']]['username'] = i['fields']['username']
        output[i['phid']]['realName'] = i['fields']['realName']
    print(len(output))
    return output

def clean_repo():
    output = dict()
    data = read_pkl('repository.pkl')
    print(len(data))
    for i in data:
        output[i['phid']] = 'http://gitlab.test.com/'+ i['fields']['name']
    print(len(output))
    return output

def clean_diff_bak(repo_data):
    output = dict()
    data = read_pkl('diffs.pkl')
    print(len(data))
    for i in data:
        output[i['phid']] = dict()
        output[i['phid']]['branch'] = 'no_data'
        output[i['phid']]['repo'] = 'no_data'
        #print(i)
        if i['fields']['repositoryPHID'] in repo_data.keys():
            output[i['phid']]['repo'] = repo_data[i['fields']['repositoryPHID']]
            output[i['phid']]['branch'] = 'unknown'
            for j in i['fields']['refs']:
                if j['type'] == 'branch':
                    output[i['phid']]['branch'] = j['name']
                    break
    print(len(output))
    return output

def clean_diff(repo_data):
    #ids_data = read_pkl('ids.pkl')
    output = dict()
    data = read_pkl('diffs.pkl')
    for i in data:
        if i['fields']['revisionPHID'] not in output.keys():
            output[i['fields']['revisionPHID']] = dict()
            output[i['fields']['revisionPHID']]['branch'] = 'no_data'
            output[i['fields']['revisionPHID']]['repo'] = 'no_data'
        if i['fields']['repositoryPHID'] in repo_data.keys():
            output[i['fields']['revisionPHID']]['repo'] = repo_data[i['fields']['repositoryPHID']]
            tmp_onto = None
            tmp_branch = None
            for j in i['fields']['refs']:
                if j['type'] == 'onto':
                    tmp_onto = j['name']
                if j['type'] == 'branch':
                    tmp_branch = j['name']
            if output[i['fields']['revisionPHID']]['branch'] == 'no_data':
                if not tmp_onto:
                    output[i['fields']['revisionPHID']]['branch'] = tmp_onto
                elif tmp_branch:
                    output[i['fields']['revisionPHID']]['branch'] = tmp_branch
                else:
                    pass
    return output

def clean_comments(user_data):
    data = read_pkl('transaction.pkl')
    print(len(data))
    output = dict()
    for case_id, comments in data.items():
        output[case_id] = list()
        for i in reversed(comments):
            if not i['type']:
                continue
            if i['authorPHID'] not in user_data.keys():
                continue
            tmp_data = dict()
            tmp_data['type'] = i['type']
            tmp_data['author'] = user_data[i['authorPHID']]
            tmp_data['comments'] = i['comments']
            tmp_data['dateCreated'] = i['dateCreated']
            output[case_id].append(tmp_data)
    print(len(output))
    return output

def clean_ids(user_data, diff_data, comments_data):
    data = read_pkl('ids.pkl')
    print(len(data))
    output = dict()
    output_single = dict()
    for i in data:
        tmp_single = dict()
        tmp_single['id'] = i['id']
        tmp_single['title'] = i['fields']['title']
        tmp_single['case_author'] = user_data[i['fields']['authorPHID']]['username']
        tmp_single['status_value'] = i['fields']['status']['value']
        tmp_single['status_name'] = i['fields']['status']['name']
        tmp_single['is_closed'] = i['fields']['status']['closed']
        tmp_single['summary'] = i['fields']['summary']
        tmp_single['testPlan'] = i['fields']['testPlan']
        tmp_single['isDraft'] = i['fields']['isDraft']
        tmp_single['holdAsDraft'] = i['fields']['holdAsDraft']
        tmp_single['dateCreated'] = time.strftime("%Y/%m/%d",time.localtime(i['fields']['dateCreated']))
        tmp_single['repo'] = diff_data[i['phid']]['repo']
        tmp_single['branch'] = diff_data[i['phid']]['branch']
        i_year = tmp_single['dateCreated'].split('/')[0]
        #
        if i_year not in output.keys():
            output[i_year] = list()
        if i_year not in output_single.keys():
            output_single[i_year] = list()
        #
        output_single[i_year].append(tmp_single)
        for j in comments_data[i['id']]:
            tmp_data = dict()
            tmp_data = copy.deepcopy(tmp_single)
            tmp_data['tr_type'] = j['type']
            tmp_data['tr_author'] = j['author']['username']
            tmp_data['tr_comments'] = str()
            for x in j['comments']:
                x_auth = user_data[x['authorPHID']]['username']
                x_content = x['content']['raw']
                tmp_data['tr_comments'] = tmp_data['tr_comments'] + x_auth + '\r\n '
                tmp_data['tr_comments'] = tmp_data['tr_comments'] + x_content + '\r\n'
            #tmp_data['tr_dateCreated'] = j['dateCreated']
            tmp_data['tr_dateCreated'] = time.strftime("%Y/%m/%d",time.localtime(j['dateCreated']))
            output[i_year].append(tmp_data)
    #print(len(output_single))
    #print(len(output))
    return output_single, output


def generate_excel(ids_data, name, status = False):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active 
    worksheet.title = name
    
    for index,value in enumerate(ids_data[0].keys()):
        worksheet.cell(1, (index + 1), str(value))
    row = 1
    for one_data in ids_data:
        if status:
            if 'tr_author' in one_data.keys() and \
               one_data['tr_author'].strip() == 'robot' and \
               (one_data['tr_type'].strip() == 'comment' or one_data['tr_type'].strip() == 'title'):
                #print(one_data)
                #print(row)
                continue
        row = row + 1
        for colum,value in enumerate(one_data.values()):
            tmp_value = ILLEGAL_CHARACTERS_RE.sub(r'', str(value))
            worksheet.cell((row), (colum + 1), tmp_value)
    print('finish') 
    workbook.save((name + '.xlsx'))

if __name__ == "__main__":
    user_data = clean_user()
    repo_data = clean_repo()
    diff_data = clean_diff(repo_data)
    comments_data = clean_comments(user_data)
    ids_single, ids_data = clean_ids(user_data, diff_data, comments_data)
    del user_data, repo_data, diff_data, comments_data
    gc.collect()
    for i,j in ids_single.items():
        generate_excel(j, 'cr_case_'+i)
    del ids_single
    gc.collect()
    for i,j in ids_data.items():
        generate_excel(j, 'cr_detail_'+i, True)
