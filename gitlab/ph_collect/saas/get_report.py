from collections import OrderedDict
import sys
import pickle
import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import copy
import time
import gc

def save_pkl(file_name,data):
    f_obj = open(file_name, 'wb')
    pickle.dump(data, f_obj)
    f_obj.close()

def read_pkl(file_name):
    f_obj = open(file_name, 'rb')
    return pickle.load(f_obj)

def clean_user():
    output = dict()
    data = read_pkl('users.pkl')
    print(len(data))
    for i in data:
        output[i['phid']] = dict()
        output[i['phid']]['id'] = i['id']
        output[i['phid']]['type'] = i['type']
        output[i['phid']]['username'] = i['fields']['username']
        output[i['phid']]['realName'] = i['fields']['realName']
    print(len(output))
    return output

def clean_repo():
    output = dict()
    data = read_pkl('repository.pkl')
    print(len(data))
    for i in data:
        output[i['phid']] = 'http://gitlab.test.com/'+ i['fields']['name']
    print(len(output))
    return output

def clean_diff_bak(repo_data):
    output = dict()
    data = read_pkl('diffs.pkl')
    print(len(data))
    for i in data:
        output[i['phid']] = dict()
        output[i['phid']]['branch'] = 'no_data'
        output[i['phid']]['repo'] = 'no_data'
        #print(i)
        if i['fields']['repositoryPHID'] in repo_data.keys():
            output[i['phid']]['repo'] = repo_data[i['fields']['repositoryPHID']]
            output[i['phid']]['branch'] = 'unknown'
            for j in i['fields']['refs']:
                if j['type'] == 'branch':
                    output[i['phid']]['branch'] = j['name']
                    break
    print(len(output))
    return output

def clean_diff(repo_data):
    #ids_data = read_pkl('ids.pkl')
    output = dict()
    data = read_pkl('diffs.pkl')
    for i in data:
        if i['fields']['revisionPHID'] not in output.keys():
            output[i['fields']['revisionPHID']] = dict()
            output[i['fields']['revisionPHID']]['branch'] = 'no_data'
            output[i['fields']['revisionPHID']]['repo'] = 'no_data'
        if i['fields']['repositoryPHID'] in repo_data.keys():
            output[i['fields']['revisionPHID']]['repo'] = repo_data[i['fields']['repositoryPHID']]
            tmp_onto = None
            tmp_branch = None
            for j in i['fields']['refs']:
                if j['type'] == 'onto':
                    tmp_onto = j['name']
                if j['type'] == 'branch':
                    tmp_branch = j['name']
            if output[i['fields']['revisionPHID']]['branch'] == 'no_data':
                if not tmp_onto:
                    output[i['fields']['revisionPHID']]['branch'] = tmp_onto
                elif tmp_branch:
                    output[i['fields']['revisionPHID']]['branch'] = tmp_branch
                else:
                    pass
    return output

def clean_comments(user_data):
    data = read_pkl('transaction.pkl')
    print(len(data))
    output = dict()
    for case_id, comments in data.items():
        output[case_id] = list()
        for i in reversed(comments):
            if not i['type']:
                continue
            if i['authorPHID'] not in user_data.keys():
                continue
            tmp_data = dict()
            tmp_data['type'] = i['type']
            tmp_data['author'] = user_data[i['authorPHID']]
            tmp_data['comments'] = i['comments']
            tmp_data['dateCreated'] = i['dateCreated']
            output[case_id].append(tmp_data)
    print(len(output))
    return output

def clean_ids(user_data, diff_data, comments_data):
    data = read_pkl('ids.pkl')
    print(len(data))
    output = list()
    output_single = list()
    for i in data:
        tmp_single = dict()
        tmp_single['id'] = i['id']
        tmp_single['title'] = i['fields']['title']
        tmp_single['case_author'] = user_data[i['fields']['authorPHID']]['username']
        tmp_single['status_value'] = i['fields']['status']['value']
        tmp_single['status_name'] = i['fields']['status']['name']
        tmp_single['is_closed'] = i['fields']['status']['closed']
        tmp_single['summary'] = i['fields']['summary']
        tmp_single['testPlan'] = i['fields']['testPlan']
        tmp_single['isDraft'] = i['fields']['isDraft']
        tmp_single['holdAsDraft'] = i['fields']['holdAsDraft']
        tmp_single['dateCreated'] = time.strftime("%Y/%m/%d",time.localtime(i['fields']['dateCreated']))
        tmp_single['repo'] = diff_data[i['phid']]['repo']
        tmp_single['branch'] = diff_data[i['phid']]['branch']
        output_single.append(tmp_single)
        for j in comments_data[i['id']]:
            tmp_data = dict()
            tmp_data = copy.deepcopy(tmp_single)
            tmp_data['tr_type'] = j['type']
            tmp_data['tr_author'] = j['author']['username']
            tmp_data['tr_comments'] = str()
            for x in j['comments']:
                x_auth = user_data[x['authorPHID']]['username']
                x_content = x['content']['raw']
                tmp_data['tr_comments'] = tmp_data['tr_comments'] + x_auth + '\r\n '
                tmp_data['tr_comments'] = tmp_data['tr_comments'] + x_content + '\r\n'
            #tmp_data['tr_dateCreated'] = j['dateCreated']
            tmp_data['tr_dateCreated'] = time.strftime("%Y/%m/%d",time.localtime(j['dateCreated']))
            output.append(tmp_data)
    print(len(output_single))
    print(len(output))
    return output_single, output


def generate_excel(ids_data, name, status = False):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active 
    worksheet.title = name
    
    for index,value in enumerate(ids_data[0].keys()):
        worksheet.cell(1, (index + 1), str(value))
    row = 1
    for one_data in ids_data:
        #print(one_data)
        if status:
            if 'tr_author' in one_data.keys() and \
               one_data['tr_author'].strip() == 'robot' and \
               (one_data['tr_type'].strip() == 'comment' or one_data['tr_type'].strip() == 'title'):
                #print(one_data)
                #print(row)
                continue
        row = row + 1
        for colum,value in enumerate(one_data.values()):
            tmp_value = ILLEGAL_CHARACTERS_RE.sub(r'', str(value))
            worksheet.cell((row), (colum + 1), tmp_value)
    print('finish') 
    workbook.save(('report/' + name + '.xlsx'))

def saas_data_parse(ids_data, repo_list):
    output = dict()
    for one_data in ids_data:
        cr_repo = one_data['repo'].strip()
        if not one_data['branch']:
            continue
        cr_branch = one_data['branch'].strip()
        cr_tr_type = one_data['tr_type'].strip()
        cr_tr_author = one_data['tr_author'].strip()
        cr_tr_dateCreated = one_data['tr_dateCreated'].strip()
        cr_tr_year = cr_tr_dateCreated.split('/')[0]
        cr_tr_month = cr_tr_dateCreated.split('/')[1]
        
        if cr_repo not in repo_list:
            continue
        if cr_repo not in output.keys():
            output[cr_repo] = dict()
        if cr_branch not in output[cr_repo].keys():
            output[cr_repo][cr_branch] = dict()
            output[cr_repo][cr_branch]['total'] = dict()
            output[cr_repo][cr_branch]['year'] = dict()
            output[cr_repo][cr_branch]['year_month'] = dict()
            # total
            output[cr_repo][cr_branch]['total']['create'] = 0
            output[cr_repo][cr_branch]['total']['request-changes'] = 0
            output[cr_repo][cr_branch]['total']['comments'] = 0
        if cr_tr_year not in output[cr_repo][cr_branch]['year'].keys():
            # by year
            output[cr_repo][cr_branch]['year'][cr_tr_year] = dict()
            output[cr_repo][cr_branch]['year'][cr_tr_year]['create'] = 0
            output[cr_repo][cr_branch]['year'][cr_tr_year]['request-changes'] = 0
            output[cr_repo][cr_branch]['year'][cr_tr_year]['comments'] = 0
        if (cr_tr_year + '_' + cr_tr_month) not in output[cr_repo][cr_branch]['year_month'].keys():
            # by month
            output[cr_repo][cr_branch]['year_month'][cr_tr_year + '_' + cr_tr_month] = dict()
            output[cr_repo][cr_branch]['year_month'][cr_tr_year + '_' + cr_tr_month]['create'] = 0
            output[cr_repo][cr_branch]['year_month'][cr_tr_year + '_' + cr_tr_month]['request-changes'] = 0
            output[cr_repo][cr_branch]['year_month'][cr_tr_year + '_' + cr_tr_month]['comments'] = 0
        if cr_tr_type == 'create':
            output[cr_repo][cr_branch]['total']['create'] += 1
            output[cr_repo][cr_branch]['year'][cr_tr_year]['create'] += 1
            output[cr_repo][cr_branch]['year_month'][cr_tr_year + '_' + cr_tr_month]['create'] += 1
        if cr_tr_type == 'request-changes':
            output[cr_repo][cr_branch]['total']['request-changes'] += 1
            output[cr_repo][cr_branch]['year'][cr_tr_year]['request-changes'] += 1
            output[cr_repo][cr_branch]['year_month'][cr_tr_year + '_' + cr_tr_month]['request-changes'] += 1
        if cr_tr_type in ['comment', 'inline'] and cr_tr_author != 'robot':
            output[cr_repo][cr_branch]['total']['comments'] += 1
            output[cr_repo][cr_branch]['year'][cr_tr_year]['comments'] += 1
            output[cr_repo][cr_branch]['year_month'][cr_tr_year + '_' + cr_tr_month]['comments'] += 1
        save_pkl('saas_output.pkl', output)

    report = dict()
    for repo, repo_data in output.items():
        for branch, branch_data in repo_data.items():
            if 'total' not in report.keys():
                report['total'] = list()
            tmp_total = dict()
            tmp_total['repo'] = repo
            tmp_total['branch'] = branch
            tmp_total['create'] = branch_data['total']['create']
            tmp_total['request-changes'] = branch_data['total']['request-changes']
            tmp_total['comments'] = branch_data['total']['comments'] 
            report['total'].append(tmp_total)
            for year, year_data in branch_data['year'].items():
                if year not in report.keys():
                    report[year] = list()
                tmp_year = dict()
                tmp_year['repo'] = repo
                tmp_year['branch'] = branch
                tmp_year['create'] = year_data['create']
                tmp_year['request-changes'] = year_data['request-changes']
                tmp_year['comments'] = year_data['comments']
                report[year].append(tmp_year)
            for year_mouth, year_mouth_data in branch_data['year_month'].items():
                if year_mouth not in report.keys():
                    report[year_mouth] = list()
                tmp_mouth = dict()
                tmp_mouth['repo'] = repo
                tmp_mouth['branch'] = branch
                tmp_mouth['create'] = year_mouth_data['create'] 
                tmp_mouth['request-changes'] = year_mouth_data['request-changes']
                tmp_mouth['comments'] = year_mouth_data['comments']
                report[year_mouth].append(tmp_mouth)

    save_pkl('saas_report.pkl', report)
    for key, value in report.items():
        generate_excel(value, key)
        
    print('not in list')
    for i in repo_list:
        if i not in output.keys():
            print(i)


if __name__ == "__main__":
    repo_list = list(map(lambda x:x.strip(), open('repo_list.txt', 'r').readlines()))
    print(repo_list)
    user_data = clean_user()
    repo_data = clean_repo()
    diff_data = clean_diff(repo_data)
    comments_data = clean_comments(user_data)
    ids_single, ids_data = clean_ids(user_data, diff_data, comments_data)
    del user_data, repo_data, diff_data, comments_data
    gc.collect()
    saas_data_parse(ids_data, repo_list)
    #generate_excel(ids_single, 'cr_case')
    #generate_excel(ids_data, 'cr_saas', True)
