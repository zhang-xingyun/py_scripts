import get_raw_data
import pickle
import os
import time
import datetime
import sys
import re
import shutil
import argparse
import textwrap
import openpyxl
import subprocess
import matplotlib.pyplot as plt
from pathlib import Path
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from xml.dom.minidom import parse
from lxml import etree
import numpy as np
import gitlab
plt.rcParams['font.sans-serif']=['SimHei'] #用来正常显示中文标签
plt.rcParams['axes.unicode_minus']=False #用来正常显示负号

class Report():
    def __init__(self,
                 ):

        self.raw_data = get_raw_data.GetData("", False)
        self.current_time = int(time.time())
        self.department_info = dict()
        self.branch_block = False
        self.pingtai = [
        ]
        self.ph_data_path = '/opt/gitlab_report/ph_raw_data/'
        self.branch_match = 'master|^develop.+|^main.+|^release-.+|^sprint-.+|^feature-.+|^bugfix-.+|^hotfix-.+|^cicd-.+|^test-.+|^tool-.+'
        #self.branch_match = '.*'
        self.gl = gitlab.Gitlab.from_config('trigger', ['python-gitlab.cfg'])
        self.jira_data = self.read_pkl(self.raw_data.raw_data_path + 'jira.pkl')
        self.ph_data = self.read_pkl(self.ph_data_path + 'ids.pkl')
        self.ph_rat_data = dict()
        self.gen_ph_data_ids()
        self.raw_projects = self.read_pkl(self.raw_data.raw_data_path + 'projects_attributes.pkl')

    def gen_ph_data_ids(self):
        for k in self.ph_data:
            title_rat_result = re.findall('.*(RAT-\d+).*', k['fields']['title'])
            summary_rat_result = re.findall('.*(RAT-\d+).*', k['fields']['summary'])
            if title_rat_result:
                self.ph_rat_data['D' + str(k['id'])] = title_rat_result[0]
            if summary_rat_result:
                self.ph_rat_data['D' + str(k['id'])] = summary_rat_result[0]

    def gen_branch_active_status(self, days):
        active_repo_list = list()
        active_num_t = 0
        total_num = 0
        for pi in self.pingtai:
            tmp_dic = dict()
            tmp_dic['total_repo'] = []
            tmp_dic['active_repo'] = []
            tmp_dic['unactive_repo'] = []
            tmp_dic['total_branch'] = []
            tmp_dic['active_branch'] = []
            tmp_dic['need_be_frozen_branch_list'] = []
            self.department_info[pi] = tmp_dic

        active_repo_list.append()
        for project_id, project_data in self.raw_projects.items():
            total_num = total_num + 1
            branch_file_path = self.raw_data.raw_data_repo_path + str(project_id) + '/branches_attributes.pkl'
            if not os.path.exists(branch_file_path):
                continue
            group_full_path = self.raw_projects[project_id]['namespace']['full_path']
            if group_full_path.find('/3rd') > 0 or \
                group_full_path.find('/thirdparty') > 0 or \
                group_full_path.find('/third_party') > 0 or \
                group_full_path.find('/third-party') > 0 or \
                group_full_path.find('/3rdparty') > 0 or \
                group_full_path.find('third_party') == 0 :
                continue
            if self.raw_projects[project_id]['archived']:
                continue
            if self.raw_projects[project_id]['namespace']['kind'] == 'user':
                continue
            belong_to = ''
            if self.raw_projects[project_id]['description']:
                description = self.raw_projects[project_id]['description'].split('\r\n')
                if len(description) >= 6:
                    if description[1].find('归属') == 0:
                        belong_to = description[1].replace('归属:', '')
            project_last_activity_at = project_data['last_activity_at'].split('T')[0]
            inactive_day = int((self.current_time - time.mktime(time.strptime(project_last_activity_at, '%Y-%m-%d'))) / (24 * 60 * 60))
            if inactive_day < 180:
                active_num_t = active_num_t + 1
            #print('inactive_day:' + str(inactive_day))
            tag_list = project_data['tag_list']
            if tag_list:
                for dep in active_repo_list:
                    if tag_list[0] == 'owner:'+dep['tag平台']:
                        dep['total'] = dep['total'] + 1
                        self.department_info[dep['tag平台']]['total_repo'].append(project_id)
                        if inactive_day < 180:
                            if not project_id in self.department_info[dep['tag平台']]['active_repo']:
                                dep['active'] = dep['active'] + 1
                                self.department_info[dep['tag平台']]['active_repo'].append(project_id)
                        else:
                            if not project_id in self.department_info[dep['tag平台']]['unactive_repo']:
                                self.department_info[dep['tag平台']]['unactive_repo'].append(project_id)
            else:
                for dep in active_repo_list:
                    if belong_to == dep['des平台']:
                        dep['total'] = dep['total'] + 1
                        self.department_info[dep['tag平台']]['total_repo'].append(project_id)
                        if inactive_day < 180:
                            if not project_id in self.department_info[dep['tag平台']]['active_repo']:
                                dep['active'] = dep['active'] + 1
                                self.department_info[dep['tag平台']]['active_repo'].append(project_id)
                        else:
                            if not project_id in self.department_info[dep['tag平台']]['unactive_repo']:
                                self.department_info[dep['tag平台']]['unactive_repo'].append(project_id)
        for dep in active_repo_list:
            if dep['total'] != 0:
                dep['rate'] = dep['active'] / dep['total']
        print('active repo num = ' + str(active_num_t))
        print('total_num repo num = ' + str(total_num))

        for k in self.department_info:
            tmp_repo = dict()
            tmp_branch = dict()
            self.department_info[k]['total_repo'] = list(set(self.department_info[k]['total_repo']))
            self.department_info[k]['active_repo'] = list(set(self.department_info[k]['active_repo']))
            br_info = self.get_branch_info(self.department_info[k]['total_repo'])
            active_br_num = 0
            total_br_num = 0
            for br in br_info:
                if br['freeze']:
                    continue
                create_times = 89
                if re.search(self.branch_match, br['branch_name']):
                    total_br_num = total_br_num + 1
                    if br['branch_create'] != 'more_unknown':
                        create_days = int(
                            (self.current_time - time.mktime(time.strptime(br['branch_create'], '%Y-%m-%d'))) / (
                                        24 * 60 * 60))
                    #print('create_days:' + str(create_days))

                    if br['inactive_day'] > 90 and create_days > 90:
                        if re.search('^sprint-.+|^feature-.+|^bugfix-.+|^hotfix-.+|^cicd-.+|^test-.+|^tool-.+',
                                     br['branch_name']):
                            br_tmp = dict()
                            br_tmp['project_id'] = br['project_id']
                            br_tmp['project_path'] = br['project_path']
                            br_tmp['branch_name'] = br['branch_name']
                            br_tmp['branch_create_at'] = br['branch_create']
                            br_tmp['create_days'] = create_days
                            br_tmp['branch_last_commit_at'] = br['branch_last_commit']
                            br_tmp['inactive_day'] = br['inactive_day']
                            self.department_info[k]['need_be_frozen_branch_list'].append(br_tmp)
                    else:
                        active_br_num = active_br_num + 1
            self.department_info[k]['total_branch'] = total_br_num
            self.department_info[k]['active_branch'] = active_br_num

        print(str(self.department_info))

    def read_pkl(self, file_name):
        f_obj = open(file_name, 'rb')
        return pickle.load(f_obj)

    def get_xml_file(self, old_tag):
        branch_name = self.include_branch
        tmp_dir = 'tmp-' + self.include_repo
        if os.path.exists(tmp_dir):
            shutil.rmtree(tmp_dir)
        os.makedirs(tmp_dir)
        git_repo_cmd = 'git clone git@gitlab.test.com:ptd/ss/' + self.include_repo + '/manifest.git ' + tmp_dir
        print(git_repo_cmd)
        ret, output = self.runcmd(git_repo_cmd)
        if not ret:
            print(git_repo_cmd, ' error')
        current_folder = os.getcwd()
        os.chdir(tmp_dir)
        print(os.getcwd())
        git_checkout_cmd = 'git checkout -b %s origin/%s' % (
            branch_name, branch_name)

        ret, output = self.runcmd(git_checkout_cmd)
        if not ret:
            print(git_checkout_cmd, 'error')
            print(output)
            self.runcmd('git checkout ' + branch_name)
        xml_dir = Path(old_tag + '.xml')
        print('xml_dir is : ' + str(xml_dir))
        if xml_dir.exists():
            print('the xml is: ' + str(xml_dir))
            os.chdir(current_folder)
            return tmp_dir + '/' + str(xml_dir)
        return ''



    def parse_xml(self, xml_file):
        dom = parse(xml_file)
        data = dom.documentElement
        projects = data.getElementsByTagName('project')
        proj_list = list()
        for proj in projects:
            # 获取标签属性值
            proj_dic = dict()
            name = proj.getAttribute('name')
            path = proj.getAttribute('path')
            revision = proj.getAttribute('revision')
            upstream = proj.getAttribute('upstream')
            proj_dic['name'] = name
            proj_dic['path'] = path
            proj_dic['revision'] = revision
            #proj_dic['upstream'] = upstream
            proj_dic['proj_id'] = self.get_query_projects(name)
            proj_list.append(proj_dic)
            #self.query_projects.append(proj_dic['proj_id'])
            #self.get_commit_info_by_hash(proj_dic)
        #print('self.query_projects : ' +str(self.query_projects))
        return proj_list


            #print('name:', name, ', path:', path, 'revision:', revision, ', upstream:', upstream)
        #print('proj_list:S ' + str(proj_list)

    def get_parent_groups(self, group_id, parent_list):
        parent_id = self.raw_groups[group_id]['parent_id']
        if parent_id:
            parent_list.append(parent_id)
            self.get_parent_groups(parent_id, parent_list)

    def get_query_groups(self):
        to_be_include = list()
        to_be_exclude = list()
        group_list = list()
        for group_id in self.raw_groups.keys():
            parent_groups = list()
            self.get_parent_groups(group_id, parent_groups)
            if self.include_group:
                for include in self.include_group:
                    if include in parent_groups:
                        to_be_include.append(group_id)
            if self.exclude_group:
                for exclude in self.exclude_group:
                    if exclude in parent_groups:
                        to_be_exclude.append(group_id)
        if self.include_group:
            #print(to_be_include)
            to_be_include.extend(self.include_group)
            group_list = to_be_include
        else:
            #print('else')
            if not self.include_repo:
                #print('else None')
                group_list = self.raw_groups.keys()
        group_list = list(set(group_list))
        #print(len(group_list))
        if self.exclude_group:
            to_be_exclude.extend(self.exclude_group)
            to_be_exclude = list(set(to_be_exclude))
            for i in to_be_exclude:
                if i in group_list:
                    group_list.remove(i)
        #print(len(to_be_exclude))
        #print(len(group_list))
        return group_list

    def get_query_projects(self, path_str):
        project_list = list()
        for project_id, project_data in self.raw_projects.items():
            #print('project_id ' + str(project_id))
            #if project_id == 7972:
               # print('project_data '+ str(project_data['path_with_namespace']))
            if project_data['path_with_namespace'] == 'ptd/ss/'+ path_str:
                #print(project_data['path_with_namespace'] + ' == ' + 'ptd/ss/'+ path_str)
                print('project_id ' + str(project_id))
                return project_id
            #if project_data['path_with_namespace'] == 'prebuilts/root_hijack':
            #    print('project_data ' + str(project_data))
            #if project_data['namespace']['id'] in self.query_groups:
              #  project_list.append(project_id)
        #project_list = list(set(project_list))
        #print(len(project_list))
        #return project_list

    def generate_excel(self,ids_data, filename, name):
        print(filename)
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = name
    
        for index,value in enumerate(ids_data[0].keys()):
            worksheet.cell(1, (index + 1), str(value))
        for row, one_data in enumerate(ids_data):
            #print(row)
            #print(one_data)
            for colum,value in enumerate(one_data.values()):
                #tmp_value = ILLEGAL_CHARACTERS_RE.sub(r'', str(value))
                tmp_value = value
                if isinstance(value,str):
                    tmp_value = ILLEGAL_CHARACTERS_RE.sub(r'', str(value))
                else:
                    tmp_value = value
                #print(type(tmp_value),tmp_value)
                worksheet.cell((row + 2), (colum + 1), tmp_value)
        file_path = '/'.join(filename.split('/')[:-1])

        print(file_path)
        if not os.path.exists(file_path):
            os.makedirs(file_path) 
        workbook.save((filename + '.xlsx'))
        print('finish')

    def generate_group_excel(self):
        output_project = list()
        for group_id in self.query_groups:
            group = self.raw_groups[group_id]
            tmp_dict = dict()
            tmp_dict['id'] = group['id']
            tmp_dict['web_url'] = group['web_url']
            tmp_dict['name'] = group['name']
            tmp_dict['path'] = group['path']
            tmp_dict['description'] = group['description']
            tmp_dict['visibility'] = group['visibility']
            tmp_dict['full_name'] = group['full_name']
            tmp_dict['full_path'] = group['full_path']
            output_project.append(tmp_dict)
        if len(output_project) > 0:
            self.generate_excel(output_project,'report/group','group')

    def generate_members_excel(self):
        output = list()
        total = 0
        total_manage = 0
        total_miss = 0
        project_user_permission_file_path = self.raw_data.raw_data_path + '/project_user_permission.pkl'
        if not os.path.exists(project_user_permission_file_path):
            return
        project_user_permission = self.read_pkl(project_user_permission_file_path)
        for project_id in self.query_projects:
            if project_id not in project_user_permission.keys():
                continue
            #if self.raw_projects[project_id]['archived']:
            #    continue
            belong_to = str()
            if self.raw_projects[project_id]['description']:
                description = self.raw_projects[project_id]['description'].split('\r\n')
                if len(description) >= 6:
                    if description[1].find('归属') == 0:
                        belong_to = description[1].replace('归属:','')
            if not self.raw_projects[project_id]['archived']:
                total += 1
            p_ret = False
            for permission, user_list in project_user_permission[project_id].items():
                for username in user_list:
                    if username in []:
                        continue
                    tmp_dict = dict()
                    tmp_dict['name_with_namespace'] = self.raw_projects[project_id]['name_with_namespace']
                    tmp_dict['project_path'] = self.raw_projects[project_id]['web_url']
                    tmp_dict['username'] = username
                    tmp_dict['permission'] = permission
                    tmp_dict['freeze'] = self.raw_projects[project_id]['archived']
                    tmp_dict['归属'] = belong_to
                    output.append(tmp_dict)
                    if permission == 'Maintainer' or permission == 'Owner':
                        p_ret = True
            if not self.raw_projects[project_id]['archived']:
                if p_ret:
                    total_miss += 1
                else:
                    total_manage += 1
        if len(output) > 0:
            self.generate_excel(output,'report/member','member')
        print(total, total_manage, total_miss)
        summary = dict()
        summary['仓库总数'] = total 
        summary['未回收总数'] = total_miss
        summary['已回收总数'] = total_manage
        summary['完成情况'] = str(int(((total_manage/total) * 100))) + '%'
        print(summary)
        self.generate_excel([summary], 'report/member_summary', 'member_summary')

    def generate_project_excel(self):
        output_project = list()
        for project_id in self.query_projects:
            project = self.raw_projects[project_id]
            project_created_at = project['created_at'].split('T')[0]
            project_last_activity_at = project['last_activity_at'].split('T')[0]
            project_group = 'http://gitlab.test.com/' + project['namespace']['full_path']
            project_path = project['web_url']
            project_namespace = project['name_with_namespace']
            tmp_project = dict()
            tmp_project['project_id'] = project_id
            tmp_project['project_namespace'] = project_namespace
            tmp_project['project_path'] = project_path
            tmp_project['project_group'] = project_group
            tmp_project['project_created_at'] = project_created_at
            tmp_project['project_last_activity_at'] = project_last_activity_at
            tmp_project['inactive_day'] = int((self.current_time - time.mktime(time.strptime(project_last_activity_at,'%Y-%m-%d')))/(24*60*60))
            tmp_project['type'] = 'normal'
            tmp_project['freeze'] = False
            if project['archived']:
                tmp_project['freeze'] = 'True'
            branch_file_path = self.raw_data.raw_data_repo_path + str(project_id) + '/branches_attributes.pkl'
            #print(branch_file_path)
            if not os.path.exists(branch_file_path):
                tmp_project['type'] = 'empty'
            group_full_path = project['namespace']['full_path']
            if group_full_path.find('/3rd') > 0 or \
                group_full_path.find('/thirdparty') > 0 or \
                group_full_path.find('/third_party') > 0 or \
                group_full_path.find('/third-party') > 0 or \
                group_full_path.find('/3rdparty') > 0 or \
                group_full_path.find('third_party') == 0 :
                tmp_project['type'] = '3rd'
            if group_full_path.find('/experimental') > 0:
                tmp_project['type'] = 'experimental'
            tmp_project['项目'] = None
            tmp_project['归属'] = None
            tmp_project['状态'] = None
            tmp_project['发布'] = None
            tmp_project['分支管控'] = None
            tmp_project['审批人'] = None
            if project['description']:
                description = project['description'].split('\r\n')
                if len(description) >= 6:
                    if description[0].find('项目:') == 0:
                        tmp_project['项目'] = description[0].replace('项目:','')
                    if description[1].find('归属') == 0:
                        tmp_project['归属'] = description[1].replace('归属:','')
                    if description[2].find('状态') == 0:
                        tmp_project['状态'] = description[2].replace('状态:','')
                    if description[3].find('发布') == 0:
                        tmp_project['发布'] = description[3].replace('发布:','')
                    if description[4].find('分支管控') == 0:
                        tmp_project['分支管控'] = description[4].replace('分支管控:','')
                    if description[5].find('审批人') == 0:
                        tmp_project['审批人'] = description[5].replace('审批人:','')
            #print(description)
            output_project.append(tmp_project)
        if len(output_project) > 0:
            self.generate_excel(output_project,'report/project','project')

    def generate_branch_excel(self):
        output = list()
        for project_id in self.query_projects:
            #print(project_id)
            release_branch = list()
            if self.raw_projects[project_id]['description']:
                description = self.raw_projects[project_id]['description'].split('\r\n')
                if len(description) >= 6:
                    if description[4].find('分支管控') == 0:
                        release_branch = description[4].replace('分支管控:','').split('|')
            #
            branches_attributes_file_path = self.raw_data.raw_data_path + '/repo/' + str(project_id) + '/branches_attributes.pkl'
            if not os.path.exists(branches_attributes_file_path):
                continue
            branches_attributes = self.read_pkl(branches_attributes_file_path)
            #print(branches_attributes)
            # /opt/gitlab_report/raw_data/repo/7827/commits.pkl
            commits = dict()
            commits_file_path = self.raw_data.raw_data_path + '/repo/' + str(project_id) + '/commits.pkl'
            if os.path.exists(commits_file_path):
                commits = self.read_pkl(commits_file_path)
            #print(commits)
            # /opt/gitlab_report/raw_data/repo/7827/events_attributes.pkl
            events = list()
            events_file_path = self.raw_data.raw_data_path + '/repo/' + str(project_id) + '/events_attributes.pkl'
            if os.path.exists(events_file_path):
                events = self.read_pkl(events_file_path)
            # 
            for branch_name, branch in branches_attributes.items():
                freeze = False
                if branch['protected']:
                    if not branch['developers_can_push'] and \
                       not branch['developers_can_merge'] and \
                       not branch['can_push'] and \
                       not branch['default']:
                        freeze = True
                if self.raw_projects[project_id]['archived']:
                    freeze = True
                branch_create = 'more_unknown'
                author_username = 'more_unknown'
                for event in events:
                    if event['action_name'] == 'pushed new' and \
                       event['push_data']['ref'] == branch_name:
                        branch_create = event['created_at'].split('T')[0]
                        author_username = event['author_username']
                tmp_dict = dict()
                tmp_dict['project_id'] = project_id
                tmp_dict['project_namespace'] = self.raw_projects[project_id]['name_with_namespace']
                tmp_dict['project_path'] = self.raw_projects[project_id]['web_url']
                tmp_dict['project_group'] = 'http://gitlab.test.com/' + self.raw_projects[project_id]['namespace']['full_path']
                tmp_dict['branch_name'] = branch_name
                tmp_dict['freeze'] = freeze
                tmp_dict['author_username'] = author_username
                tmp_dict['branch_create'] = branch_create
                tmp_dict['branch_last_commit'] = str()
                if branch_name in commits.keys():
                    for i in commits[branch['name']]:
                        if i['date'] > tmp_dict['branch_last_commit']:
                            tmp_dict['branch_last_commit'] = i['date']
                if len(tmp_dict['branch_last_commit']) == 0:
                    if branch_name in commits.keys():
                        tmp_dict['branch_last_commit'] = commits[branch['name']][-1]['date']
                    else:
                        if len(commits) > 0:
                            tmp_dict['branch_last_commit'] = commits[list(commits.keys())[0]][-1]['date']
                        else:
                            tmp_dict['branch_last_commit'] = branch_create
                tmp_dict['inactive_day'] = int((self.current_time - time.mktime(time.strptime(tmp_dict['branch_last_commit'],'%Y-%m-%d')))/(24*60*60))
                if self.branch_block and not tmp_dict['freeze']:
                    block_branch_ret = False
                    if branch_create == 'more_unknown':
                        block_branch_ret = True
                    else:
                        block_branch_create_day = int((self.current_time - time.mktime(time.strptime(tmp_dict['branch_create'],'%Y-%m-%d')))/(24*60*60))
                        if block_branch_create_day > 365:
                            block_branch_ret = True
                    if tmp_dict['inactive_day'] > 365 and block_branch_ret:
                        print('block branch', self.raw_projects[project_id]['web_url'], branch_name)
                        try:
                            ctl_project = self.gl.projects.get(project_id, retry_transient_errors=True)
                            ctl_branch = ctl_project.branches.get(branch_name, retry_transient_errors=True)
                            ctl_branch.unprotect(retry_transient_errors=True)
                            ctl_project.protectedbranches.create({'name':branch_name, 'merge_access_level': 0, 'push_access_level': 0}, retry_transient_errors=True)
                        except Exception as e:
                            print('block branch error', self.raw_projects[project_id]['web_url'], branch_name, str(e))
                #sys.exit(1)
                tmp_dict['标记'] = None
                if branch_name in release_branch:
                    tmp_dict['标记'] = '主干/发布'
                if branch_name.lower().find('rel') == 0 or branch_name.lower().find('sprint') == 0:
                    tmp_dict['标记'] = '主干/发布'
                if branch_name in ['dev', 'master', 'develop']:
                    tmp_dict['标记'] = '主干/发布'
                output.append(tmp_dict)
        if len(output) > 0:
            self.generate_excel(output,'report/branch','branch')

    def get_branch_info(self,query_projects):
        output = list()
        for project_id in query_projects:
            #print(project_id)
            release_branch = list()
            if self.raw_projects[project_id]['description']:
                description = self.raw_projects[project_id]['description'].split('\r\n')
                if len(description) >= 6:
                    if description[4].find('分支管控') == 0:
                        release_branch = description[4].replace('分支管控:','').split('|')
            #
            branches_attributes_file_path = self.raw_data.raw_data_path + '/repo/' + str(project_id) + '/branches_attributes.pkl'
            if not os.path.exists(branches_attributes_file_path):
                continue
            branches_attributes = self.read_pkl(branches_attributes_file_path)
            #print(branches_attributes)
            # /opt/gitlab_report/raw_data/repo/7827/commits.pkl
            commits = dict()
            commits_file_path = self.raw_data.raw_data_path + '/repo/' + str(project_id) + '/commits.pkl'
            if os.path.exists(commits_file_path):
                commits = self.read_pkl(commits_file_path)
            #print(commits)
            # /opt/gitlab_report/raw_data/repo/7827/events_attributes.pkl
            events = list()
            events_file_path = self.raw_data.raw_data_path + '/repo/' + str(project_id) + '/events_attributes.pkl'
            if os.path.exists(events_file_path):
                events = self.read_pkl(events_file_path)
            #
            for branch_name, branch in branches_attributes.items():
                freeze = False
                if branch['protected']:
                    if not branch['developers_can_push'] and \
                       not branch['developers_can_merge'] and \
                       not branch['can_push'] and \
                       not branch['default']:
                        freeze = True
                if self.raw_projects[project_id]['archived']:
                    freeze = True
                branch_create = 'more_unknown'
                author_username = 'more_unknown'
                for event in events:
                    if event['action_name'] == 'pushed new' and \
                       event['push_data']['ref'] == branch_name:
                        branch_create = event['created_at'].split('T')[0]
                        author_username = event['author_username']
                tmp_dict = dict()
                tmp_dict['project_id'] = project_id
                tmp_dict['project_namespace'] = self.raw_projects[project_id]['name_with_namespace']
                tmp_dict['project_path'] = self.raw_projects[project_id]['web_url']
                tmp_dict['project_group'] = 'http://gitlab.test.com/' + self.raw_projects[project_id]['namespace']['full_path']
                tmp_dict['branch_name'] = branch_name
                tmp_dict['freeze'] = freeze
                tmp_dict['author_username'] = author_username
                tmp_dict['branch_create'] = branch_create
                tmp_dict['branch_last_commit'] = str()
                if branch_name in commits.keys():
                    for i in commits[branch['name']]:
                        #print(str(i))
                        if i['submitter_data'] > tmp_dict['branch_last_commit']:
                            tmp_dict['branch_last_commit'] = i['submitter_data']
                if len(tmp_dict['branch_last_commit']) == 0:
                    if branch_name in commits.keys():
                        tmp_dict['branch_last_commit'] = commits[branch['name']][-1]['submitter_data']
                    else:
                        if len(commits) > 0:
                            tmp_dict['branch_last_commit'] = commits[list(commits.keys())[0]][-1]['submitter_data']
                        else:
                            tmp_dict['branch_last_commit'] = branch_create
                tmp_dict['inactive_day'] = int((self.current_time - time.mktime(time.strptime(tmp_dict['branch_last_commit'],'%Y-%m-%d')))/(24*60*60))
                if self.branch_block and not tmp_dict['freeze']:
                    block_branch_ret = False
                    if branch_create == 'more_unknown':
                        block_branch_ret = True
                    else:
                        block_branch_create_day = int((self.current_time - time.mktime(time.strptime(tmp_dict['branch_create'],'%Y-%m-%d')))/(24*60*60))
                        if block_branch_create_day > 365:
                            block_branch_ret = True
                    if tmp_dict['inactive_day'] > 365 and block_branch_ret:
                        print('block branch', self.raw_projects[project_id]['web_url'], branch_name)
                        try:
                            ctl_project = self.gl.projects.get(project_id, retry_transient_errors=True)
                            ctl_branch = ctl_project.branches.get(branch_name, retry_transient_errors=True)
                            ctl_branch.unprotect(retry_transient_errors=True)
                            ctl_project.protectedbranches.create({'name':branch_name, 'merge_access_level': 0, 'push_access_level': 0}, retry_transient_errors=True)
                        except Exception as e:
                            print('block branch error', self.raw_projects[project_id]['web_url'], branch_name, str(e))
                #sys.exit(1)
                tmp_dict['标记'] = None
                if branch_name in release_branch:
                    tmp_dict['标记'] = '主干/发布'
                if branch_name.lower().find('rel') == 0 or branch_name.lower().find('sprint') == 0:
                    tmp_dict['标记'] = '主干/发布'
                if branch_name in ['dev', 'master', 'develop']:
                    tmp_dict['标记'] = '主干/发布'
                output.append(tmp_dict)
        return output
        #if len(output) > 0:
            #self.generate_excel(output,'report/branch','branch')

    def generate_count_excel(self):
        output = list()
        for project_id in self.query_projects:
            #print(project_id)
            release_branch = list()
            if self.raw_projects[project_id]['description']:
                description = self.raw_projects[project_id]['description'].split('\r\n')
                if len(description) >= 6:
                    if description[4].find('分支管控') == 0:
                        release_branch = description[4].replace('分支管控:','').split('|')
            #
            conuts_attributes_file_path = self.raw_data.raw_data_path + '/repo/' + str(project_id) + '/counts.pkl'
            if not os.path.exists(conuts_attributes_file_path):
                continue
            conuts_attributes = self.read_pkl(conuts_attributes_file_path)
            #print(conuts_attributes)
            for branch, count_data in conuts_attributes.items():
                for count_type, count in count_data.items():
                    tmp_dict = dict()
                    tmp_dict['project_id'] = project_id
                    tmp_dict['project_namespace'] = self.raw_projects[project_id]['name_with_namespace']
                    tmp_dict['project_path'] = self.raw_projects[project_id]['web_url']
                    tmp_dict['project_group'] = 'http://gitlab.test.com/' + self.raw_projects[project_id]['namespace']['full_path']
                    tmp_dict['branch_name'] = branch
                    tmp_dict['count_type'] = count_type
                    tmp_dict['Files'] = int(count['Files'])
                    tmp_dict['Lines'] = int(count['Lines'])
                    tmp_dict['Blank'] = int(count['Blank'])
                    tmp_dict['Comment'] = int(count['Comment'])
                    tmp_dict['Code'] = int(count['Code'])
                    tmp_dict['标记'] = None
                    if branch in release_branch:
                        tmp_dict['标记'] = 'Release'
                    if branch.find('rel') == 0 or branch.find('sprint') == 0:
                        tmp_dict['标记'] = 'Release'
                    output.append(tmp_dict)
        if len(output) > 0:
            self.generate_excel(output,'report/count','count')

    def generate_commit_excel(self):
        for project_id in self.query_projects:
            output = dict()
            #
            commits_attributes_file_path = self.raw_data.raw_data_path + '/repo/' + str(project_id) + '/commits.pkl'
            if not os.path.exists(commits_attributes_file_path):
                continue
            commits_attributes = self.read_pkl(commits_attributes_file_path)
            for branch, commit_data in commits_attributes.items():
                for commit in commit_data:
                    ret, author = self.parse_commit_author(commit['author'])
                    if not ret:
                        continue
                    tmp_dict = dict()
                    tmp_dict['project_id'] = project_id
                    tmp_dict['project_path'] = self.raw_projects[project_id]['web_url']
                    tmp_dict['project_group'] = 'http://gitlab.test.com/' + self.raw_projects[project_id]['namespace']['full_path']
                    tmp_dict['project_namespace'] = self.raw_projects[project_id]['name_with_namespace']
                    tmp_dict['branch_name'] = branch
                    tmp_dict['commit'] = commit['commit']
                    tmp_dict['author'] = author
                    tmp_dict['date'] = commit['date']
                    tmp_dict['files'] = int(commit['files'])
                    tmp_dict['additions'] = int(commit['additions'])
                    tmp_dict['deletions'] = int(commit['deletions'])
                    tmp_dict['total'] = tmp_dict['additions'] + tmp_dict['deletions']
                    tmp_dict['cr_ph'] = ','.join(map(lambda x:'D'+x, commit['cr_ph']))
                    tmp_dict['jira_id'] = list()
                    tmp_dict['jira_type'] = list()
                    for key, value in commit['jira_data'].items():
                        tmp_dict['jira_id'].append(key)
                        tmp_dict['jira_type'].append(value['fields']['issuetype']['name'])
                    tmp_dict['jira_id'] = ','.join(tmp_dict['jira_id'])
                    tmp_dict['jira_type'] = ','.join(tmp_dict['jira_type'])
                    year = tmp_dict['date'].split('-')[0]
                    if year not in output.keys():
                        output[year] = list()
                    output[year].append(tmp_dict)
            for year, data in output.items():
                if len(data) > 0:
                    self.generate_excel(data,'report/commit/' + year + '/' + str(project_id), 'commit')

    def generate_single_commit_excel(self):
        print('generate_single_commit_excel')
        output = dict()
        for project_id in self.query_projects:
            #
            single_commits_attributes_file_path = self.raw_data.raw_data_path + 'repo/' + str(project_id) + '/single_commits.pkl'
            #print(single_commits_attributes_file_path)
            if not os.path.exists(single_commits_attributes_file_path):
                continue
            single_commits_attributes = self.read_pkl(single_commits_attributes_file_path)
            for single, commit in single_commits_attributes.items():
                ret, author = self.parse_commit_author(commit['author'])
                if not ret:
                    continue
                tmp_dict = dict()
                tmp_dict['project_id'] = project_id
                tmp_dict['project_path'] = self.raw_projects[project_id]['web_url']
                tmp_dict['project_group'] = 'http://gitlab.test.com/' + self.raw_projects[project_id]['namespace']['full_path']
                tmp_dict['project_namespace'] = self.raw_projects[project_id]['name_with_namespace']
                tmp_dict['commit'] = commit['commit']
                tmp_dict['author'] = author
                tmp_dict['date'] = commit['date']
                tmp_dict['files'] = int(commit['files'])
                tmp_dict['additions'] = int(commit['additions'])
                tmp_dict['deletions'] = int(commit['deletions'])
                tmp_dict['total'] = tmp_dict['additions'] + tmp_dict['deletions']
                tmp_dict['cr_ph'] = ','.join(map(lambda x:'D'+x, commit['cr_ph']))
                tmp_dict['jira_id'] = list()
                tmp_dict['jira_type'] = list()
                for key, value in commit['jira_data'].items():
                    tmp_dict['jira_id'].append(key)
                    tmp_dict['jira_type'].append(value['fields']['issuetype']['name'])
                tmp_dict['jira_id'] = ','.join(tmp_dict['jira_id'])
                tmp_dict['jira_type'] = ','.join(tmp_dict['jira_type'])
                year = tmp_dict['date'].split('-')[0]
                if year not in output.keys():
                    output[year] = list()
                output[year].append(tmp_dict)
        for year, data in output.items():
            if len(data) > 0:
                self.generate_excel(data,'report/single_commit_' + year,'single_commit')

    def parse_commit_author(self, author):
        tmp_author = author
        ret = True
        if tmp_author.find('test') < 0:
            ret = False
        if tmp_author.find('<') > 0:
            tmp_author = tmp_author.split('<')[1]
        if tmp_author.find('@') > 0:
            tmp_author = tmp_author.split('@')[0]
        return ret, tmp_author

    def generate_mrs_excel(self):
        output = list()
        output_pipeline = list()
        output_discussion = list()
        output_participant = list()
        for project_id in self.query_projects:
            #
            mrs_attributes_file_path = self.raw_data.raw_data_path + '/repo/' + str(project_id) + '/mrs_attributes.pkl'
            if not os.path.exists(mrs_attributes_file_path):
                continue
            mrs_attributes = self.read_pkl(mrs_attributes_file_path)
            for mr in mrs_attributes:
                #print(mr['data'])
                tmp_dict = dict()
                tmp_dict['project_id'] = project_id
                tmp_dict['web_url'] = mr['data']['web_url']
                tmp_dict['title'] = mr['data']['title']
                tmp_dict['state'] = mr['data']['state']
                tmp_dict['created_at'] = mr['data']['created_at'].split('T')[0]
                tmp_dict['closed_or_merged'] = None
                tmp_dict['updated_at'] = mr['data']['updated_at'].split('T')[0]
                if mr['data']['merged_by'] is None:
                    tmp_dict['merged_by'] = None
                else:
                    tmp_dict['merged_by'] = mr['data']['merged_by']['username']
                tmp_dict['target_branch'] = mr['data']['target_branch']
                tmp_dict['source_branch'] = mr['data']['source_branch']
                tmp_dict['author'] = mr['data']['author']['username']
                tmp_dict['sha'] = mr['data']['sha']
                tmp_dict['project_path'] = self.raw_projects[project_id]['web_url']
                #output.append(tmp_dict)
                for pipeline in mr['pipeline']:
                    tmp_pipeline = {'mr_url': mr['data']['web_url']}
                    tmp_pipeline.update(pipeline)
                    output_pipeline.append(tmp_pipeline)
                #
                tmp_output_participant_list = list()
                for participant in mr['participant']:
                    tmp_output_participant_list.append(participant['username'])
                tmp_output_participant_list = list(set(tmp_output_participant_list))
                for i in tmp_output_participant_list:
                    if i == 'robot':
                        continue
                    tmp_participant = dict()
                    tmp_participant['mr_url'] = mr['data']['web_url']
                    tmp_participant['people'] = i
                    tmp_participant['project_path'] = self.raw_projects[project_id]['web_url']
                    output_participant.append(tmp_participant)
                #
                for discussion in mr['discussion']:
                    #print('------------------------------')
                    #print(discussion)
                    if discussion['system']:
                        if discussion['body'] == 'merged' or discussion['body'] == 'closed':
                            tmp_dict['closed_or_merged'] = discussion['updated_at'].split('T')[0]
                        continue
                    if discussion['author']['username'] == 'robot':
                        continue
                    if discussion['body'] == 'Jenkins please retry a build':
                        continue
                    tmp_discussion = dict()
                    tmp_discussion['mr_url'] = mr['data']['web_url']
                    tmp_discussion['body'] = discussion['body']
                    tmp_discussion['author'] = discussion['author']['username']
                    tmp_discussion['created_at'] = discussion['created_at'].split('T')[0]
                    tmp_discussion['updated_at'] = discussion['updated_at'].split('T')[0]
                    tmp_discussion['project_path'] = self.raw_projects[project_id]['web_url']
                    #tmp_pipeline['system'] = discussion['system']
                    output_discussion.append(tmp_discussion)
                output.append(tmp_dict)
                    
        if len(output) > 0:
            self.generate_excel(output,'report/mr','mr')
        if len(output_pipeline) > 0:
            self.generate_excel(output_pipeline,'report/mr_pipeline','mr_pipeline')
        if len(output_discussion) > 0:
            self.generate_excel(output_discussion,'report/mr_discussion','mr_discussion')
        if len(output_participant) > 0:
            self.generate_excel(output_participant,'report/mr_participant','mr_participant')

    def generate_issues_excel(self):
        output = list()
        for project_id in self.query_projects:
            #
            issues_attributes_file_path = self.raw_data.raw_data_path + '/repo/' + str(project_id) + '/issues_attributes.pkl'
            if not os.path.exists(issues_attributes_file_path):
                continue
            issues_attributes = self.read_pkl(issues_attributes_file_path)
            for issue in issues_attributes:
                tmp_dict = dict()
                tmp_dict['project_id'] = project_id
                tmp_dict['project_path'] = self.raw_projects[project_id]['web_url']
                tmp_dict['title'] = issue['title']
                tmp_dict['state'] = issue['state']
                tmp_dict['created_at'] = issue['created_at'].split('T')[0]
                tmp_dict['updated_at'] = issue['updated_at'].split('T')[0]
                if issue['closed_at'] is None:
                    tmp_dict['closed_at'] = None
                else:
                    tmp_dict['closed_at'] = issue['closed_at'].split('T')[0]
                if issue['closed_by'] is None:
                    tmp_dict['closed_by'] = None
                else:
                    tmp_dict['closed_by'] = issue['closed_by']['username']
                tmp_dict['author'] = issue['author']['username']
                tmp_dict['web_url'] = issue['web_url']
                output.append(tmp_dict)
        if len(output) > 0:
            self.generate_excel(output,'report/issue','issue')

    def collect_info_with_old_and_new_project(self, old_project_list, new_project_list):
        coll_proj = list()
        for new_p in new_project_list:
            pp = dict()
            for old_p in old_project_list:
                if old_p['proj_id'] == new_p['proj_id'] and old_p['upstream'] == new_p['upstream']:
                    pp['proj_id'] = new_p['proj_id']
                    pp['new_commit'] = new_p['revision']
                    pp['path'] = new_p['path']
                    pp['upstream'] = new_p['upstream']
                    pp['old_commit'] = old_p['revision']
            coll_proj.append(pp)
        return coll_proj

    def runcmd(self, command):
        # print('run cmd---- ', command)
        try:
            ret = subprocess.run(
                command, shell=True,
                stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        except Exception as e:
            print(e)
            print('run cmd exception change to gpk ---- ', command)
            return (False, '')
        if ret.returncode != 0:
            # print('error', command, ret.stdout)
            print("error:", command, ret.returncode)
            # sys.exit(1)
            return (False, ''.join(map(chr, ret.stdout)))
        return (True, ''.join(map(chr, ret.stdout)))
    def find_jira_type(self, jira_id):
        for project_id, project_data in self.jira_data.items():
            if project_id == jira_id:
                jira_type = project_data['fields']['issuetype']['name']
                print('Jira type :' + jira_type)
                return jira_type
        return ''

    def get_commit_info_between_hash(self,proj_list):
        total_result = list()
        for project in proj_list:
            output = dict()
            output['proj'] = project
            #
            commits_attributes_file_path = self.raw_data.raw_data_path + '/repo/' + str(project['proj_id']) + '/commits.pkl'
            commits_attributes = self.read_pkl(commits_attributes_file_path)
            for branch, commit_data in commits_attributes.items():
                if branch == project['upstream']:
                    start_to_count = dict()
                    start_to_count['count_tag'] = False
                    start_to_count['found_old_hash'] = False
                    start_to_count['found_new_hash'] = False
                    count_result = list()
                    for commit in sorted(commit_data, key = lambda i: i['submitter_data']):
                        if commit['commit'] == project['old_commit']:
                            start_to_count['found_old_hash'] = True
                            start_to_count['count_tag'] = True
                        if commit['commit'] == project['new_commit']:
                            start_to_count['found_new_hash'] = True
                            start_to_count['count_tag'] = False
                        ret, author = self.parse_commit_author(commit['author'])
                        if not ret:
                            continue
                        if start_to_count['count_tag']:
                            tmp_dict = dict()
                            tmp_dict['project_id'] = project['proj_id']
                            tmp_dict['project_path'] = self.raw_projects[project['proj_id']]['web_url']
                            tmp_dict['project_group'] = 'http://gitlab.test.com/' + \
                                                        self.raw_projects[project['proj_id']]['namespace']['full_path']
                            tmp_dict['project_namespace'] = self.raw_projects[project['proj_id']]['name_with_namespace']
                            tmp_dict['branch_name'] = branch
                            tmp_dict['commit'] = commit['commit']
                            tmp_dict['author'] = author
                            tmp_dict['date'] = commit['submitter_data']
                            tmp_dict['files'] = int(commit['files'])
                            tmp_dict['additions'] = int(commit['additions'])
                            tmp_dict['deletions'] = int(commit['deletions'])
                            tmp_dict['total'] = tmp_dict['additions'] + tmp_dict['deletions']
                            tmp_dict['cr_ph'] = ','.join(map(lambda x: 'D' + x, commit['cr_ph']))
                            tmp_dict['jira_id'] = list()
                            tmp_dict['jira_type'] = list()
                            for key in commit['jira_data']:
                                tmp_dict['jira_id'].append(key)
                                # tmp_dict['jira_type'].append(value['fields']['issuetype']['name'])
                            tmp_dict['jira_id'] = ','.join(tmp_dict['jira_id'])
                            tmp_dict['jira_type'] = ','.join(tmp_dict['jira_type'])
                            if tmp_dict['jira_id']:
                                tmp_dict['jira_type'] = self.find_jira_type(tmp_dict['jira_id'])
                            count_result.append(tmp_dict)
                    if start_to_count['found_old_hash'] and start_to_count['found_new_hash']:
                        total_result.extend(count_result)

            #total_jira_ids.extend(count_result['jira_id_list'])
        print('\n---------------TOTAL RESULT:---------------\n')
        #print('total result is: ' + str(total_result))
        self.generate_excel(total_result, 'report/summary', 'between')


            #for year, data in output.items():
                #if len(data) > 0:
                    #self.generate_excel(data,'report/commit/' + year + '/' + str(project_id), 'commit')
    def get_commit_info_in_several_month(self,month_num, proj_list):
        total_result = list()
        DateNow = datetime.datetime.today()
        DateNow = DateNow.strftime("%Y-%m")
        DateYear = datetime.datetime.today()
        DateYear = DateYear.strftime("%Y")
        DateMonth = datetime.datetime.today()
        DateMonth = DateMonth.strftime("%m")
        print(DateYear)
        print(DateMonth)
        print(DateNow)
        month_list = list()
        month_data = dict()

        for m in range(int(DateMonth)-int(month_num), int(DateMonth)+1):
            m_name = str(DateYear) + '-' + str(m).zfill(2)
            month_list.append(m_name)
            month_data[m_name] = dict()
            month_data[m_name]['commit_mun'] = 0
            month_data[m_name]['jira_num'] = 0
            month_data[m_name]['rat_num'] = 0
        print(month_list)
        for project in proj_list:
            output = dict()
            output['proj'] = project
            #
            commits_attributes_file_path = self.raw_data.raw_data_path + '/repo/' + str(project) + '/commits.pkl'
            commits_attributes = self.read_pkl(commits_attributes_file_path)
            #print('start to get data from: ' + str(project))
            for branch, commit_data in commits_attributes.items():
                #if branch == project['upstream']:
                    start_to_count = dict()
                    count_result = list()
                    for commit in sorted(commit_data, key = lambda i: i['submitter_data']):
                        ret, author = self.parse_commit_author(commit['author'])
                        if not ret:
                            continue
                        for m_date in month_list:
                            if re.search(m_date, commit['submitter_data']):
                                #print(commit['submitter_data'])
                                tmp_dict = dict()
                                tmp_dict['project_id'] = project
                                tmp_dict['project_path'] = self.raw_projects[project]['web_url']
                                tmp_dict['project_group'] = 'http://gitlab.test.com/' + \
                                                            self.raw_projects[project]['namespace']['full_path']
                                tmp_dict['project_namespace'] = self.raw_projects[project]['name_with_namespace']
                                tmp_dict['branch_name'] = branch
                                tmp_dict['commit'] = commit['commit']
                                tmp_dict['author'] = author
                                tmp_dict['date'] = commit['submitter_data']
                                tmp_dict['files'] = int(commit['files'])
                                tmp_dict['additions'] = int(commit['additions'])
                                tmp_dict['deletions'] = int(commit['deletions'])
                                tmp_dict['total'] = tmp_dict['additions'] + tmp_dict['deletions']
                                tmp_dict['cr_ph'] = ','.join(map(lambda x: 'D' + x, commit['cr_ph']))
                                tmp_dict['jira_id'] = list()
                                tmp_dict['jira_type'] = list()
                                for key in commit['jira_data']:
                                    tmp_dict['jira_id'].append(key)
                                    # tmp_dict['jira_type'].append(value['fields']['issuetype']['name'])
                                tmp_dict['jira_id'] = ','.join(tmp_dict['jira_id'])
                                tmp_dict['jira_type'] = ','.join(tmp_dict['jira_type'])
                                if tmp_dict['jira_id']:
                                    month_data[m_date]['jira_num'] = month_data[m_date]['jira_num'] + 1
                                    if re.search('RAT', tmp_dict['jira_id']):
                                        month_data[m_date]['rat_num'] = month_data[m_date]['rat_num'] + 1
                                    #tmp_dict['jira_type'] = self.find_jira_type(tmp_dict['jira_id'])
                                count_result.append(tmp_dict)
                                month_data[m_date]['commit_mun'] = month_data[m_date]['commit_mun'] + 1

                    total_result.extend(count_result)

            #total_jira_ids.extend(count_result['jira_id_list'])
        #print('\n---------------TOTAL RESULT:---------------\n')
        #print('total result is: ' + str(month_data))
        #return month_data
        self.generate_excel(total_result, 'report/commit_by_month', 'between')


            #for year, data in output.items():
                #if len(data) > 0:
                    #self.generate_excel(data,'report/commit/' + year + '/' + str(project_id), 'commit')

    def get_commit_info_by_month(self,proj_list):
        total_result = list()
        DateNow = datetime.datetime.today()
        DateNow = DateNow.strftime("%Y-%m")
        DateYear = datetime.datetime.today()
        DateYear = DateYear.strftime("%Y")
        DateMonth = datetime.datetime.today()
        DateMonth = DateMonth.strftime("%m")
        print(DateYear)
        print(DateMonth)
        print(DateNow)
        month_list = list()
        month_data = dict()

        for m in range(1, int(DateMonth)+1):
            m_name = str(DateYear) + '-' + str(m).zfill(2)
            month_list.append(m_name)
            month_data[m_name] = dict()
            month_data[m_name]['commit_mun'] = 0
            month_data[m_name]['jira_num'] = 0
            month_data[m_name]['rat_num'] = 0
        print(month_list)
        for project in proj_list:
            output = dict()
            output['proj'] = project
            #
            commits_attributes_file_path = self.raw_data.raw_data_path + '/repo/' + str(project) + '/commits.pkl'
            commits_attributes = self.read_pkl(commits_attributes_file_path)
            #print('start to get data from: ' + str(project))
            for branch, commit_data in commits_attributes.items():
                if re.search(self.branch_match, branch):
                    start_to_count = dict()
                    count_result = list()
                    for commit in sorted(commit_data, key = lambda i: i['submitter_data']):
                        ret, author = self.parse_commit_author(commit['author'])
                        if not ret:
                            continue
                        for m_date in month_list:
                            if re.search(m_date, commit['submitter_data']):
                                #print(commit['submitter_data'])
                                tmp_dict = dict()
                                tmp_dict['project_id'] = project
                                tmp_dict['project_path'] = self.raw_projects[project]['web_url']
                                tmp_dict['project_group'] = 'http://gitlab.test.com/' + \
                                                            self.raw_projects[project]['namespace']['full_path']
                                tmp_dict['project_namespace'] = self.raw_projects[project]['name_with_namespace']
                                tmp_dict['branch_name'] = branch
                                tmp_dict['commit'] = commit['commit']
                                tmp_dict['author'] = author
                                tmp_dict['date'] = commit['submitter_data']
                                tmp_dict['files'] = int(commit['files'])
                                tmp_dict['additions'] = int(commit['additions'])
                                tmp_dict['deletions'] = int(commit['deletions'])
                                tmp_dict['total'] = tmp_dict['additions'] + tmp_dict['deletions']
                                tmp_dict['cr_ph'] = ','.join(map(lambda x: 'D' + x, commit['cr_ph']))
                                tmp_dict['jira_id'] = list()
                                tmp_dict['jira_type'] = list()
                                for key in commit['jira_data']:
                                    tmp_dict['jira_id'].append(key)
                                    # tmp_dict['jira_type'].append(value['fields']['issuetype']['name'])
                                tmp_dict['jira_id'] = ','.join(tmp_dict['jira_id'])
                                tmp_dict['jira_type'] = ','.join(tmp_dict['jira_type'])
                                if commit['commit'] == 'b148da81ddcfe3ef3c89486db1ffbc7dfd9044a1':
                                    print(str(commit))
                                    print(str(tmp_dict))
                                if tmp_dict['jira_id']:
                                    month_data[m_date]['jira_num'] = month_data[m_date]['jira_num'] + 1
                                    if re.search('RAT', tmp_dict['jira_id']):
                                        print('rat: ' + tmp_dict['jira_id'])
                                        month_data[m_date]['rat_num'] = month_data[m_date]['rat_num'] + 1
                                    #tmp_dict['jira_type'] = self.find_jira_type(tmp_dict['jira_id'])
                                else:
                                    if tmp_dict['cr_ph'] in self.ph_rat_data.keys():
                                        tmp_dict['jira_id'] = str(self.ph_rat_data[tmp_dict['cr_ph']])
                                        print('no jira RAT: ' + str(self.ph_rat_data[tmp_dict['cr_ph']]))
                                        month_data[m_date]['rat_num'] = month_data[m_date]['rat_num'] + 1
                                count_result.append(tmp_dict)
                                month_data[m_date]['commit_mun'] = month_data[m_date]['commit_mun'] + 1

                    total_result.extend(count_result)

            #total_jira_ids.extend(count_result['jira_id_list'])
        #print('\n---------------TOTAL RESULT:---------------\n')
        #print('total result is: ' + str(month_data))
        return month_data
        #self.generate_excel(total_result, 'report/commit_by_month', 'between')


            #for year, data in output.items():
                #if len(data) > 0:
                    #self.generate_excel(data,'report/commit/' + year + '/' + str(project_id), 'commit')

    def get_commit_info_by_week(self,proj_list):
        total_result = list()
        DateNow = datetime.datetime.today()
        DateNow = DateNow.strftime("%Y-%m")
        DateYear = datetime.datetime.today()
        DateYear = DateYear.strftime("%Y")
        DateMonth = datetime.datetime.today()
        DateMonth = DateMonth.strftime("%m")
        print(DateYear)
        print(DateMonth)
        print(DateNow)
        month_list = list()
        month_data = dict()

        for m in range(1, int(DateMonth)+1):
            m_name = str(DateYear) + '-' + str(m).zfill(2)
            month_list.append(m_name)
            month_data[m_name] = dict()
            month_data[m_name]['commit_mun'] = 0
            month_data[m_name]['jira_num'] = 0
            month_data[m_name]['rat_num'] = 0
        print(month_list)
        for project in proj_list:
            output = dict()
            output['proj'] = project
            #
            commits_attributes_file_path = self.raw_data.raw_data_path + '/repo/' + str(project) + '/commits.pkl'
            commits_attributes = self.read_pkl(commits_attributes_file_path)
            #print('start to get data from: ' + str(project))
            for branch, commit_data in commits_attributes.items():
                if re.search(self.branch_match, branch):
                    start_to_count = dict()
                    count_result = list()
                    for commit in sorted(commit_data, key = lambda i: i['submitter_data']):
                        ret, author = self.parse_commit_author(commit['author'])
                        if not ret:
                            continue
                        for m_date in month_list:
                            if re.search(m_date, commit['submitter_data']):
                                #print(commit['submitter_data'])
                                tmp_dict = dict()
                                tmp_dict['project_id'] = project
                                tmp_dict['project_path'] = self.raw_projects[project]['web_url']
                                tmp_dict['project_group'] = 'http://gitlab.test.com/' + \
                                                            self.raw_projects[project]['namespace']['full_path']
                                tmp_dict['project_namespace'] = self.raw_projects[project]['name_with_namespace']
                                tmp_dict['branch_name'] = branch
                                tmp_dict['commit'] = commit['commit']
                                tmp_dict['author'] = author
                                tmp_dict['date'] = commit['submitter_data']
                                tmp_dict['files'] = int(commit['files'])
                                tmp_dict['additions'] = int(commit['additions'])
                                tmp_dict['deletions'] = int(commit['deletions'])
                                tmp_dict['total'] = tmp_dict['additions'] + tmp_dict['deletions']
                                tmp_dict['cr_ph'] = ','.join(map(lambda x: 'D' + x, commit['cr_ph']))
                                tmp_dict['jira_id'] = list()
                                tmp_dict['jira_type'] = list()
                                for key in commit['jira_data']:
                                    tmp_dict['jira_id'].append(key)
                                    # tmp_dict['jira_type'].append(value['fields']['issuetype']['name'])
                                tmp_dict['jira_id'] = ','.join(tmp_dict['jira_id'])
                                tmp_dict['jira_type'] = ','.join(tmp_dict['jira_type'])
                                if tmp_dict['jira_id']:
                                    month_data[m_date]['jira_num'] = month_data[m_date]['jira_num'] + 1
                                    if re.search('RAT', tmp_dict['jira_id']):
                                        print('rat: ' + tmp_dict['jira_id'])
                                        print('branch_name:' + branch)
                                        month_data[m_date]['rat_num'] = month_data[m_date]['rat_num'] + 1
                                    #tmp_dict['jira_type'] = self.find_jira_type(tmp_dict['jira_id'])
                                count_result.append(tmp_dict)
                                month_data[m_date]['commit_mun'] = month_data[m_date]['commit_mun'] + 1

                    total_result.extend(count_result)

            #total_jira_ids.extend(count_result['jira_id_list'])
        #print('\n---------------TOTAL RESULT:---------------\n')
        #print('total result is: ' + str(month_data))
        return month_data
        #self.generate_excel(total_result, 'report/commit_by_month', 'between')


            #for year, data in output.items():
                #if len(data) > 0:
                    #self.generate_excel(data,'report/commit/' + year + '/' + str(project_id), 'commit')

    def gen_active_repo_branch_data(self):
        repo_biao = dict()
        branch_biao = dict()
        active_repo_list = list()
        active_branch_list = list()
        toatl_active_branch = dict()
        total_active_repo = dict()
        total_active_repo['项目'] = '总项目'
        total_active_repo['总仓库'] = 0
        total_active_repo['活仓库数'] = 0
        total_active_branch = dict()
        total_active_branch['项目'] = '总分支'
        total_active_branch['总分支'] = 0
        total_active_branch['活分支数'] = 0
        for k in self.department_info:
            tmp_repo = dict()
            tmp_repo['项目'] = k
            tmp_repo['总仓库数'] = len(self.department_info[k]['total_repo'])
            tmp_repo['活仓库数'] = len(self.department_info[k]['active_repo'])
            tmp_repo['活跃度'] = 100 * len(self.department_info[k]['active_repo']) / len(
                self.department_info[k]['total_repo'])
            total_active_repo['总仓库'] = total_active_repo['总仓库'] + len(self.department_info[k]['total_repo'])
            total_active_repo['活仓库数'] = total_active_repo['活仓库数'] + len(self.department_info[k]['active_repo'])
            active_repo_list.append(tmp_repo)
            repo_biao[k] = tmp_repo['活跃度']

            tmp_branch = dict()
            tmp_branch['项目'] = k
            tmp_branch['总分支数'] = self.department_info[k]['total_branch']
            tmp_branch['活分支数'] = self.department_info[k]['active_branch']
            total_active_branch['总分支'] = total_active_branch['总分支'] + self.department_info[k]['total_branch']
            total_active_branch['活分支数'] = total_active_branch['活分支数'] + self.department_info[k]['active_branch']
            tmp_branch['活跃度'] = 100 * self.department_info[k]['active_branch'] / self.department_info[k]['total_branch']
            active_branch_list.append(tmp_branch)
            branch_biao[k] = tmp_branch['活跃度']

            if self.department_info[k]['need_be_frozen_branch_list']:
                self.generate_excel(self.department_info[k]['need_be_frozen_branch_list'],'report/' + k + '-need_be_frozen_branch_list-{}'.format(time.strftime("%Y-%m-%d")), 'between')

        active_repo_list.append(total_active_repo)
        active_branch_list.append(total_active_branch)
        total_repo_active_rate = 100*total_active_repo['活仓库数'] / total_active_repo['总仓库']
        total_branch_active_rate = 100*total_active_branch['活分支数'] / total_active_branch['总分支']
        total_active_repo['活跃度'] = total_repo_active_rate
        total_active_branch['活跃度'] = total_branch_active_rate
        repo_biao['总仓库'] = total_repo_active_rate
        branch_biao['总分支'] = total_branch_active_rate

        self.generate_excel(active_repo_list, 'report/repo_active', 'between')
        plt.figure(1)
        plt.title(u'仓库活跃度')
        plt.xlabel(u"项目")  # x轴名称
        plt.ylabel(u"百分比")  # y 轴名称
        repo_rate_p = plt.bar(repo_biao.keys(), repo_biao.values())
        self.autolabel(repo_rate_p)
        plt.savefig('report/repo_active.jpg')
        plt.close(1)

        self.generate_excel(active_branch_list, 'report/active_branch_list', 'between')
        plt.figure(2)
        plt.title(u'分支活跃度')
        plt.xlabel(u"项目")  # x轴名称
        plt.ylabel(u"百分比")  # y 轴名称
        branch_rate_p = plt.bar(branch_biao.keys(), branch_biao.values())
        self.autolabel(branch_rate_p)
        plt.savefig('report/active_branch_list.jpg')
        plt.close(2)

    def autolabel(self, rects):
        for rect in rects:
            height = rect.get_height()
            plt.text(rect.get_x() + rect.get_width() / 2. - 0.2, 1.03 * height, '%s' % int(height))

    def get_project_month_data(self):
        #print(str(self.jira_data))
        self.gen_branch_active_status(90)
        self.gen_active_repo_branch_data()
        total_dep_commit_data = list()
        for k in self.department_info:
            commit_jira_month = list()
            month_data = self.get_commit_info_by_month(self.department_info[k]['total_repo'])
            tmp_data_dic_total = dict()
            tmp_data_dic_total['月份'] = '总计'
            tmp_data_dic_total['commit'] = 0
            tmp_data_dic_total['jira'] = 0
            tmp_data_dic_total['RAT'] = 0
            for mon in month_data:
                tmp_data_dic = dict()
                tmp_data_dic['月份'] = mon
                tmp_data_dic['commit'] = month_data[mon]['commit_mun']
                tmp_data_dic['jira'] = month_data[mon]['jira_num']
                tmp_data_dic['RAT'] = month_data[mon]['rat_num']
                tmp_data_dic_total['commit'] = tmp_data_dic_total['commit'] + month_data[mon]['commit_mun']
                tmp_data_dic_total['jira'] = tmp_data_dic_total['jira'] + month_data[mon]['jira_num']
                tmp_data_dic_total['RAT'] = tmp_data_dic_total['RAT'] + month_data[mon]['rat_num']
                commit_jira_month.append(tmp_data_dic)
            # 部门整体commit jira RAT数据dict
            dep_commit_data = dict()
            dep_commit_data['部门'] = k
            dep_commit_data['commit'] = tmp_data_dic_total['commit']
            dep_commit_data['jira'] = tmp_data_dic_total['jira']
            dep_commit_data['RAT'] = tmp_data_dic_total['RAT']
            total_dep_commit_data.append(dep_commit_data)

            commit_jira_month.append(tmp_data_dic_total)
            commit_jira_data_name = k +'_commit_jira_data'
            self.generate_excel(commit_jira_month, 'report/'+commit_jira_data_name, 'commit_jira_data')

            #做折线图
            plt.figure('commit jira 月度图')
            plt.title(u''+k+' commit jira 月度图')
            plt.xlabel(u"月份")  # x轴名称
            plt.ylabel(u"数量")  # y 轴名称
            commit_jira_month_commit = [item['commit'] for item in commit_jira_month[:len(commit_jira_month)-1]]
            commit_jira_month_jira = [item['jira'] for item in commit_jira_month[:len(commit_jira_month) - 1]]
            commit_jira_month_rat = [item['RAT'] for item in commit_jira_month[:len(commit_jira_month) - 1]]
            commit_jira_month_month = [item['月份'][5:] for item in commit_jira_month[:len(commit_jira_month)-1]]
            plt.plot(commit_jira_month_month, commit_jira_month_commit,color='red',linewidth=2.0,linestyle='--')
            plt.plot(commit_jira_month_month, commit_jira_month_jira,color='blue',linewidth=3.0,linestyle='-.')
            plt.plot(commit_jira_month_month, commit_jira_month_rat, color='yellow', linewidth=3.0, linestyle='-.')
            plt.legend()  # 让图例生效
            #plt.xticks(x, names)  # 让x轴的刻度以names标签显示
            for i in range(len(commit_jira_month_commit)):
                plt.text(commit_jira_month_month[i], commit_jira_month_commit[i] + 0.5, '%s' % round(commit_jira_month_commit[i], 3), ha='center', fontsize=10)
            for i in range(len(commit_jira_month_jira)):
                plt.text(commit_jira_month_month[i], commit_jira_month_jira[i] + 3, '%s' % round(commit_jira_month_jira[i], 3), ha='center', fontsize=10, va='bottom')
            for i in range(len(commit_jira_month_rat)):
                plt.text(commit_jira_month_month[i], commit_jira_month_rat[i] - 3, '%s' % round(commit_jira_month_rat[i], 3), ha='center', fontsize=10, va='bottom')
            plt.savefig('report/'+ commit_jira_data_name + '.jpg')
            plt.close('commit jira 月度图')

        #整体部门的并列柱状图(commit, jira, RAT)
        self.generate_excel(total_dep_commit_data, 'report/commit_jira_各部门数据','total')
        plt.figure('commit jira 整体图')
        plt.title('各部门jira/RAT关联图')
        size = len(total_dep_commit_data)
        x = np.arange(size)
        commit_list = [item['commit'] for item in total_dep_commit_data]
        jira_list = [item['jira'] for item in total_dep_commit_data]
        rat_list = [item['RAT'] for item in total_dep_commit_data]
        rep_list = [item['部门'] for item in total_dep_commit_data]

        total_width, n = 0.8, 3
        width = total_width / n
        x = x - (total_width - width) / 2
        plt.xticks(x, labels=rep_list)

        t_commit_p = plt.bar(x, commit_list, width=width, label='commit')
        self.autolabel(t_commit_p)
        t_jira_p = plt.bar(x + width, jira_list, width=width, label='jira')
        self.autolabel(t_jira_p)
        t_rat_p = plt.bar(x + 2 * width, rat_list, width=width, label='RAT')
        self.autolabel(t_rat_p)
        plt.legend()
        plt.savefig('report/commit_jira_各部门图.jpg')
        plt.close('commit jira 整体图')

    def protect_unactive_branches(self):
        for k in self.department_info:
            if self.department_info[k]['need_be_frozen_branch_list']:
                for br in self.department_info[k]['need_be_frozen_branch_list']:
                    self.protect_branch(br['project_id'], br['branch_name'])

    def freeze_unactive_project(self):
        for k in self.department_info:
            if self.department_info[k]['unactive_repo']:
                self.freeze_project_list(self.department_info[k]['unactive_repo'])

    def protect_branch(self, id):
        gitlab_project = None
        failed = list()
        success = list()
        try:
            gitlab_project = self.gl.projects.get(id)
            # print('project is:' + str(gitlab_project))
        except:
            print('cannot find the gitlab_project with id:' + str(id))
            failed.append(id)
        if gitlab_project is not None:
            branch = gitlab_project.branches.get(branch_name)
            branch.protect()
            gitlab_project.save()
            print('id: ' + str(id) + 'branch: ' + str(branch_name) + ' was protected successfully!')
            success.append(id)
        if len(failed):
            raise Exception(print("The following ids failed:" + str(failed)))

    def freeze_project_list(self, id_list):
        failed = list()
        success = list()
        for id in id_list:
            gitlab_project = None
            try:
                gitlab_project = self.gl.projects.get(id)
                print('project is:' + str(gitlab_project))
            except:
                print('cannot find the gitlab_project with id:'+str(id))
                failed.append(id)
            if gitlab_project is not None:
                gitlab_project.archive()
                gitlab_project.save()
                print('id: '+ str(id) +' was frozen successfully!')
                success.append(id)
        if len(failed):
            raise Exception(print("The following ids failed:" + str(self.failed)))

def main():
    parser = argparse.ArgumentParser(
        formatter_class=argparse.RawDescriptionHelpFormatter,
        description=textwrap.dedent('''\
                     input include group repo and exclude group repo
                     '''))
    parser.add_argument('-r', '--include_repo',
                        metavar=str(), required=False,
                        help="input include repo",
                        type=str)
    parser.add_argument('-b', '--include_branch',
                        metavar=str(), required=False,
                        help="input include branch",
                        type=str)
    parser.add_argument('-ot', '--include_old_tag',
                        metavar=str(), required=False,
                        help="input include old tag",
                        type=str)
    parser.add_argument('-nt', '--include_new_tag',
                        metavar=str(), required=False,
                        help="input include new tag",
                        type=str)

    args = parser.parse_args()

    report = Report()
    #report.freeze_project_list([244])

    report.get_project_month_data()
    #report.freeze_unactive_project()
    #report.gen_branch_active_status(90)
    #report.get_commit_info_by_month([5888])
    #report.protect_branch(7178, 'feature-compile')
    #report.protect_unactive_branches()
    #report.generate_group_excel()
    #report.generate_project_excel()
    #report.generate_members_excel()
    #report.generate_branch_excel()
    #report.generate_count_excel()
    #if ('COMMIT_INFO' in os.environ.keys()) and (os.environ['COMMIT_INFO'] == 'true'):
    #    report.generate_commit_excel()
    #report.generate_single_commit_excel()
    #report.generate_mrs_excel()
    #report.generate_issues_excel()

if __name__ == "__main__":
    main()

