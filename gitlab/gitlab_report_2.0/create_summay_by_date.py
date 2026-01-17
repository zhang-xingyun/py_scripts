import get_raw_data
import pickle
import os
import re
import time
import sys
import shutil
import threading
import argparse
import textwrap
import openpyxl
import calendar
import datetime
import subprocess
import matplotlib.pyplot as plt
from pathlib import Path
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from xml.dom.minidom import parse
from lxml import etree
import gitlab


class Report():
    def __init__(self,
                 thread_num,
                 start_from,
                 ):
        # 开始获取数据的日期'2021-01'
        if start_from:
            self.date = start_from
        else:
            self.date = '2021-01'
        # 并发线程数量设置
        if thread_num:
            self.thread_num = int(thread_num)
        else:
            self.thread_num = 20
        self.month_list = self.gen_month_list(self.date)
        self.raw_data = get_raw_data.GetData('', False,False)
        self.br = self.raw_data.init_br()
        self.jira_issue_path = '/opt/gitlab_report/jira_data/issue/'
        self.people_data = self.read_pkl(self.raw_data.raw_data_path + 'people.pkl')['data']['records']
        self.current_time = int(time.time())
        self.raw_projects = self.read_pkl(self.raw_data.raw_data_path + 'projects_attributes.pkl')
        self.commit_job = dict()
        self.commit_job['branch'] = ''
        self.commit_job['project'] = dict()
        self.commit_job['commit'] = list()
        self.collect_info = [{'proj_id': 7552}, {'proj_id': 7429}, {'proj_id': 6790}, {'proj_id': 7814}, {'proj_id': 4488}, {'proj_id': 7564}, {'proj_id': 7565}, {'proj_id': 7438}, {'proj_id': 400}, {'proj_id': 7570}, {'proj_id': 3346}, {'proj_id': 6676}, {'proj_id': 1690}, {'proj_id': 7578}, {'proj_id': 7964}, {'proj_id': 6941}, {'proj_id': 5150}, {'proj_id': 7967}, {'proj_id': 7584}, {'proj_id': 7842}, {'proj_id': 6948}, {'proj_id': 8101}, {'proj_id': 7333}, {'proj_id': 8104}, {'proj_id': 7343}, {'proj_id': 303}, {'proj_id': 7987}, {'proj_id': 7347}, {'proj_id': 6964}, {'proj_id': 5043}, {'proj_id': 5687}, {'proj_id': 7099}, {'proj_id': 7484}, {'proj_id': 7102}, {'proj_id': 4671}, {'proj_id': 8128}, {'proj_id': 6907}, {'proj_id': 6908}, {'proj_id': 7877}, {'proj_id': 7623}, {'proj_id': 6984}, {'proj_id': 7370}, {'proj_id': 5963}, {'proj_id': 8140}, {'proj_id': 1996}, {'proj_id': 2384}, {'proj_id': 7637}, {'proj_id': 7639}, {'proj_id': 5722}, {'proj_id': 7518}, {'proj_id': 7647}, {'proj_id': 8159}, {'proj_id': 6625}, {'proj_id': 7650}, {'proj_id': 5984}, {'proj_id': 5982}, {'proj_id': 7805}, {'proj_id': 7785}, {'proj_id': 2155}, {'proj_id': 3437}, {'proj_id': 494}, {'proj_id': 3824}, {'proj_id': 2545}, {'proj_id': 240}, {'proj_id': 1909}, {'proj_id': 7670}, {'proj_id': 7031}, {'proj_id': 7672}, {'proj_id': 3195}, {'proj_id': 3452}, {'proj_id': 8189}, {'proj_id': 7679}, {'proj_id': 7552}, {'proj_id': 8067}, {'proj_id': 7814}, {'proj_id': 6790}, {'proj_id': 4488}, {'proj_id': 7438}, {'proj_id': 400}, {'proj_id': 6676}, {'proj_id': 1690}, {'proj_id': 7964}, {'proj_id': 6941}, {'proj_id': 5150}, {'proj_id': 7967}, {'proj_id': 7584}, {'proj_id': 6948}, {'proj_id': 7333}, {'proj_id': 7973}, {'proj_id': 7974}, {'proj_id': 7975}, {'proj_id': 7976}, {'proj_id': 8105}, {'proj_id': 7978}, {'proj_id': 7979}, {'proj_id': 7980}, {'proj_id': 7977}, {'proj_id': 7982}, {'proj_id': 7343}, {'proj_id': 303}, {'proj_id': 7347}, {'proj_id': 7988}, {'proj_id': 5043}, {'proj_id': 5687}, {'proj_id': 7877}, {'proj_id': 6984}, {'proj_id': 1996}, {'proj_id': 2384}, {'proj_id': 7637}, {'proj_id': 7639}, {'proj_id': 5722}, {'proj_id': 5982}, {'proj_id': 7647}, {'proj_id': 5984}, {'proj_id': 6625}, {'proj_id': 7650}, {'proj_id': 2155}, {'proj_id': 3437}, {'proj_id': 494}, {'proj_id': 3824}, {'proj_id': 2545}, {'proj_id': 8178}, {'proj_id': 240}, {'proj_id': 1909}, {'proj_id': 7031}, {'proj_id': 7672}, {'proj_id': 3195}, {'proj_id': 3452}, {'proj_id': 7679}, {'proj_id': 7939}, {'proj_id': 7944}, {'proj_id': 7945}, {'proj_id': 7692}, {'proj_id': 7949}, {'proj_id': 7455}, {'proj_id': 7083}, {'proj_id': 7863}, {'proj_id': 7864}, {'proj_id': 6975}, {'proj_id': 6977}, {'proj_id': 6978}, {'proj_id': 7242}, {'proj_id': 7380}, {'proj_id': 7128}, {'proj_id': 7385}, {'proj_id': 7784}, {'proj_id': 7786}, {'proj_id': 7659}, {'proj_id': 8172}, {'proj_id': 8173}, {'proj_id': 8174}, {'proj_id': 8175}, {'proj_id': 8176}, {'proj_id': 7791}, {'proj_id': 7277}, {'proj_id': 7677}, {'proj_id': 3073}, {'proj_id': 6147}, {'proj_id': 6148}, {'proj_id': 7178}, {'proj_id': 7693}, {'proj_id': 4631}, {'proj_id': 4641}, {'proj_id': 2603}, {'proj_id': 2096}, {'proj_id': 3666}, {'proj_id': 7255}, {'proj_id': 6239}, {'proj_id': 4202}, {'proj_id': 2167}, {'proj_id': 4729}, {'proj_id': 3713}, {'proj_id': 3715}, {'proj_id': 7305}, {'proj_id': 7306}, {'proj_id': 7307}, {'proj_id': 3732}, {'proj_id': 4253}, {'proj_id': 7334}, {'proj_id': 7338}, {'proj_id': 7356}, {'proj_id': 4285}, {'proj_id': 6860}, {'proj_id': 7373}, {'proj_id': 1742}, {'proj_id': 6864}, {'proj_id': 7891}, {'proj_id': 7892}, {'proj_id': 7893}, {'proj_id': 3285}, {'proj_id': 7392}, {'proj_id': 7394}, {'proj_id': 5870}, {'proj_id': 6897}, {'proj_id': 7410}, {'proj_id': 5878}, {'proj_id': 5879}, {'proj_id': 7416}, {'proj_id': 5881}, {'proj_id': 5880}, {'proj_id': 5882}, {'proj_id': 5883}, {'proj_id': 5884}, {'proj_id': 5885}, {'proj_id': 5886}, {'proj_id': 5888}, {'proj_id': 5887}, {'proj_id': 5890}, {'proj_id': 5891}, {'proj_id': 5892}, {'proj_id': 3848}, {'proj_id': 6928}, {'proj_id': 6929}, {'proj_id': 4394}, {'proj_id': 3376}, {'proj_id': 4913}, {'proj_id': 7993}, {'proj_id': 4922}, {'proj_id': 3390}, {'proj_id': 7495}, {'proj_id': 3919}, {'proj_id': 7520}, {'proj_id': 4453}, {'proj_id': 3431}, {'proj_id': 2409}, {'proj_id': 2412}, {'proj_id': 4468}, {'proj_id': 7546}, {'proj_id': 5505}, {'proj_id': 6018}, {'proj_id': 3972}, {'proj_id': 7561}, {'proj_id': 3468}, {'proj_id': 8097}, {'proj_id': 2473}, {'proj_id': 2474}, {'proj_id': 2475}, {'proj_id': 2476}, {'proj_id': 7601}, {'proj_id': 3506}, {'proj_id': 1457}, {'proj_id': 7621}, {'proj_id': 4046}, {'proj_id': 3030}, {'proj_id': 4056}, {'proj_id': 2529}, {'proj_id': 2530}, {'proj_id': 7651}, {'proj_id': 4580}, {'proj_id': 2551}, {'proj_id': 7676}]
        self.collect_info_excel = [{'proj_id': 8097}, {'proj_id': 7993}, {'proj_id': 7949}, {'proj_id': 7893}, {'proj_id': 7892}, {'proj_id': 7891}, {'proj_id': 7864}, {'proj_id': 7863}, {'proj_id': 7793}, {'proj_id': 7791}, {'proj_id': 7789}, {'proj_id': 7786}, {'proj_id': 7784}, {'proj_id': 7693}, {'proj_id': 7692}, {'proj_id': 7677}, {'proj_id': 7676}, {'proj_id': 7675}, {'proj_id': 7663}, {'proj_id': 7659}, {'proj_id': 7651}, {'proj_id': 7645}, {'proj_id': 7621}, {'proj_id': 7601}, {'proj_id': 7596}, {'proj_id': 7594}, {'proj_id': 7593}, {'proj_id': 7561}, {'proj_id': 7546}, {'proj_id': 7544}, {'proj_id': 7520}, {'proj_id': 7511}, {'proj_id': 7510}, {'proj_id': 7509}, {'proj_id': 7508}, {'proj_id': 7507}, {'proj_id': 7506}, {'proj_id': 7505}, {'proj_id': 7504}, {'proj_id': 7503}, {'proj_id': 7502}, {'proj_id': 7501}, {'proj_id': 7499}, {'proj_id': 7498}, {'proj_id': 7495}, {'proj_id': 7487}, {'proj_id': 7485}, {'proj_id': 7483}, {'proj_id': 7455}, {'proj_id': 7431}, {'proj_id': 7416}, {'proj_id': 7410}, {'proj_id': 7402}, {'proj_id': 7401}, {'proj_id': 7400}, {'proj_id': 7399}, {'proj_id': 7398}, {'proj_id': 7397}, {'proj_id': 7396}, {'proj_id': 7395}, {'proj_id': 7394}, {'proj_id': 7393}, {'proj_id': 7392}, {'proj_id': 7385}, {'proj_id': 7380}, {'proj_id': 7373}, {'proj_id': 7356}, {'proj_id': 7338}, {'proj_id': 7334}, {'proj_id': 7314}, {'proj_id': 7307}, {'proj_id': 7306}, {'proj_id': 7305}, {'proj_id': 7277}, {'proj_id': 7272}, {'proj_id': 7263}, {'proj_id': 7262}, {'proj_id': 7255}, {'proj_id': 7243}, {'proj_id': 7242}, {'proj_id': 7218}, {'proj_id': 7216}, {'proj_id': 7215}, {'proj_id': 7201}, {'proj_id': 7178}, {'proj_id': 7137}, {'proj_id': 7134}, {'proj_id': 7128}, {'proj_id': 7114}, {'proj_id': 7113}, {'proj_id': 7083}, {'proj_id': 7056}, {'proj_id': 7051}, {'proj_id': 7047}, {'proj_id': 7039}, {'proj_id': 7038}, {'proj_id': 7030}, {'proj_id': 7016}, {'proj_id': 7014}, {'proj_id': 7003}, {'proj_id': 6991}, {'proj_id': 6978}, {'proj_id': 6977}, {'proj_id': 6975}, {'proj_id': 6972}, {'proj_id': 6971}, {'proj_id': 6967}, {'proj_id': 6937}, {'proj_id': 6929}, {'proj_id': 6928}, {'proj_id': 6921}, {'proj_id': 6920}, {'proj_id': 6897}, {'proj_id': 6893}, {'proj_id': 6892}, {'proj_id': 6886}, {'proj_id': 6864}, {'proj_id': 6860}, {'proj_id': 6815}, {'proj_id': 6805}, {'proj_id': 6797}, {'proj_id': 6763}, {'proj_id': 6759}, {'proj_id': 6574}, {'proj_id': 6239}, {'proj_id': 6153}, {'proj_id': 6152}, {'proj_id': 6150}, {'proj_id': 6148}, {'proj_id': 6147}, {'proj_id': 6143}, {'proj_id': 6142}, {'proj_id': 6141}, {'proj_id': 6140}, {'proj_id': 6049}, {'proj_id': 6018}, {'proj_id': 8121}, {'proj_id': 7967}, {'proj_id': 7964}, {'proj_id': 7877}, {'proj_id': 7855}, {'proj_id': 7814}, {'proj_id': 7672}, {'proj_id': 7513}, {'proj_id': 7333}, {'proj_id': 6941}, {'proj_id': 5687}, {'proj_id': 4488}, {'proj_id': 6819}, {'proj_id': 6816}, {'proj_id': 6088}, {'proj_id': 6008}, {'proj_id': 6007}, {'proj_id': 5982}, {'proj_id': 5722}, {'proj_id': 5593}, {'proj_id': 5533}, {'proj_id': 3271}, {'proj_id': 2937}, {'proj_id': 2794}, {'proj_id': 1822}, {'proj_id': 1729}, {'proj_id': 1693}, {'proj_id': 1690}, {'proj_id': 1647}, {'proj_id': 1341}, {'proj_id': 1257}, {'proj_id': 1256}, {'proj_id': 1183}, {'proj_id': 873}, {'proj_id': 781}, {'proj_id': 757}, {'proj_id': 755}, {'proj_id': 640}, {'proj_id': 624}, {'proj_id': 605}, {'proj_id': 604}, {'proj_id': 534}, {'proj_id': 503}, {'proj_id': 502}, {'proj_id': 429}, {'proj_id': 388}, {'proj_id': 212}, {'proj_id': 186}, {'proj_id': 176}, {'proj_id': 7682}, {'proj_id': 7657}, {'proj_id': 7650}, {'proj_id': 7639}, {'proj_id': 7637}, {'proj_id': 7584}, {'proj_id': 7552}, {'proj_id': 7516}, {'proj_id': 7512}, {'proj_id': 7347}, {'proj_id': 7116}, {'proj_id': 7031}, {'proj_id': 6913}, {'proj_id': 6809}, {'proj_id': 6675}, {'proj_id': 6099}, {'proj_id': 6003}, {'proj_id': 5935}, {'proj_id': 5911}, {'proj_id': 5859}, {'proj_id': 5776}, {'proj_id': 5773}, {'proj_id': 5737}, {'proj_id': 5626}, {'proj_id': 5625}, {'proj_id': 5569}, {'proj_id': 5417}, {'proj_id': 5413}, {'proj_id': 5005}, {'proj_id': 4664}, {'proj_id': 3824}, {'proj_id': 3452}, {'proj_id': 3437}, {'proj_id': 3354}, {'proj_id': 3195}, {'proj_id': 2882}, {'proj_id': 1996}, {'proj_id': 1909}, {'proj_id': 868}, {'proj_id': 656}, {'proj_id': 494}, {'proj_id': 410}, {'proj_id': 7647}, {'proj_id': 7438}, {'proj_id': 7343}, {'proj_id': 7244}, {'proj_id': 7140}, {'proj_id': 6984}, {'proj_id': 6861}, {'proj_id': 6793}, {'proj_id': 6790}, {'proj_id': 6625}, {'proj_id': 5304}, {'proj_id': 3340}, {'proj_id': 2545}, {'proj_id': 2155}, {'proj_id': 5984}, {'proj_id': 7679}, {'proj_id': 7589}, {'proj_id': 6825}, {'proj_id': 6032}, {'proj_id': 5958}, {'proj_id': 5931}, {'proj_id': 5821}, {'proj_id': 5820}, {'proj_id': 5705}, {'proj_id': 5704}, {'proj_id': 4312}, {'proj_id': 4205}, {'proj_id': 4027}, {'proj_id': 4024}, {'proj_id': 3903}, {'proj_id': 3902}, {'proj_id': 3901}, {'proj_id': 3849}, {'proj_id': 2717}, {'proj_id': 2686}, {'proj_id': 2680}, {'proj_id': 2135}, {'proj_id': 1425}, {'proj_id': 1405}, {'proj_id': 1164}, {'proj_id': 1159}, {'proj_id': 1156}, {'proj_id': 1136}, {'proj_id': 779}, {'proj_id': 692}, {'proj_id': 631}, {'proj_id': 630}, {'proj_id': 602}, {'proj_id': 580}, {'proj_id': 517}, {'proj_id': 507}, {'proj_id': 484}, {'proj_id': 475}, {'proj_id': 400}, {'proj_id': 387}, {'proj_id': 358}, {'proj_id': 336}, {'proj_id': 303}, {'proj_id': 240}, {'proj_id': 123}, {'proj_id': 98}, {'proj_id': 5}, {'proj_id': 2}, {'proj_id': 7562}, {'proj_id': 7261}, {'proj_id': 7181}, {'proj_id': 7096}, {'proj_id': 7090}, {'proj_id': 7044}, {'proj_id': 6865}, {'proj_id': 6836}, {'proj_id': 6801}, {'proj_id': 6799}, {'proj_id': 6320}, {'proj_id': 6106}, {'proj_id': 6082}, {'proj_id': 5920}, {'proj_id': 5185}, {'proj_id': 4772}, {'proj_id': 4470}, {'proj_id': 4366}, {'proj_id': 4339}, {'proj_id': 4280}, {'proj_id': 4216}, {'proj_id': 3978}, {'proj_id': 3779}, {'proj_id': 3778}, {'proj_id': 3620}, {'proj_id': 3540}, {'proj_id': 3400}, {'proj_id': 1402}, {'proj_id': 1118}, {'proj_id': 647}, {'proj_id': 7422}, {'proj_id': 7415}, {'proj_id': 7359}, {'proj_id': 7309}, {'proj_id': 6948}, {'proj_id': 6676}, {'proj_id': 5153}, {'proj_id': 5150}, {'proj_id': 5043}, {'proj_id': 4622}, {'proj_id': 4621}, {'proj_id': 4620}, {'proj_id': 4619}, {'proj_id': 4618}, {'proj_id': 4616}, {'proj_id': 4615}, {'proj_id': 4613}, {'proj_id': 4612}, {'proj_id': 4611}, {'proj_id': 3135}, {'proj_id': 2723}, {'proj_id': 2384}]
        self.collect_info_excel_dep = []
        new_collect = self.plus_collect()
        print(str(new_collect))
        self.data_info_in_month = list()
        self.get_commit_info_from_date(new_collect, self.month_list)
        self.generate_excel(self.data_info_in_month, 'report/commit', 'between')
        self.save_pkl('report/year_data.pkl', self.data_info_in_month)

    def read_pkl(self, file_name):
        f_obj = open(file_name, 'rb')
        return pickle.load(f_obj)

    def multi_thread_run(self, thread_num):
        threadLock = threading.Lock()
        threads = list()
        for i in range(thread_num):
            tr = threading.Thread(target=self.run_thread)
            threads.append(tr)
        for i in threads:
            i.start()
        for i in threads:
            i.join()

    def run_thread(self):
        print("using thread: {}".format(threading.current_thread().name))
        while True:
            try:
                commit = self.commit_job['commit'].pop()
                branch = self.commit_job['branch']
                project = self.commit_job['project']
                data = self.filter_commit(project, self.month_list, commit, branch)
                # print("data=" + str(data))
                if data:
                    self.data_info_in_month.append(data)
            except Exception as e:
                print('there is error:' + str(e))
                if len(self.commit_job['commit']) == 0:
                    break
                continue


    def plus_collect(self):
        new_set = set()
        collect_info_set = set()
        collect_info_excel_dep_set = set()
        new_coll = list()
        filt_coll = list()
        for coll in self.collect_info:
            collect_info_set.add(coll['proj_id'])
            new_set.add(coll['proj_id'])
        for coll in self.collect_info_excel_dep:
            collect_info_excel_dep_set.add(coll['proj_id'])
            new_set.add(coll['proj_id'])
        for s in list(new_set):
            tmp = dict()
            tmp['proj_id'] = s
            new_coll.append(tmp)
        for s in list(collect_info_set):
            if s not in list(collect_info_excel_dep_set):
                tmp = dict()
                tmp['proj_id'] = s
                filt_coll.append(tmp)
        print(str(new_coll))
        return new_coll

    def save_pkl(self, file_name, data):
        # print(time.ctime(), 'to be write', file_name)
        f_obj = open(file_name, 'wb')
        pickle.dump(data, f_obj)
        f_obj.close()

    def get_xml_file(self, old_tag):
        branch_name = self.include_branch
        #repo_list = ['hobot-system', 'hobot-release']
        tmp_dir = 'tmp-' + self.include_repo
        if os.path.exists(tmp_dir):
            shutil.rmtree(tmp_dir)
        os.makedirs(tmp_dir)
        git_repo_cmd = 'git clone git@gitlab.test.com:ptd/ss/' + self.include_repo + '/manifest.git ' + tmp_dir
        print(git_repo_cmd)
        ret, output = self.runcmd(git_repo_cmd)
        if not ret:
            print(git_repo_cmd, ' error')
        current_folder = os.getcwd()
        os.chdir(tmp_dir)
        print(os.getcwd())
        git_checkout_cmd = 'git checkout -b %s origin/%s' % (
            branch_name, branch_name)

        ret, output = self.runcmd(git_checkout_cmd)
        if not ret:
            print(git_checkout_cmd, 'error')
            print(output)
            self.runcmd('git checkout ' + branch_name)
        xml_dir = Path(old_tag + '.xml')
        print('xml_dir is : ' + str(xml_dir))
        if xml_dir.exists():
            print('the xml is: ' + str(xml_dir))
            os.chdir(current_folder)
            return tmp_dir + '/' + str(xml_dir)
        return ''



    def parse_xml(self, xml_file):
        dom = parse(xml_file)
        data = dom.documentElement
        projects = data.getElementsByTagName('project')
        proj_list = list()
        for proj in projects:
            # 获取标签属性值
            proj_dic = dict()
            name = proj.getAttribute('name')
            path = proj.getAttribute('path')
            revision = proj.getAttribute('revision')
            upstream = proj.getAttribute('upstream')
            proj_dic['name'] = name
            proj_dic['path'] = path
            proj_dic['revision'] = revision
            proj_dic['upstream'] = upstream
            proj_dic['proj_id'] = self.get_query_projects(name)
            proj_list.append(proj_dic)
        return proj_list

    def get_parent_groups(self, group_id, parent_list):
        parent_id = self.raw_groups[group_id]['parent_id']
        if parent_id:
            parent_list.append(parent_id)
            self.get_parent_groups(parent_id, parent_list)

    def generate_excel(self,ids_data, filename, name):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = name
        #print(str(ids_data))
        for index,value in enumerate(ids_data[0].keys()):
            worksheet.cell(1, (index + 1), str(value))
        for row, one_data in enumerate(ids_data):
            #print(row)
            #print(one_data)
            for colum,value in enumerate(one_data.values()):
                #tmp_value = ILLEGAL_CHARACTERS_RE.sub(r'', str(value))
                tmp_value = value
                if isinstance(value,str):
                    tmp_value = ILLEGAL_CHARACTERS_RE.sub(r'', str(value))
                else:
                    tmp_value = value
                #print(type(tmp_value),tmp_value)
                worksheet.cell((row + 2), (colum + 1), tmp_value)
        file_path = '/'.join(filename.split('/')[:-1])

        print(file_path)
        if not os.path.exists(file_path):
            os.makedirs(file_path) 
        workbook.save((filename + '.xlsx'))
        print('finish')

    def runcmd(self, command):
        # print('run cmd---- ', command)
        try:
            ret = subprocess.run(
                command, shell=True,
                stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        except Exception as e:
            print(e)
            print('run cmd exception change to gpk ---- ', command)
            return (False, '')
        if ret.returncode != 0:
            # print('error', command, ret.stdout)
            print("error:", command, ret.returncode)
            # sys.exit(1)
            return (False, ''.join(map(chr, ret.stdout)))
        return (True, ''.join(map(chr, ret.stdout)))

    def find_jira_type(self, jira_id):
        #print('jira_id:' + jira_id)
        proj = jira_id.strip().split("-")[0]
        issue_pkl = self.jira_issue_path + proj + '.pkl'
        jira_issue = dict()
        if proj != '' and os.path.exists(issue_pkl):
            jira_issue = self.read_pkl(issue_pkl)
        for issue_id, issue_data in jira_issue.items():
            if issue_id == jira_id:
                if issue_id == 'JSP-1362':
                    print("issue_data:" + str(issue_data))
                jira_type = issue_data['fields']['issuetype']['name']
                fixVersions = issue_data['fields']['fixVersions']
                proj_name = issue_data['fields']['project']['name']
                jira_key = issue_data['fields']['project']['key']
                component_list = ''
                component_hash_list = list(issue_data['fields']['components'])
                if component_hash_list:
                    for component_hash in component_hash_list:
                        component_list = component_list + component_hash['name'] + '|'
                if fixVersions:
                    fixVersions = [fix['name'].strip() for fix in fixVersions][0]
                else:
                    fixVersions = ''
                #print(str(fixVersions))
                return jira_key, proj_name, jira_type, fixVersions, component_list
        return '','', '', '',''

    def gen_month_list(self, date):
        year, month = date.split('-')
        year = int(year)
        month = int(month)
        DateNow = datetime.datetime.today()
        DateNow = DateNow.strftime("%Y-%m")
        DateYear = datetime.datetime.today()
        DateYear = DateYear.strftime("%Y")
        # DateYear = 2021
        DateMonth = datetime.datetime.today()
        DateMonth = DateMonth.strftime("%m")
        # DateMonth = 12
        print(DateYear)
        print(DateMonth)
        print(DateNow)
        month_list = list()
        if year < int(DateYear):
            for m in range(month, 13):
                m_name = str(year) + '-' + str(m).zfill(2)
                month_list.append(m_name)
            for m in range(1, int(DateMonth) + 1):
                m_name = str(DateYear) + '-' + str(m).zfill(2)
                month_list.append(m_name)
            for y in range(year + 1, int(DateYear)):
                for m in range(1, 13):
                    m_name = str(y) + '-' + str(m).zfill(2)
                    month_list.append(m_name)
        if year == int(DateYear):
            for m in range(month, int(DateMonth) + 1):
                m_name = str(DateYear) + '-' + str(m).zfill(2)
                month_list.append(m_name)
        print(str(month_list))
        return month_list

    def get_commit_info_from_date(self, proj_list, month_list):
        for project in proj_list:
            output = dict()
            print("start to:" + str(project['proj_id']))
            #
            if not project:
                continue
            # if project['proj_id'] != 7892:
            #    continue
            output['proj'] = project['proj_id']
            commits_attributes_file_path = self.raw_data.raw_data_path + '/repo/' + str(project['proj_id']) + '/commits.pkl'
            if not os.path.isfile(commits_attributes_file_path):
                continue
            commits_attributes = self.read_pkl(commits_attributes_file_path)
            for branch, commit_data in commits_attributes.items():
                print(str(project['proj_id']) + " --> " + branch)
                self.commit_job['project'] = project
                self.commit_job['branch'] = branch
                self.commit_job['commit'] = commit_data
                self.multi_thread_run(20)

    def filter_commit(self, project, month_list, commit, branch):
        tmp_dict = dict()
        ret, author = self.parse_commit_author(commit['author'])
        if not ret or int(commit['additions']) > 5000 or int(
                                commit['deletions']) > 5000:
            return tmp_dict
        if 'submitter_data' not in commit.keys() or commit['3rd_skip']:
            return tmp_dict
        commit_month = commit['submitter_data'][0:7]
        # print(commit_month)
        if commit_month in self.month_list:
            tmp_dict['project_id'] = project['proj_id']
            tmp_dict['project_path'] = self.raw_projects[project['proj_id']]['web_url']
            tmp_dict['project_group'] = 'http://gitlab.test.com/' + \
                                        self.raw_projects[project['proj_id']]['namespace']['full_path']
            tmp_dict['project_namespace'] = self.raw_projects[project['proj_id']]['name_with_namespace']
            tmp_dict['branch_name'] = branch
            tmp_dict['commit'] = commit['commit']
            tmp_dict['author'] = author
            tmp_dict['date'] = commit['submitter_data']
            tmp_dict['month'] = commit['submitter_data'][0:7]
            tmp_dict['week'] = datetime.datetime.strptime(commit['submitter_data'],
                                                          '%Y-%m-%d').strftime('%Y-%W')
            tmp_dict['files'] = int(commit['files'])
            tmp_dict['additions'] = int(commit['additions'])
            tmp_dict['deletions'] = int(commit['deletions'])
            tmp_dict['total'] = tmp_dict['additions'] + tmp_dict['deletions']
            tmp_dict['cr_ph'] = ','.join(map(lambda x: 'D' + x, commit['cr_ph']))
            tmp_dict['jira_id'] = list()
            tmp_dict['jira_type'] = list()
            tmp_dict['businessUnit'], tmp_dict['division'], tmp_dict[
                'department'] = self.get_department(author)
            for key in commit['jira_data']:
                tmp_dict['jira_id'].append(key)
            tmp_dict['jira_id'] = ','.join(tmp_dict['jira_id'])
            tmp_dict['jira_type'] = ','.join(tmp_dict['jira_type'])
            tmp_dict['fixVersions'] = ''
            tmp_dict['jira_key'] = ''
            tmp_dict['proj_name'] = ''
            tmp_dict['component_list'] = ''
            if tmp_dict['jira_id']:
                tmp_dict['jira_key'], tmp_dict['proj_name'], tmp_dict['jira_type'], tmp_dict[
                    'fixVersions'], tmp_dict['component_list'] = self.find_jira_type(tmp_dict['jira_id'])
            if commit['postfix_list']:
                tmp_dict['code_type'] = self.get_code_type(commit['postfix_list'])
        return tmp_dict

    def get_department(self, user_id):
        for pe in self.people_data:
            if pe['hobotId'] == user_id:
                return pe['businessUnit'], pe['division'], pe['department']
        return '', '', ''

    def get_code_type(self, fix_list):
        code_type_set = set()
        #print(str(fix_list))
        for fix in fix_list:
            if fix == '.c':
                code_type_set.add('c')
            if fix == '.cc' or fix == '.cpp':
                code_type_set.add('c++')
            if fix == '.py':
                code_type_set.add('python')
            if fix == '.go':
                code_type_set.add('go')
            if fix == '.java':
                code_type_set.add('java')
        type_string = ''
        for type in list(code_type_set):
            type_string = type + '|'
        return type_string

    def parse_commit_author(self, author):
        tmp_author = author
        ret = True
        if tmp_author.find('horizon') < 0 and \
           tmp_author.find('hobot') < 0 and \
           tmp_author.find('hochip') < 0 and \
           tmp_author.find('hogpu') < 0 :
            ret = False
        if tmp_author.find('<') > 0:
            tmp_author = tmp_author.split('<')[1]
        if tmp_author.find('@') > 0:
            tmp_author = tmp_author.split('@')[0]
        return ret, tmp_author

def main():
    parser = argparse.ArgumentParser(
        formatter_class=argparse.RawDescriptionHelpFormatter,
        description=textwrap.dedent('''\
                     input thread number and start_from
                     '''))
    parser.add_argument('-r', '--thread_num',
                        metavar=str(), required=False,
                        help="input thread number like: 20",
                        type=str)
    parser.add_argument('-b', '--start_from',
                        metavar=str(), required=False,
                        help="input start from like 2021-01",
                        type=str)

    args = parser.parse_args()

    report = Report(args.thread_num, args.start_from)

if __name__ == "__main__":
    main()
