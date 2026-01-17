import gitlab
import pickle
import os
import time
import sys
import shutil
import subprocess
import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import threading
import gitlab

class GetData():
    def __init__(self, query_type, query_project_list, query_group_list, exclude_repo):
        self.gl = gitlab.Gitlab.from_config('trigger', 'python-gitlab.cfg')
        self.gl.auth()
        self.projects_attributes = dict()
        self.branches_attributes = dict()
        self.groups_attributes = dict()
        self.events_attributes = dict()
        self.mrs_attributes = dict()
        self.issues_attributes = dict()
        self.members_attributes = dict()
        self.query_type = query_type
        self.query_project_list = query_project_list
        self.query_group_list = query_group_list
        self.language = list()
        self.counts = dict()
        self.commits = dict()
        self.single_commits = dict()
        self.exclude_repo = exclude_repo
        self.current_time = int(time.time())
        self.git_cmd_thread = 120

    def get_all_groups(self, group_id):
        group = self.gl.groups.get(group_id)
        self.groups_attributes[group.attributes['id']] = group.attributes
        subgroups = group.subgroups.list(all=True, retry_transient_errors=True)
        for subgroup in subgroups:
            self.get_all_groups(subgroup.id)
    
    def get_all_groups_reverse(self, group_id):
        group = self.gl.groups.get(group_id)
        self.groups_attributes[group.attributes['id']] = group.attributes
        print(group.attributes['parent_id'])
        if group.attributes['parent_id']:
            self.get_all_groups_reverse(group.attributes['parent_id'])

    def get_groups(self):
        print(time.ctime(), 'get_groups')
        if not os.path.exists('data'):
            os.mkdir('data')
        for group_id in self.query_group_list:
            self.get_all_groups(group_id)
        print('group list:',self.groups_attributes.keys())
        for group_id in self.groups_attributes.keys():
            print('group:', group_id)
            group = self.gl.groups.get(group_id, retry_transient_errors=True)
            projects = group.projects.list(all=True, retry_transient_errors=True)
            for single_project in projects:
                print('project:', single_project.id)
                try:
                    project = self.gl.projects.get(single_project.id, retry_transient_errors=True)
                except Exception as e:
                    print(e)
                    continue
                print(project.attributes)
                self.projects_attributes[project.attributes['id']] = project.attributes
                # issue
                self.get_events(project)
                # mrs
                self.get_mrs(project)
                # issues
                self.get_issues(project)
                # members
                self.get_members(project)
        self.save_pkl('data/projects_attributes.pkl', self.projects_attributes)
        self.save_pkl('data/groups_attributes.pkl', self.groups_attributes)
        self.save_pkl('data/events_attributes.pkl', self.events_attributes)
        self.save_pkl('data/mrs_attributes.pkl', self.mrs_attributes)
        self.save_pkl('data/issues_attributes.pkl', self.issues_attributes)
        self.save_pkl('data/members_attributes.pkl',self.members_attributes)

    def get_projects(self):
        print('get_projects')
        if not os.path.exists('data'):
            os.mkdir('data')
        for project_id in self.query_project_list:
            print('project:', project_id)
            try:
                project = self.gl.projects.get(project_id, retry_transient_errors=True)
            except Exception as e:
                print(e)
                continue
            print(project.attributes)
            self.projects_attributes[project.attributes['id']] = project.attributes
            self.get_all_groups_reverse(project.attributes['namespace']['id'])
            print('project:', project.id)
            self.projects_attributes[project.attributes['id']] = project.attributes
            #issue
            self.get_events(project)
            # mrs
            self.get_mrs(project)
            # issues
            self.get_issues(project)
            # members
            self.get_members(project)
        self.save_pkl('data/projects_attributes.pkl', self.projects_attributes)
        self.save_pkl('data/groups_attributes.pkl', self.groups_attributes)
        self.save_pkl('data/events_attributes.pkl', self.events_attributes)
        self.save_pkl('data/mrs_attributes.pkl', self.mrs_attributes)
        self.save_pkl('data/issues_attributes.pkl', self.issues_attributes)
        self.save_pkl('data/members_attributes.pkl',self.members_attributes)

    def get_members(self, project):
        print('project member:', project.id)
        self.members_attributes[project.attributes['id']] = list()
        members = project.members.all(all=True, retry_transient_errors=True)
        for member in members:
            self.members_attributes[project.attributes['id']].append(member.attributes)

    def get_events(self, project):
        print('project event:', project.id)
        self.events_attributes[project.attributes['id']] = list()
        events = project.events.list(all=True, retry_transient_errors=True)
        for event in events:
            self.events_attributes[project.attributes['id']].append(event.attributes)

    def get_mrs(self, project):
        print('project mrs:', project.id)
        self.mrs_attributes[project.attributes['id']] = list()
        try:
            mrs = project.mergerequests.list(all=True, retry_transient_errors=True)
            for mr in mrs:
                tmp_mr = dict()
                tmp_mr['data'] = mr.attributes
                tmp_mr['discussion'] = list()
                tmp_mr['pipeline'] = list()
                tmp_mr['participant'] = list()
                for discussion in mr.discussions.list(all=True, retry_transient_errors=True):
                    if 'notes' not in discussion.attributes.keys():
                        continue
                    tmp_mr['discussion'].extend(discussion.attributes['notes'])
                #for commit in mr.commits():
                #    print(commit)
                for participant in mr.participants():
                    tmp_mr['participant'].append(participant)
                for pipeline in mr.pipelines():
                    tmp_mr['pipeline'].append(pipeline)
                self.mrs_attributes[project.attributes['id']].append(tmp_mr)
        except Exception as e:
            print(e)

    def get_issues(self, project):
        print('project issues:', project.id)
        self.issues_attributes[project.attributes['id']] = list()
        try:
            issues = project.issues.list(all=True, retry_transient_errors=True)
            for issue in issues:
                self.issues_attributes[project.attributes['id']].append(issue.attributes)
        except Exception as e:
            print(e)

    def get_branches(self):
        print(time.ctime(), 'get_branches')
        for project_id in self.projects_attributes.keys():
            print('get_branches', self.projects_attributes[project_id]['web_url'])
            if project_id in self.exclude_repo:
                print('exclude repo')
                continue
            try:
                project = self.gl.projects.get(project_id)
            except Exception as e:
                print(e)
                print('project exception')
            #print(project.attributes)
            # ignore 3rd
            group_full_path = project.attributes['namespace']['full_path']
            if group_full_path.find('/3rd') > 0 or \
               group_full_path.find('/thirdparty') > 0 or \
               group_full_path.find('/third_party') > 0 or \
               group_full_path.find('/third-party') > 0 or \
               group_full_path.find('/3rdparty') > 0 or \
               group_full_path.find('third_party') == 0 :
               #group_full_path.find('/experimental') > 0:
                print('3rd ignore')
                continue
            branches = project.branches.list(all=True, retry_transient_errors=True)
            if project_id not in self.branches_attributes.keys():
                self.branches_attributes[project_id] = list()
            for branch in branches:
                if branch.name == 'HEAD':
                    continue
                if project_id == 7403 and branch.name == 'harman_citation_64bit':
                    continue
                self.branches_attributes[project_id].append(branch.attributes)
        self.save_pkl('data/branches_attributes.pkl', self.branches_attributes)

    def runcmd(self, command):
        #print('run cmd---- ', command)
        try:
            ret = subprocess.run(command,shell=True,stdout=subprocess.PIPE,stderr=subprocess.PIPE)
        except Exception as e:
            print(e)
            print('run cmd exception change to gpk ---- ', command)
        if ret.returncode != 0:
            #print('error', command, ret.stdout)
            print("error:",command, ret.returncode)
            #sys.exit(1)
        return ''.join(map(chr, ret.stdout))

    def parse_loc_count(self, loc_info):
        print(time.ctime(), 'parse_loc_count')
        tmp_dict = dict()
        for line in loc_info.split('\n'):
            line_parse = line.split()
            if len(line_parse) != 6:
                continue
            if line_parse[0] == 'Language':
                continue
            #print(line_parse)
            if line_parse[0] not in self.language:
                self.language.append(line_parse[0])
            if line_parse[0] not in tmp_dict.keys():
                tmp_dict[line_parse[0]] = dict()
            tmp_dict[line_parse[0]]['Files'] = line_parse[1]
            tmp_dict[line_parse[0]]['Lines'] = line_parse[2]
            tmp_dict[line_parse[0]]['Blank'] = line_parse[3]
            tmp_dict[line_parse[0]]['Comment'] = line_parse[4]
            tmp_dict[line_parse[0]]['Code'] = line_parse[5]
        return tmp_dict

    def run_thread(self, commit_id_list, threadLock, commit_list, project_id):
        while True:
            try:
                commit_id = commit_id_list.pop()
                if commit_id in self.single_commits[project_id].keys():
                    #print('commit exist')
                    commit_list.append(self.single_commits[project_id][commit_id])
                    #print(len(commit_list))
                    continue
            except:
                if len(commit_id_list) == 0:
                    break
                continue
            #print(commit_id)
            commit_show_cmd = ' '.join(['git', '--no-pager', 'show','--shortstat', commit_id])
            #print(commit_show_cmd)
            commit_message = self.runcmd(commit_show_cmd).split('\n')
            #print(commit_message)
            if commit_message[1].find('Merge:') == 0:
                continue
            author = commit_message[1].replace('Author:','').strip()
            #if author.find('horizon') < 0 and \
            #   author.find('hogpu') < 0 and \
            #   author.find('hogpu') < 0 :
            #    continue
            date = commit_message[2].replace('Date:','').split('+')[0].strip()
            date = date.split('-')[0].strip()
            date = time.strftime("%Y-%m-%d", time.strptime(date))
            files = 0
            additions = 0
            deletions = 0
            for i in commit_message[-2].split(','):
                tmp_i = i.strip()
                #print(tmp_i)
                if tmp_i.find('changed') > 0:
                    files = tmp_i.split()[0]
                if tmp_i.find('insertion') > 0:
                    additions = tmp_i.split()[0]
                if tmp_i.find('deletion') > 0:
                    deletions = tmp_i.split()[0]
            #print(author,date,files,additions,deletions)
            tmp_dict = dict()
            tmp_dict['author'] = author
            tmp_dict['date'] = date
            tmp_dict['files'] = files
            tmp_dict['additions'] = additions
            tmp_dict['deletions'] = deletions
            tmp_dict['commit'] = commit_id
            commit_list.append(tmp_dict)
            self.single_commits[project_id][commit_id] = tmp_dict
            if len(commit_id_list) == 0:
                break

    def commit_parse_thread(self, commit_id_list, project_id):
        print(time.ctime(), 'commit_id_list', len(commit_id_list))
        #print(commit_id_list[0])
        #while len(commit_id_list) != 0:
        #    print(commit_id_list.pop())
        threadLock = threading.Lock()
        threads = list()
        data = dict()
        output = list()
        for i in range(self.git_cmd_thread):
            data[i] = list()
            tr = threading.Thread(target=self.run_thread, args=(commit_id_list, threadLock, data[i], project_id))
            threads.append(tr)
        for i in threads:
            i.start()
        for i in threads:
            i.join()
        for i,j in data.items():
            #print(i,len(j))
            output.extend(j)
        return output

    def get_commit_list(self, project_id):
        print(time.ctime(), 'get_commit_list')
        commit_id_list = self.runcmd('git --no-pager log --pretty=format:"%H"').split('\n')
        commit_list = self.commit_parse_thread(commit_id_list, project_id)
        return commit_list
 
    def get_branches_commits(self):
        print(time.ctime(), 'get_branches_commits')
        for project_id, branches in self.branches_attributes.items():
            #print(self.projects_attributes[project_id])
            ssh_url_to_repo = self.projects_attributes[project_id]['ssh_url_to_repo']
            if os.path.exists('tmp_repo'):
                shutil.rmtree('tmp_repo')
            print(ssh_url_to_repo)
            git_repo_cmd = ' '.join(['git', 'clone', ssh_url_to_repo, 'tmp_repo'])
            print(git_repo_cmd)
            self.runcmd(git_repo_cmd)
            os.chdir('tmp_repo')
            for branch in branches:
                #print(branch)
                print('branch name', branch['name'])
                # debug
                #if branch['name'] != 'll-fix-taskprogress':
                #    continue
                git_checkout_cmd = ' '.join(['git', 'checkout', '-b', branch['name'], 'origin/'+branch['name']])
                print(git_checkout_cmd)
                try:
                    self.runcmd(git_checkout_cmd)
                except Exception as e:
                    print(e)
                    self.runcmd('git checkout ' + branch['name'])
                count_info = self.runcmd('loc')
                if project_id not in self.counts.keys():
                    self.counts[project_id] = dict()
                self.counts[project_id][branch['name']] = self.parse_loc_count(count_info)
                if project_id not in self.commits.keys():
                    self.commits[project_id] = dict()
                if project_id not in self.single_commits.keys():
                    self.single_commits[project_id] = dict()
                self.commits[project_id][branch['name']] = self.get_commit_list(project_id)
                print(len(self.commits[project_id][branch['name']]))
                print(len(self.single_commits[project_id]))
            os.chdir('..')

    def save_pkl(self, file_name, data):
        print(time.ctime(), 'to be write')
        f_obj = open(file_name, 'wb')
        pickle.dump(data, f_obj)
        f_obj.close()

    def read_pkl(self, file_name):
        f_obj = open(file_name, 'rb')
        return pickle.load(f_obj)

    def generate_excel(self,ids_data, filename, name):
        print(filename)
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = name
    
        for index,value in enumerate(ids_data[0].keys()):
            worksheet.cell(1, (index + 1), str(value))
        for row, one_data in enumerate(ids_data):
            #print(row)
            #print(one_data)
            for colum,value in enumerate(one_data.values()):
                #tmp_value = ILLEGAL_CHARACTERS_RE.sub(r'', str(value))
                tmp_value = value
                if isinstance(value,str):
                    tmp_value = ILLEGAL_CHARACTERS_RE.sub(r'', str(value))
                else:
                    tmp_value = value
                #print(type(tmp_value),tmp_value)
                worksheet.cell((row + 2), (colum + 1), tmp_value)
        print('finish')
        workbook.save((filename + '.xlsx'))

    def generate_project_excel(self):
        if not os.path.exists('report'):
            os.mkdir('report')
        output_project = list()
        for project_id, project in self.projects_attributes.items():
            #print(project)
            project_created_at = project['created_at'].split('T')[0]
            project_last_activity_at = project['last_activity_at'].split('T')[0]
            project_group = 'http://gitlab.test.com/' + project['namespace']['full_path']
            project_path = project['web_url']
            project_namespace = project['name_with_namespace']
            tmp_project = dict()
            tmp_project['project_id'] = project_id
            tmp_project['project_namespace'] = project_namespace
            tmp_project['project_path'] = project_path
            tmp_project['project_group'] = project_group
            tmp_project['project_created_at'] = project_created_at
            tmp_project['project_last_activity_at'] = project_last_activity_at
            tmp_project['inactive_day'] = int((self.current_time - time.mktime(time.strptime(project_last_activity_at,'%Y-%m-%d')))/(24*60*60))
            tmp_project['type'] = 'normal'
            tmp_project['freeze'] = False
            if project['archived']:
                tmp_project['freeze'] = 'True'
            if project_id not in self.branches_attributes.keys():
                tmp_project['type'] = 'empty'
            if project_id in self.branches_attributes.keys() and \
               len(self.branches_attributes[project_id]) == 0:
                tmp_project['type'] = 'empty'
            group_full_path = project['namespace']['full_path']
            if group_full_path.find('/3rd') > 0 or \
                group_full_path.find('/thirdparty') > 0 or \
                group_full_path.find('/third_party') > 0 or \
                group_full_path.find('/third-party') > 0 or \
                group_full_path.find('/3rdparty') > 0 or \
                group_full_path.find('third_party') == 0 :
                tmp_project['type'] = '3rd'
            if group_full_path.find('/experimental') > 0:
                tmp_project['type'] = 'experimental'
            tmp_project['归属部门'] = None
            tmp_project['标记'] = None
            for one_tag in project['tag_list']:
                if one_tag.find('归属部门:') == 0:
                    tmp_project['归属部门'] = one_tag.split(':')[1]
                if one_tag.find('标记:') == 0:
                    tmp_project['标记'] = one_tag.split(':')[1]
            output_project.append(tmp_project)
        if len(output_project) > 0:
            self.generate_excel(output_project,'report/project','project')

    def generate_branch_excel(self):
        if not os.path.exists('report'):
            os.mkdir('report')
        output = list()
        for project_id, branch_list in self.branches_attributes.items():
            for branch in branch_list:
                #print(branch)
                freeze = False
                if branch['protected']:
                    if not branch['developers_can_push'] and \
                       not branch['developers_can_merge'] and \
                       not branch['can_push'] and \
                       not branch['default']:
                        freeze = True
                if self.projects_attributes[project_id]['archived']:
                    freeze = True
                branch_name = branch['name']
                branch_create = 'more_unknown'
                author_username = 'more_unknown'
                for event in self.events_attributes[project_id]:
                    if event['action_name'] == 'pushed new' and \
                       event['push_data']['ref'] == branch_name:
                        branch_create = event['created_at'].split('T')[0]
                        author_username = event['author_username']
                tmp_dict = dict()
                tmp_dict['project_id'] = project_id
                tmp_dict['project_namespace'] = self.projects_attributes[project_id]['name_with_namespace']
                tmp_dict['project_path'] = self.projects_attributes[project_id]['web_url']
                tmp_dict['project_group'] = 'http://gitlab.test.com/' + self.projects_attributes[project_id]['namespace']['full_path']
                tmp_dict['branch_name'] = branch_name
                tmp_dict['freeze'] = freeze
                tmp_dict['author_username'] = author_username
                tmp_dict['branch_create'] = branch_create
                tmp_dict['branch_last_commit'] = str()
                for i in self.commits[project_id][branch['name']]:
                    if i['date'] > tmp_dict['branch_last_commit']:
                        tmp_dict['branch_last_commit'] = i['date']
                tmp_dict['inactive_day'] = int((self.current_time - time.mktime(time.strptime(tmp_dict['branch_last_commit'],'%Y-%m-%d')))/(24*60*60))
                tmp_dict['归属部门'] = None
                tmp_dict['标记'] = None
                for one_tag in self.projects_attributes[project_id]['tag_list']:
                    if one_tag.find('归属部门:') == 0:
                        tmp_dict['归属部门'] = one_tag.split(':')[1]
                    if one_tag.find('标记:') == 0:
                        tmp_dict['标记'] = one_tag.split(':')[1]
                #print(tmp_dict['branch_last_commit']) 
                #sys.exit(1)
                #tmp_dict['branch_last_commit'] = self.commits[project_id][branch['name']][0]['date']
                output.append(tmp_dict)
        if len(output) > 0:
            self.generate_excel(output,'report/branch','branch')

    def generate_count_excel(self):
        if not os.path.exists('report'):
            os.mkdir('report')
        output = list()
        for project_id, branch_list in self.counts.items():
            for branch, count_data in branch_list.items():
                for count_type, count in count_data.items():
                    tmp_dict = dict()
                    tmp_dict['project_id'] = project_id
                    tmp_dict['project_namespace'] = self.projects_attributes[project_id]['name_with_namespace']
                    tmp_dict['project_path'] = self.projects_attributes[project_id]['web_url']
                    tmp_dict['project_group'] = 'http://gitlab.test.com/' + self.projects_attributes[project_id]['namespace']['full_path']
                    tmp_dict['branch_name'] = branch
                    tmp_dict['count_type'] = count_type
                    tmp_dict['Files'] = int(count['Files'])
                    tmp_dict['Lines'] = int(count['Lines'])
                    tmp_dict['Blank'] = int(count['Blank'])
                    tmp_dict['Comment'] = int(count['Comment'])
                    tmp_dict['Code'] = int(count['Code'])
                    output.append(tmp_dict)
        if len(output) > 0:
            self.generate_excel(output,'report/count','count')

    def generate_commit_excel(self):
        if not os.path.exists('report/commit'):
            os.mkdir('report/commit')
        for project_id, branch_data in self.commits.items():
            output = dict()
            for branch, commit_data in branch_data.items():
                for commit in commit_data:
                    ret, author = self.parse_commit_author(commit['author'])
                    if not ret:
                        continue
                    tmp_dict = dict()
                    tmp_dict['project_id'] = project_id
                    tmp_dict['project_path'] = self.projects_attributes[project_id]['web_url']
                    tmp_dict['project_group'] = 'http://gitlab.test.com/' + self.projects_attributes[project_id]['namespace']['full_path']
                    tmp_dict['project_namespace'] = self.projects_attributes[project_id]['name_with_namespace']
                    tmp_dict['branch_name'] = branch
                    tmp_dict['commit'] = commit['commit']
                    tmp_dict['author'] = author
                    tmp_dict['date'] = commit['date']
                    tmp_dict['files'] = int(commit['files'])
                    tmp_dict['additions'] = int(commit['additions'])
                    tmp_dict['deletions'] = int(commit['deletions'])
                    tmp_dict['total'] = tmp_dict['additions'] + tmp_dict['deletions']
                    year = tmp_dict['date'].split('-')[0]
                    if year not in output.keys():
                        output[year] = list()
                    output[year].append(tmp_dict)
            for year, data in output.items():
                if len(data) > 0:
                    self.generate_excel(data,'report/commit/' + str(project_id) + '-' + year, 'commit')

    def generate_single_commit_excel(self):
        if not os.path.exists('report'):
            os.mkdir('report')
        output = dict()
        for project_id, commit_data in self.single_commits.items():
            for single, commit in commit_data.items():
                ret, author = self.parse_commit_author(commit['author'])
                if not ret:
                    continue
                tmp_dict = dict()
                tmp_dict['project_id'] = project_id
                tmp_dict['project_path'] = self.projects_attributes[project_id]['web_url']
                tmp_dict['project_group'] = 'http://gitlab.test.com/' + self.projects_attributes[project_id]['namespace']['full_path']
                tmp_dict['project_namespace'] = self.projects_attributes[project_id]['name_with_namespace']
                tmp_dict['commit'] = commit['commit']
                tmp_dict['author'] = author
                tmp_dict['date'] = commit['date']
                tmp_dict['files'] = int(commit['files'])
                tmp_dict['additions'] = int(commit['additions'])
                tmp_dict['deletions'] = int(commit['deletions'])
                tmp_dict['total'] = tmp_dict['additions'] + tmp_dict['deletions']
                year = tmp_dict['date'].split('-')[0]
                if year not in output.keys():
                    output[year] = list()
                output[year].append(tmp_dict)
        for year, data in output.items():
            if len(data) > 0:
                self.generate_excel(data,'report/single_commit_' + year,'single_commit')

    def parse_commit_author(self, author):
        tmp_author = author
        ret = True
        if tmp_author.find('horizon') < 0 and \
           tmp_author.find('hobot') < 0 and \
           tmp_author.find('hochip') < 0 and \
           tmp_author.find('hogpu') < 0 :
            ret = False
        if tmp_author.find('<') > 0:
            tmp_author = tmp_author.split('<')[1]
        if tmp_author.find('@') > 0:
            tmp_author = tmp_author.split('@')[0]
        return ret, tmp_author
        
    def generate_mrs_excel(self):
        output = list()
        output_pipeline = list()
        output_discussion = list()
        output_participant = list()
        for project_id, mr_list in self.mrs_attributes.items():
            for mr in mr_list:
                #print(mr['data'])
                tmp_dict = dict()
                tmp_dict['web_url'] = mr['data']['web_url']
                tmp_dict['title'] = mr['data']['title']
                tmp_dict['state'] = mr['data']['state']
                tmp_dict['created_at'] = mr['data']['created_at'].split('T')[0]
                tmp_dict['closed_or_merged'] = None
                tmp_dict['updated_at'] = mr['data']['updated_at'].split('T')[0]
                if mr['data']['merged_by'] is None:
                    tmp_dict['merged_by'] = None
                else:
                    tmp_dict['merged_by'] = mr['data']['merged_by']['username']
                tmp_dict['target_branch'] = mr['data']['target_branch']
                tmp_dict['source_branch'] = mr['data']['source_branch']
                tmp_dict['author'] = mr['data']['author']['username']
                tmp_dict['sha'] = mr['data']['sha']
                tmp_dict['project_path'] = self.projects_attributes[project_id]['web_url']
                #output.append(tmp_dict)
                for pipeline in mr['pipeline']:
                    tmp_pipeline = {'mr_url': mr['data']['web_url']}
                    tmp_pipeline.update(pipeline)
                    output_pipeline.append(tmp_pipeline)
                #
                tmp_output_participant_list = list()
                for participant in mr['participant']:
                    tmp_output_participant_list.append(participant['username'])
                tmp_output_participant_list = list(set(tmp_output_participant_list))
                for i in tmp_output_participant_list:
                    if i == 'robot':
                        continue
                    tmp_participant = dict()
                    tmp_participant['mr_url'] = mr['data']['web_url']
                    tmp_participant['people'] = i
                    tmp_participant['project_path'] = self.projects_attributes[project_id]['web_url']
                    output_participant.append(tmp_participant)
                #
                for discussion in mr['discussion']:
                    #print('------------------------------')
                    #print(discussion)
                    if discussion['system']:
                        if discussion['body'] == 'merged' or discussion['body'] == 'closed':
                            tmp_dict['closed_or_merged'] = discussion['updated_at'].split('T')[0]
                        continue
                    if discussion['author']['username'] == 'robot':
                        continue
                    if discussion['body'] == 'Jenkins please retry a build':
                        continue
                    tmp_discussion = dict()
                    tmp_discussion['mr_url'] = mr['data']['web_url']
                    tmp_discussion['body'] = discussion['body']
                    tmp_discussion['author'] = discussion['author']['username']
                    tmp_discussion['created_at'] = discussion['created_at'].split('T')[0]
                    tmp_discussion['updated_at'] = discussion['updated_at'].split('T')[0]
                    tmp_discussion['project_path'] = self.projects_attributes[project_id]['web_url']
                    #tmp_pipeline['system'] = discussion['system']
                    output_discussion.append(tmp_discussion)
                output.append(tmp_dict)
                    
        if len(output) > 0:
            self.generate_excel(output,'report/mr','mr')
        if len(output_pipeline) > 0:
            self.generate_excel(output_pipeline,'report/mr_pipeline','mr_pipeline')
        if len(output_discussion) > 0:
            self.generate_excel(output_discussion,'report/mr_discussion','mr_discussion')
        if len(output_participant) > 0:
            self.generate_excel(output_participant,'report/mr_participant','mr_participant')

    def generate_issues_excel(self):
        output = list()
        for project_id, issues_list in self.issues_attributes.items():
            for issue in issues_list:
                tmp_dict = dict()
                tmp_dict['project_path'] = self.projects_attributes[project_id]['web_url']
                tmp_dict['title'] = issue['title']
                tmp_dict['state'] = issue['state']
                tmp_dict['created_at'] = issue['created_at'].split('T')[0]
                tmp_dict['updated_at'] = issue['updated_at'].split('T')[0]
                if issue['closed_at'] is None:
                    tmp_dict['closed_at'] = None
                else:
                    tmp_dict['closed_at'] = issue['closed_at'].split('T')[0]
                if issue['closed_by'] is None:
                    tmp_dict['closed_by'] = None
                else:
                    tmp_dict['closed_by'] = issue['closed_by']['username']
                tmp_dict['author'] = issue['author']['username']
                tmp_dict['web_url'] = issue['web_url']
                output.append(tmp_dict)
        if len(output) > 0:
            self.generate_excel(output,'report/issue','issue')

    def generate_members_excel(self):
        output = list()
        for project_id, members_list in self.members_attributes.items():
            for member in members_list:
                #print(member)
                #continue
                if member['state'] != 'active':
                    continue
                tmp_dict = dict()
                tmp_dict['project_path'] = self.projects_attributes[project_id]['web_url']
                tmp_dict['username'] = member['username']
                tmp_dict['access_level'] = member['access_level']
                output.append(tmp_dict)
        if len(output) > 0:
            self.generate_excel(output,'report/member','member')
            
def main(query_type, query_project_list, query_group_list, exclude_repo):
    get_data = GetData(query_type, query_project_list, query_group_list, exclude_repo)
    if query_type == 'group':
        get_data.get_groups()
    else:
        get_data.get_projects()
    print('total project number is %d' % len(get_data.projects_attributes))
    get_data.get_branches()
    get_data.get_branches_commits()
    get_data.save_pkl('data/language.pkl', get_data.language)
    get_data.save_pkl('data/counts.pkl', get_data.counts)
    get_data.save_pkl('data/commits.pkl', get_data.commits)
    get_data.save_pkl('data/single_commits.pkl', get_data.single_commits)
    #excel
    get_data.generate_project_excel()
    get_data.generate_branch_excel()
    get_data.generate_count_excel()
    get_data.generate_commit_excel()
    get_data.generate_mrs_excel()
    get_data.generate_issues_excel()
    get_data.generate_members_excel()
    get_data.generate_single_commit_excel()


if __name__ == "__main__":
    query_type = sys.argv[1]
    query_list = sys.argv[2]
    try:
        exclude_repo = sys.argv[3]
    except:
        exclude_repo = list()
    print(query_type, query_list, exclude_repo)
    if query_type != 'repo' and query_type != 'group':
        print('query type only support repo and group')
    try:
        query_list = list(map(lambda x:int(x.strip()), query_list.split(',')))
    except Exception as e:
        print(e)
        print('only support number with comma split, example 1,2,3,4')
    try:
        exclude_repo = list(map(lambda x:int(x.strip()), exclude_repo.split(',')))
    except Exception as e:
        print(e)
        print('exclude_repo only support number with comma split, example 1,2,3,4')
        exclude_repo = list()
    if os.path.exists('data'):
        shutil.rmtree('data')
    if os.path.exists('report'):
        shutil.rmtree('report')
    if query_type == 'repo':
        main(query_type, query_list, [], exclude_repo)
    else:
        main(query_type, [], query_list, exclude_repo)
